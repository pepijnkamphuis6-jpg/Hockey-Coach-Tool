import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import time
import uuid
import json
import os
import re
import hashlib
import hmac
from io import BytesIO
from textwrap import dedent
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

try:
    from supabase import create_client
except Exception:
    create_client = None

try:
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors as _rl_colors
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

st.set_page_config(
    page_title="Coach Studio",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# ==================================================
# AUTHENTICATIE & TEAMS (multi-team met wachtwoord per team)
# ==================================================
def _hash_password(password: str, salt: str = "coach-studio-v1") -> str:
    """Simpele gezouten SHA-256 hash voor teamwachtwoorden."""
    raw = (salt + ":" + (password or "")).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def _verify_password(password: str, stored_hash: str) -> bool:
    if not stored_hash:
        return False
    return hmac.compare_digest(_hash_password(password), stored_hash)


def _teams_local_file() -> str:
    """Pad naar de lokale fallback-file (naast app.py).

    Werkt zowel lokaal, in Cowork als op Streamlit Cloud — het file staat
    dan altijd naast de script-file, waar we schrijf-rechten hebben.
    """
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        base_dir = os.getcwd()
    return os.path.join(base_dir, "_teams_local.json")


def _load_local_teams() -> list:
    try:
        with open(_teams_local_file(), "r", encoding="utf-8") as f:
            return json.load(f) or []
    except Exception:
        return []


def _save_local_teams(teams: list) -> None:
    try:
        with open(_teams_local_file(), "w", encoding="utf-8") as f:
            json.dump(teams, f, ensure_ascii=False, indent=2)
    except Exception as err:
        log_cloud_error("teams lokaal opslaan", err)


@st.cache_data(ttl=120, show_spinner=False)
def list_teams() -> list:
    """Haal alle teams op (uit Supabase, fallback lokaal)."""
    client = get_supabase_client()
    if client is not None:
        try:
            response = client.table("teams").select("id,name,password_hash,created_at").order("name").execute()
            rows = response.data or []
            mark_cloud_ok()
            return [
                {"id": r.get("id"), "name": r.get("name", ""), "password_hash": r.get("password_hash", "")}
                for r in rows if r.get("name")
            ]
        except Exception as err:
            log_cloud_error("teams ophalen", err)
    return _load_local_teams()


def create_team(name: str, password: str) -> tuple[bool, str]:
    """Maak een nieuw team aan. Retourneert (success, message)."""
    name = (name or "").strip()
    password = (password or "").strip()
    if not name:
        return False, "Geef een teamnaam op."
    if len(password) < 4:
        return False, "Kies een wachtwoord van minstens 4 tekens."
    existing = list_teams()
    if any(t.get("name", "").lower() == name.lower() for t in existing):
        return False, f"Team '{name}' bestaat al."
    team_id = str(uuid.uuid4())
    password_hash = _hash_password(password)
    row = {"id": team_id, "name": name, "password_hash": password_hash}
    client = get_supabase_client()
    if client is None:
        # Supabase kan niet worden gestart — vertel waarom
        import traceback
        try:
            # probeer expliciet wat er misgaat bij het starten
            from supabase import create_client as _cc
            _cc(st.secrets.get("SUPABASE_URL", ""), st.secrets.get("SUPABASE_KEY", ""))
            reason = "create_client geeft None terug"
        except Exception as e:
            reason = f"{type(e).__name__}: {e}"
        teams = _load_local_teams()
        teams.append(row)
        _save_local_teams(teams)
        return False, f"CLOUD KAN NIET GESTART WORDEN — {reason}"
    try:
        client.table("teams").insert({**row, "created_at": time.time()}).execute()
        mark_cloud_ok()
        return True, f"Team '{name}' aangemaakt."
    except Exception as err:
        log_cloud_error("team aanmaken", err)
        teams = _load_local_teams()
        teams.append(row)
        _save_local_teams(teams)
        return False, f"CLOUD FOUT — {type(err).__name__}: {err}"


def delete_team(team_id: str) -> None:
    client = get_supabase_client()
    if client is not None:
        try:
            client.table("teams").delete().eq("id", team_id).execute()
            mark_cloud_ok()
        except Exception as err:
            log_cloud_error("team verwijderen", err)
    teams = [t for t in _load_local_teams() if t.get("id") != team_id]
    _save_local_teams(teams)


def require_password() -> None:
    # Session state defaults voor auth
    st.session_state.setdefault("authenticated", False)
    st.session_state.setdefault("user_role", None)
    st.session_state.setdefault("active_team_id", None)
    st.session_state.setdefault("active_team_name", None)
    st.session_state.setdefault("active_tool", None)
    st.session_state.setdefault("login_mode", "login")  # 'login' | 'new_team'

    if st.session_state.authenticated and st.session_state.active_team_id:
        return

    st.markdown(
        '<div class="login-shell">'
        '<div class="login-logo">CS</div>'
        '<div class="login-title">Coach Studio</div>'
        '<div class="login-sub">Kies je team om verder te gaan</div>'
        '</div>',
        unsafe_allow_html=True,
    )

    teams = list_teams()

    _, mid, _ = st.columns([1, 2, 1])
    with mid:
        # Diagnose-panel: laat direct zien of de cloud werkt en hoeveel teams er
        # gevonden zijn. Klaps zichzelf open als er iets mis is.
        render_login_diagnostics()

        tabs = st.tabs(["Inloggen bij team", "Nieuw team aanmaken"])

        # ------ TAB 1: Inloggen ------
        with tabs[0]:
            if not teams:
                st.info("Nog geen teams aangemaakt. Ga naar de tab hiernaast om je eerste team aan te maken.")
            else:
                team_names = [t["name"] for t in teams]
                picked_name = st.selectbox(
                    "Team",
                    team_names,
                    key="login_team_pick",
                )
                password = st.text_input(
                    "Team-wachtwoord",
                    type="password",
                    key="login_team_pw",
                    placeholder="Wachtwoord",
                )
                login_clicked = st.button(
                    "Inloggen",
                    use_container_width=True,
                    type="primary",
                    key="login_submit_btn",
                )
                if login_clicked:
                    picked = next((t for t in teams if t["name"] == picked_name), None)
                    if picked is None:
                        st.error("Team niet gevonden.")
                    elif _verify_password(password, picked.get("password_hash", "")):
                        st.session_state.authenticated = True
                        st.session_state.user_role = "coach"
                        st.session_state.active_team_id = picked["id"]
                        st.session_state.active_team_name = picked["name"]
                        st.session_state.active_tool = None
                        # Reset wedstrijd-data zodat we niet data van vorig team tonen
                        st.session_state.events = []
                        st.session_state.video_clips = []
                        # Reset wisselschema-state (als die bestond voor ander team)
                        for k in ("subs_players", "subs_match", "subs_attendance", "subs_schema"):
                            if k in st.session_state:
                                del st.session_state[k]
                        # Nieuw match-id dat aan dit team hangt
                        prefix = f"T-{picked['id'][:8]}__"
                        st.session_state.match_id = f"{prefix}wedstrijd-{uuid.uuid4().hex[:6]}"
                        st.session_state.team_name = picked["name"]
                        st.rerun()
                    else:
                        st.error("Onjuist wachtwoord voor dit team.")

        # ------ TAB 2: Nieuw team ------
        with tabs[1]:
            new_name = st.text_input("Naam van je team", key="new_team_name",
                                     placeholder="Bijv. MO16-1 Hockey Club Xerxes")
            new_pw = st.text_input("Kies een wachtwoord", type="password",
                                   key="new_team_pw", placeholder="Minstens 4 tekens")
            new_pw_confirm = st.text_input("Wachtwoord bevestigen", type="password",
                                           key="new_team_pw2", placeholder="Nogmaals")
            create_clicked = st.button(
                "Team aanmaken",
                use_container_width=True,
                type="primary",
                key="create_team_btn",
            )
            if create_clicked:
                if new_pw != new_pw_confirm:
                    st.error("Wachtwoorden komen niet overeen.")
                else:
                    ok, msg = create_team(new_name, new_pw)
                    if ok:
                        st.success(msg + " Je kan nu inloggen op de andere tab.")
                    else:
                        st.error(msg)

    st.stop()


def render_logout_button() -> None:
    active_tool = st.session_state.get("active_tool")
    _, btn_col1, btn_col2, _ = st.columns([6, 1, 1, 0.1])
    with btn_col1:
        label = "🏠  Home" if active_tool else "🏠  Home"
        if st.button(label, use_container_width=True, key="back_to_tools_btn",
                     help="Terug naar het tool-overzicht"):
            st.session_state.active_tool = None
            st.rerun()
    with btn_col2:
        if st.button("↩  Uitloggen", use_container_width=True, key="logout_btn",
                     help="Uitloggen en terug naar teamkeuze"):
            st.session_state.authenticated = False
            st.session_state.user_role = None
            st.session_state.active_team_id = None
            st.session_state.active_team_name = None
            st.session_state.active_tool = None
            st.rerun()


def has_edit_rights() -> bool:
    # Zowel coach als assistent mogen taggen en analyseren
    return st.session_state.get("user_role") in ["coach", "assistent"]


def is_viewer() -> bool:
    # Géén rollen meer met "kijkmodus" — behouden voor compatibiliteit, retourneert altijd False
    return False


def is_coach() -> bool:
    """Volledige rechten — mag wedstrijd resetten en gevoelige acties uitvoeren."""
    return st.session_state.get("user_role") == "coach"


# (require_password() wordt onderaan in MAIN aangeroepen, zodat alle
# functies — inclusief inject_custom_css — eerst gedefinieerd zijn.)


# ==================================================
# DEFAULTS
# ==================================================
DEFAULTS = {
    "events": [],
    "video_clips": [],
    "timer_running": False,
    "start_time": None,
    "elapsed_before_run": 0,
    "quarter": "Q1",
    "team_name": "Ons team",
    "opponent_name": "Tegenstander",
    "score_team": 0,
    "score_opponent": 0,
    "match_id": "wedstrijd-1",
    "last_sync_time": None,
    "last_sync_count": 0,
    "auto_notes": "",
    "active_screen": "LIVE",
    "pending_event": None,
    "pending_team": None,
    "field_team": None,
    "field_quarter": "Alles",
    "field_layers": ["Cirkelentry", "Schot", "Goal"],
    "device_mode": "iPad",
    "halftime_report": "",
    "confirm_reset": False,
    "ui_team_name": "Ons team",
    "ui_opponent_name": "Tegenstander",
    "ui_quarter": "Q1",
    "ui_match_id": "wedstrijd-1",
    "ui_device_mode": "iPad",
    "uploaded_video_name": "",
    # NIEUW voor videospeler
    "video_source_type": "upload",        # 'upload', 'url', 'youtube'
    "video_url": "",
    "video_bytes_b64": "",                # lokaal geuploade video als data-URL
    "pushoff_offsets": {"Q1": None, "Q2": None, "Q3": None, "Q4": None},
    "current_video_time": 0.0,            # door speler teruggegeven tijd
    "jump_to_video_time": None,           # wanneer gezet: speler springt erheen
    "video_tag_pending": None,            # tijdelijke tag tijdens taggen-via-video
    # Cloud-fout tracker — laat zien wanneer er iets mis ging met opslaan
    "cloud_errors": [],                   # lijst met recente cloud-fouten
    "last_cloud_ok": True,                # laatste cloud-actie geslaagd?
    # Undo-historie: laatste N verwijderde events, zodat je ze kan terughalen
    "undo_stack": [],                     # lijst met zojuist verwijderde events
    "last_undo_msg": "",                  # feedback na undo
    # Wedstrijden-manager (oude in-page picker — blijft bestaan voor backwards compat)
    "show_match_picker": False,
    "show_new_match_form": False,
    # Nieuwe wedstrijd-picker op het tool-overzicht (home)
    "home_show_match_picker": False,
    "home_show_new_match": False,
    # Highlight-reel (video analyse uitbreiding)
    "reel_active": False,                 # staat de reel-speler aan?
    "reel_fragments": [],                 # lijst met {start_sec, label}
    "reel_share_links": [],               # lijst met deelbare YouTube-links
    # Clip-delen
    "clip_share_links": {},               # {clip_id: url}
    # Tekenen op pauze
    "drawn_snapshots": [],                # [{id, clip_id, image_b64, created_at, note}]
}
for key, value in DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = value


# ==================================================
# CONSTANTEN
# ==================================================
QUARTERS = ["Q1", "Q2", "Q3", "Q4"]
FIELD_ZONES = ["Linksvoor", "Middenvoor", "Rechtsvoor"]
EVENT_NEEDS_ZONE = {"Cirkelentry"}
VIDEO_TAGS = [
    "Opbouw",
    "Press",
    "Restverdediging",
    "Omschakeling aanval",
    "Omschakeling verdediging",
    "Cirkelentry",
    "Cirkelbezetting",
    "Verdedigende organisatie",
    "Strafcorner",
    "Goal voor",
    "Goal tegen",
    "Positief voorbeeld",
    "Leerclip",
]

# Premium kleurschema — diep navy, sportief, professioneel
ACCENT = "#3b82f6"            # vibrant blauw — hoofdaccent
ACCENT_SOFT = "#60a5fa"       # lichtblauw voor hover/highlights
ACCENT_GLOW = "rgba(59,130,246,0.18)"  # glow-effect
TEAM_BLUE = "#3b82f6"         # eigen team
OPP_RED = "#f43f5e"           # tegenstander
SUCCESS_GREEN = "#10b981"     # succes
WARNING_ORANGE = "#f59e0b"    # waarschuwing
CARD_BG = "#0f1624"           # kaart-achtergrond (diep navy)
CARD_BG_ELEVATED = "#141d2f"  # iets lichtere kaart voor hover
CARD_BORDER = "#1a2540"       # kaart-rand
CARD_BORDER_SOFT = "#1e2d47"  # zachte rand
TEXT_MAIN = "#f1f5f9"         # primaire tekst
TEXT_SUB = "#94a3b8"          # secundaire tekst
TEXT_MUTED = "#64748b"        # gedempte tekst
PAGE_BG_1 = "#080c18"         # paginakleur — diep donker navy
PAGE_BG_2 = "#080c18"         # zelfde — flat


# ==================================================
# KERN HELPERS
# ==================================================
def current_elapsed_seconds() -> int:
    if st.session_state.timer_running and st.session_state.start_time is not None:
        return int(
            st.session_state.elapsed_before_run
            + (time.time() - st.session_state.start_time)
        )
    return int(st.session_state.elapsed_before_run)


def current_time_str() -> str:
    total = current_elapsed_seconds()
    return f"{total // 60:02d}:{total % 60:02d}"


def parse_mmss(value: str) -> int:
    try:
        mm, ss = str(value).split(":")
        return int(mm) * 60 + int(ss)
    except Exception:
        return 0


def format_seconds_to_mmss(seconds) -> str:
    try:
        seconds = int(max(0, float(seconds)))
    except Exception:
        seconds = 0
    return f"{seconds // 60:02d}:{seconds % 60:02d}"


def percent(numerator: int, denominator: int) -> float:
    return (numerator / denominator * 100) if denominator > 0 else 0.0


def is_probable_video_url(value: str) -> bool:
    text = str(value).strip().lower()
    if not text:
        return False
    return text.startswith("http://") or text.startswith("https://")


def is_youtube_url(value: str) -> bool:
    text = str(value).strip().lower()
    return ("youtube.com" in text) or ("youtu.be" in text)


def extract_youtube_id(value: str) -> str:
    text = str(value).strip()
    m = re.search(r"(?:v=|youtu\.be/|embed/)([A-Za-z0-9_-]{11})", text)
    return m.group(1) if m else ""


def build_youtube_share_url(video_id: str, start_sec: float, end_sec: float | None = None) -> str:
    """Bouw een deelbare YouTube-link met start (en eventueel eind) timestamp.
    Gebruik youtu.be-formaat voor eenvoud en betere mobiele ondersteuning.
    """
    if not video_id:
        return ""
    start = max(0, int(round(float(start_sec))))
    url = f"https://youtu.be/{video_id}?t={start}"
    if end_sec is not None:
        end = max(start + 1, int(round(float(end_sec))))
        url = f"https://www.youtube.com/watch?v={video_id}&start={start}&end={end}"
    return url


def log_cloud_error(action: str, err: Exception) -> None:
    """Onthoud een cloud-fout zodat de gebruiker m kan zien (in plaats van stil falen)."""
    msg = f"{time.strftime('%H:%M:%S')} • {action}: {type(err).__name__}"
    errors = st.session_state.get("cloud_errors", [])
    errors.append(msg)
    # Hou alleen de laatste 10 bij, anders groeit de sessie te groot
    st.session_state.cloud_errors = errors[-10:]
    st.session_state.last_cloud_ok = False


def mark_cloud_ok() -> None:
    st.session_state.last_cloud_ok = True


def normalize_event_row(row: dict) -> dict:
    return {
        "id": row.get("id", str(uuid.uuid4())),
        "match_id": row.get("match_id", st.session_state.match_id),
        "quarter": row.get("quarter", "Q1"),
        "time": row.get("time", "00:00"),
        "team": row.get("team", ""),
        "event": row.get("event", ""),
        "zone": row.get("zone", ""),
        "notes": row.get("notes", ""),
        "created_at": row.get("created_at", time.time()),
        "source": row.get("source", "live"),
        "video_time_sec": row.get("video_time_sec", None),
        "player_id": row.get("player_id", None),
    }


def build_df() -> pd.DataFrame:
    cols = [
        "id", "match_id", "quarter", "time", "team", "event", "zone",
        "notes", "created_at", "source", "video_time_sec", "player_id",
    ]
    if not st.session_state.events:
        return pd.DataFrame(columns=cols)
    df = pd.DataFrame(st.session_state.events)
    for col in cols:
        if col not in df.columns:
            df[col] = "" if col != "video_time_sec" else None
    return df[cols]


def build_clips_df() -> pd.DataFrame:
    cols = [
        "id", "match_id", "video_name", "clip_title", "tag", "team_focus",
        "quarter", "start_sec", "end_sec", "start_time", "end_time",
        "duration_sec", "tactical_note", "coaching_action", "created_at",
        "snapshot_name",
    ]
    if not st.session_state.video_clips:
        return pd.DataFrame(columns=cols)
    df = pd.DataFrame(st.session_state.video_clips)
    for col in cols:
        if col not in df.columns:
            df[col] = ""
    return df[cols]


def count_events(df: pd.DataFrame, team: str, event: str, quarter: str | None = None) -> int:
    if df.empty:
        return 0
    mask = (df["team"] == team) & (df["event"] == event)
    if quarter is not None:
        mask = mask & (df["quarter"] == quarter)
    return int(mask.sum())


def dominant_zone_text(df, team, quarter=None, event="Cirkelentry") -> str:
    if df.empty:
        return "onbekend"
    mask = (df["team"] == team) & (df["event"] == event)
    if quarter is not None:
        mask = mask & (df["quarter"] == quarter)
    zone_counts = df.loc[mask, "zone"].value_counts()
    if zone_counts.empty:
        return "onbekend"
    return str(zone_counts.idxmax()).lower()


def set_new_match_id() -> None:
    base = f"wedstrijd-{uuid.uuid4().hex[:6]}"
    # Scope aan huidig team, zodat match-lijsten gescheiden blijven per team
    st.session_state.match_id = scope_match_id(base) if st.session_state.get("active_team_id") else base
    # Zorg dat wisselschema de nieuwe wedstrijd oppikt (auto-laad of vers starten)
    st.session_state["subs_linked_match_id"] = None
    st.session_state["subs_schema"] = None


def recalc_score() -> None:
    df = build_df()
    st.session_state.score_team = count_events(df, st.session_state.team_name, "Goal")
    st.session_state.score_opponent = count_events(df, st.session_state.opponent_name, "Goal")


# ==================================================
# EVENT MUTATIES (waren ontbrekend — nu gereconstrueerd)
# ==================================================
def add_event(team: str, event: str, zone: str = "", source: str = "live",
              video_time_sec: float | None = None, notes: str = "",
              player_id: str | None = None) -> dict:
    """Voeg een event toe aan de lijst en optioneel aan de cloud."""
    row = normalize_event_row({
        "id": str(uuid.uuid4()),
        "match_id": st.session_state.match_id,
        "quarter": st.session_state.quarter,
        "time": current_time_str() if source == "live" else (
            format_seconds_to_mmss(game_time_from_video(video_time_sec)) if video_time_sec is not None else "00:00"
        ),
        "team": team,
        "event": event,
        "zone": zone,
        "notes": notes,
        "created_at": time.time(),
        "source": source,
        "video_time_sec": video_time_sec,
        "player_id": player_id,
    })
    st.session_state.events.append(row)
    try:
        save_event_to_cloud(row)
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("event opslaan", err)
    refresh_derived_state()
    return row


def _maybe_mark_goal_scorer_pending(row: dict, team: str, event: str) -> None:
    """Zet last_goal_event_id als dit een goal van het eigen team is.

    De UI kan dan een scorer-picker tonen voor dit event.
    """
    if event != "Goal":
        return
    own = st.session_state.get("team_name", "")
    if own and team == own:
        st.session_state["last_goal_event_id"] = row.get("id")


def start_smart_tag(team: str, event: str) -> None:
    """Start een tag-flow. Als event een zone nodig heeft, wacht op zonekeuze; anders direct opslaan."""
    if event in EVENT_NEEDS_ZONE:
        st.session_state.pending_event = event
        st.session_state.pending_team = team
    else:
        row = add_event(team=team, event=event, zone="", source="live")
        _maybe_mark_goal_scorer_pending(row, team, event)
        st.session_state.pending_event = None
        st.session_state.pending_team = None


def add_smart_event(team: str, event: str, zone: str = "") -> None:
    """Sla event op met zone en ruim pending op."""
    row = add_event(team=team, event=event, zone=zone, source="live")
    _maybe_mark_goal_scorer_pending(row, team, event)
    st.session_state.pending_event = None
    st.session_state.pending_team = None


def clear_pending_tag() -> None:
    st.session_state.pending_event = None
    st.session_state.pending_team = None


def remove_last_event() -> None:
    if not st.session_state.events:
        st.session_state.last_undo_msg = "Geen events om ongedaan te maken."
        return
    try:
        delete_last_event_cloud()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("laatste event verwijderen", err)
    removed = st.session_state.events.pop()
    # Bewaar in undo-stack (max 20) zodat herstel mogelijk is
    stack = st.session_state.get("undo_stack", [])
    stack.append(removed)
    st.session_state.undo_stack = stack[-20:]
    st.session_state.last_undo_msg = (
        f"Ongedaan gemaakt: {removed.get('team', '?')} • {removed.get('event', '?')} "
        f"({removed.get('quarter', '?')} {removed.get('time', '?')})"
    )
    refresh_derived_state()


def redo_last_event() -> None:
    """Haal het laatst verwijderde event terug."""
    stack = st.session_state.get("undo_stack", [])
    if not stack:
        st.session_state.last_undo_msg = "Niets om te herstellen."
        return
    row = stack.pop()
    st.session_state.undo_stack = stack
    st.session_state.events.append(row)
    try:
        save_event_to_cloud(row)
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("event herstellen in cloud", err)
    refresh_derived_state()
    st.session_state.last_undo_msg = (
        f"Hersteld: {row.get('team', '?')} • {row.get('event', '?')} "
        f"({row.get('quarter', '?')} {row.get('time', '?')})"
    )


def reset_all() -> None:
    try:
        reset_match_cloud()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("wedstrijd resetten in cloud", err)
    st.session_state.events = []
    st.session_state.video_clips = []
    st.session_state.pending_event = None
    st.session_state.pending_team = None
    st.session_state.halftime_report = ""
    st.session_state.auto_notes = ""
    st.session_state.confirm_reset = False
    st.session_state.pushoff_offsets = {q: None for q in QUARTERS}
    reset_timer()
    refresh_derived_state()


def add_video_clip(video_name: str, clip_title: str, tag: str, team_focus: str,
                   quarter: str, start_sec: int, end_sec: int, tactical_note: str,
                   coaching_action: str, snapshot_name: str = "") -> None:
    start_sec = int(max(0, start_sec))
    end_sec = int(max(start_sec, end_sec))
    row = {
        "id": str(uuid.uuid4()),
        "match_id": st.session_state.match_id,
        "video_name": video_name or "geen video",
        "clip_title": clip_title or f"Clip {len(st.session_state.video_clips) + 1}",
        "tag": tag,
        "team_focus": team_focus,
        "quarter": quarter,
        "start_sec": start_sec,
        "end_sec": end_sec,
        "start_time": format_seconds_to_mmss(start_sec),
        "end_time": format_seconds_to_mmss(end_sec),
        "duration_sec": end_sec - start_sec,
        "tactical_note": tactical_note,
        "coaching_action": coaching_action,
        "created_at": time.time(),
        "snapshot_name": snapshot_name,
    }
    st.session_state.video_clips.append(row)


def remove_last_clip() -> None:
    if st.session_state.video_clips:
        st.session_state.video_clips.pop()


# ==================================================
# PUSH-OFF / VIDEO-TIJD HELPERS
# ==================================================
def get_pushoff_offset(quarter: str) -> float | None:
    """Videotijd (in seconden) die overeenkomt met wedstrijdtijd 0:00 voor dit kwart."""
    return st.session_state.pushoff_offsets.get(quarter)


def video_time_from_game(game_time_sec: int, quarter: str) -> float | None:
    """Zet wedstrijdtijd om naar videotijd voor een gegeven kwart."""
    offset = get_pushoff_offset(quarter)
    if offset is None:
        return None
    return float(offset) + float(game_time_sec)


def game_time_from_video(video_time_sec: float | None, quarter: str | None = None) -> int:
    """Zet videotijd om naar wedstrijdtijd (seconden). Gebruikt actieve kwart als niet opgegeven."""
    if video_time_sec is None:
        return 0
    q = quarter or st.session_state.quarter
    offset = get_pushoff_offset(q)
    if offset is None:
        return int(max(0, video_time_sec))
    return int(max(0, float(video_time_sec) - float(offset)))


# ==================================================
# ANALYSE HELPERS
# ==================================================
def build_kpi_summary(df: pd.DataFrame) -> dict:
    team = st.session_state.team_name
    opp = st.session_state.opponent_name
    team_entries = count_events(df, team, "Cirkelentry")
    opp_entries = count_events(df, opp, "Cirkelentry")
    team_shots = count_events(df, team, "Schot")
    team_shots_on_goal = count_events(df, team, "Schot op goal")
    opp_shots = count_events(df, opp, "Schot")
    opp_shots_on_goal = count_events(df, opp, "Schot op goal")
    team_total_attempts = team_shots + team_shots_on_goal
    opp_total_attempts = opp_shots + opp_shots_on_goal
    team_goals = count_events(df, team, "Goal")
    opp_goals = count_events(df, opp, "Goal")
    team_high_wins = count_events(df, team, "Hoge balverovering")
    team_turnovers_own = count_events(df, team, "Turnover eigen helft")
    team_counters_against = count_events(df, team, "Counter tegen na balverlies")
    team_press_success = count_events(df, team, "Press succes")
    team_build_fail = count_events(df, team, "Opbouw mislukt")
    return {
        "team_entries": team_entries,
        "opp_entries": opp_entries,
        "team_shots": team_shots,
        "team_shots_on_goal": team_shots_on_goal,
        "team_total_attempts": team_total_attempts,
        "opp_shots": opp_shots,
        "opp_shots_on_goal": opp_shots_on_goal,
        "opp_total_attempts": opp_total_attempts,
        "team_goals": team_goals,
        "opp_goals": opp_goals,
        "team_high_wins": team_high_wins,
        "team_turnovers_own": team_turnovers_own,
        "team_counters_against": team_counters_against,
        "team_press_success": team_press_success,
        "team_build_fail": team_build_fail,
        "team_entry_to_shot_pct": percent(team_total_attempts, team_entries),
        "opp_entry_to_shot_pct": percent(opp_total_attempts, opp_entries),
        "team_on_goal_pct": percent(team_shots_on_goal, team_total_attempts),
        "opp_on_goal_pct": percent(opp_shots_on_goal, opp_total_attempts),
        "team_shot_to_goal_pct": percent(team_goals, team_shots_on_goal),
        "opp_shot_to_goal_pct": percent(opp_goals, opp_shots_on_goal),
        "team_highwin_to_entry_pct": percent(team_entries, team_high_wins),
        "team_turnover_to_counter_pct": percent(team_counters_against, team_turnovers_own),
    }


def generate_tactical_patterns(df: pd.DataFrame) -> list[str]:
    if df.empty:
        return []
    team = st.session_state.team_name
    patterns = []
    entries = df[(df["team"] == team) & (df["event"] == "Cirkelentry")]
    if not entries.empty:
        zone_counts = entries["zone"].value_counts()
        total = zone_counts.sum()
        top_zone = zone_counts.idxmax()
        top_pct = percent(zone_counts.max(), total)
        if top_pct >= 50:
            patterns.append(
                f"{top_pct:.0f}% van de cirkelentries van {team} kwam via {str(top_zone).lower()}."
            )
    if count_events(df, team, "Opbouw mislukt") >= 3:
        patterns.append(f"{team} heeft meerdere mislukte opbouwmomenten onder druk.")
    if count_events(df, team, "Press succes") >= 3:
        patterns.append(f"De press van {team} levert herhaald succes op.")
    return patterns


def detect_momentum(df: pd.DataFrame) -> list[str]:
    if df.empty:
        return []
    moments = []
    entries = df[df["event"] == "Cirkelentry"].copy()
    if len(entries) >= 3:
        entries["sec"] = entries["time"].apply(parse_mmss)
        entries = entries.sort_values("sec")
        for i in range(len(entries) - 2):
            if entries.iloc[i + 2]["sec"] - entries.iloc[i]["sec"] <= 120:
                moments.append("3 cirkelentries binnen 2 minuten → sterke aanvalsfase")
                break
    if (
        count_events(df, st.session_state.team_name, "Turnover eigen helft") >= 2
        and count_events(df, st.session_state.team_name, "Counter tegen na balverlies") >= 1
    ):
        moments.append("Balverlies eigen helft leidt tot counters tegen")
    return moments


def build_entry_heatmap(df: pd.DataFrame) -> pd.DataFrame:
    entries = df[df["event"] == "Cirkelentry"]
    zones = {
        "Linksvoor": len(entries[entries["zone"] == "Linksvoor"]),
        "Middenvoor": len(entries[entries["zone"] == "Middenvoor"]),
        "Rechtsvoor": len(entries[entries["zone"] == "Rechtsvoor"]),
    }
    total = sum(zones.values())
    return pd.DataFrame(
        [{"zone": z, "entries": v, "pct": round(percent(v, total), 1)} for z, v in zones.items()]
    )


def build_quarter_stats_df(df: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "Kwart", "Entries voor", "Schoten", "Schoten op goal", "Totaal pogingen",
        "Goals voor", "Entry->poging %", "Op goal %", "Shot on goal->goal %",
        "Press succes", "Hoge balverovering", "Turnover eigen helft",
        "Counter tegen", "Entries tegen", "Schoten tegen", "Schoten op goal tegen",
        "Totaal pogingen tegen", "Goals tegen", "Tegen entry->poging %",
        "Tegen op goal %", "Tegen shot on goal->goal %",
    ]
    if df.empty:
        return pd.DataFrame(columns=cols)
    rows = []
    team = st.session_state.team_name
    opp = st.session_state.opponent_name
    for quarter in QUARTERS:
        team_entries = count_events(df, team, "Cirkelentry", quarter)
        team_shots = count_events(df, team, "Schot", quarter)
        team_shots_on_goal = count_events(df, team, "Schot op goal", quarter)
        team_total_attempts = team_shots + team_shots_on_goal
        team_goals = count_events(df, team, "Goal", quarter)
        opp_entries = count_events(df, opp, "Cirkelentry", quarter)
        opp_shots = count_events(df, opp, "Schot", quarter)
        opp_shots_on_goal = count_events(df, opp, "Schot op goal", quarter)
        opp_total_attempts = opp_shots + opp_shots_on_goal
        opp_goals = count_events(df, opp, "Goal", quarter)
        rows.append({
            "Kwart": quarter,
            "Entries voor": team_entries,
            "Schoten": team_shots,
            "Schoten op goal": team_shots_on_goal,
            "Totaal pogingen": team_total_attempts,
            "Goals voor": team_goals,
            "Entry->poging %": round(percent(team_total_attempts, team_entries), 1),
            "Op goal %": round(percent(team_shots_on_goal, team_total_attempts), 1),
            "Shot on goal->goal %": round(percent(team_goals, team_shots_on_goal), 1),
            "Press succes": count_events(df, team, "Press succes", quarter),
            "Hoge balverovering": count_events(df, team, "Hoge balverovering", quarter),
            "Turnover eigen helft": count_events(df, team, "Turnover eigen helft", quarter),
            "Counter tegen": count_events(df, team, "Counter tegen na balverlies", quarter),
            "Entries tegen": opp_entries,
            "Schoten tegen": opp_shots,
            "Schoten op goal tegen": opp_shots_on_goal,
            "Totaal pogingen tegen": opp_total_attempts,
            "Goals tegen": opp_goals,
            "Tegen entry->poging %": round(percent(opp_total_attempts, opp_entries), 1),
            "Tegen op goal %": round(percent(opp_shots_on_goal, opp_total_attempts), 1),
            "Tegen shot on goal->goal %": round(percent(opp_goals, opp_shots_on_goal), 1),
        })
    return pd.DataFrame(rows)


def build_event_summary_per_quarter(df: pd.DataFrame) -> dict:
    summary = {}
    if df.empty:
        return summary
    for quarter in QUARTERS:
        qdf = df[df["quarter"] == quarter].copy()
        quarter_summary = {
            st.session_state.team_name: [],
            st.session_state.opponent_name: [],
        }
        for team in [st.session_state.team_name, st.session_state.opponent_name]:
            tdf = qdf[qdf["team"] == team].copy()
            if tdf.empty:
                quarter_summary[team] = []
                continue
            counts = (
                tdf.groupby(["event", "zone"], dropna=False)
                .size()
                .reset_index(name="count")
                .sort_values(["event", "zone"])
            )
            items = []
            for _, row in counts.iterrows():
                event_name = str(row["event"])
                zone_name = str(row["zone"]).strip() if pd.notna(row["zone"]) else ""
                count_value = int(row["count"])
                if zone_name and zone_name != "":
                    items.append(f"{event_name} ({zone_name}): {count_value}")
                else:
                    items.append(f"{event_name}: {count_value}")
            quarter_summary[team] = items
        summary[quarter] = quarter_summary
    return summary


def build_report_sections(df: pd.DataFrame) -> dict:
    if df.empty:
        return {
            "Aanval": ["Nog geen data."],
            "Press": ["Nog geen data."],
            "Omschakeling": ["Nog geen data."],
            "Verdediging": ["Nog geen data."],
            "Actiepunt": ["Nog geen data."],
        }
    team = st.session_state.team_name
    opp = st.session_state.opponent_name
    kpi = build_kpi_summary(df)
    aanval = [
        f"{team} had {kpi['team_entries']} cirkelentries, {kpi['team_shots']} schoten ({kpi['team_shots_on_goal']} op goal) en {kpi['team_goals']} goals.",
        f"Entry → poging: {kpi['team_entry_to_shot_pct']:.0f}% • op goal: {kpi['team_on_goal_pct']:.0f}% • shot on goal → goal: {kpi['team_shot_to_goal_pct']:.0f}%.",
        f"Dominante entryzone: {dominant_zone_text(df, team, event='Cirkelentry')}.",
    ]
    press = [
        f"Hoge balveroveringen: {kpi['team_high_wins']} • press successen: {kpi['team_press_success']}.",
        f"Hoge balwinst → entry: {kpi['team_highwin_to_entry_pct']:.0f}%.",
    ]
    omschakeling = [
        f"Turnovers eigen helft: {kpi['team_turnovers_own']} • counters tegen: {kpi['team_counters_against']}.",
        f"Turnover eigen helft → counter tegen: {kpi['team_turnover_to_counter_pct']:.0f}%.",
    ]
    verdediging = [
        f"{opp} had {kpi['opp_entries']} cirkelentries, {kpi['opp_shots']} schoten ({kpi['opp_shots_on_goal']} op goal) en {kpi['opp_goals']} goals.",
        f"Entries tegen → poging: {kpi['opp_entry_to_shot_pct']:.0f}% • op goal tegen: {kpi['opp_on_goal_pct']:.0f}% • shot on goal tegen → goal: {kpi['opp_shot_to_goal_pct']:.0f}%.",
    ]
    actiepunt = []
    if kpi["team_entry_to_shot_pct"] < 40 and kpi["team_entries"] > 0:
        actiepunt.append("Sneller handelen na entry en eerder tot een doelpoging komen.")
    if kpi["team_on_goal_pct"] < 40 and kpi["team_total_attempts"] > 0:
        actiepunt.append("Meer schoten tussen de palen krijgen.")
    if kpi["team_turnover_to_counter_pct"] >= 50 and kpi["team_turnovers_own"] > 0:
        actiepunt.append("Veiliger opbouwen in eigen helft.")
    if kpi["opp_entry_to_shot_pct"] > 50:
        actiepunt.append("Eerder druk zetten bij entry tegen.")
    if not actiepunt:
        actiepunt.append("Huidige balans behouden en details blijven monitoren.")
    return {
        "Aanval": aanval,
        "Press": press,
        "Omschakeling": omschakeling,
        "Verdediging": verdediging,
        "Actiepunt": actiepunt,
    }


def generate_auto_notes(df: pd.DataFrame) -> str:
    if df.empty:
        return "Nog geen data."
    team = st.session_state.team_name
    opp = st.session_state.opponent_name
    kpi = build_kpi_summary(df)
    patterns = generate_tactical_patterns(df)
    sections = build_report_sections(df)
    quarter_df = build_quarter_stats_df(df)
    quarter_event_summary = build_event_summary_per_quarter(df)
    lines = [
        f"Wedstrijd: {team} - {opp}",
        f"Score: {kpi['team_goals']}-{kpi['opp_goals']}",
        "",
    ]
    for title, items in sections.items():
        lines.append(title.upper())
        lines.extend([f"- {x}" for x in items])
        lines.append("")
    lines.append("TACTISCHE PATRONEN")
    if patterns:
        lines.extend([f"- {p}" for p in patterns])
    else:
        lines.append("- Nog geen duidelijke patronen zichtbaar.")
    lines.append("")
    lines.append("STATISTIEKEN PER KWART")
    if quarter_df.empty:
        lines.append("- Nog geen kwartstatistieken beschikbaar.")
    else:
        for _, row in quarter_df.iterrows():
            lines.append(
                f"- {row['Kwart']}: "
                f"entries voor {int(row['Entries voor'])}, "
                f"schoten {int(row['Schoten'])}, "
                f"schoten op goal {int(row['Schoten op goal'])}, "
                f"totaal pogingen {int(row['Totaal pogingen'])}, "
                f"goals {int(row['Goals voor'])}, "
                f"entries tegen {int(row['Entries tegen'])}, "
                f"schoten tegen {int(row['Schoten tegen'])}, "
                f"schoten op goal tegen {int(row['Schoten op goal tegen'])}, "
                f"totaal pogingen tegen {int(row['Totaal pogingen tegen'])}, "
                f"goals tegen {int(row['Goals tegen'])}"
            )
    lines.append("")
    lines.append("ALLE GETAGDE EVENTS PER KWART")
    for quarter in QUARTERS:
        lines.append("")
        lines.append(quarter)
        q_summary = quarter_event_summary.get(quarter, {team: [], opp: []})
        team_items = q_summary.get(team, [])
        opp_items = q_summary.get(opp, [])
        lines.append(f"- {team}:")
        if team_items:
            lines.extend([f"  - {item}" for item in team_items])
        else:
            lines.append("  - Geen events")
        lines.append(f"- {opp}:")
        if opp_items:
            lines.extend([f"  - {item}" for item in opp_items])
        else:
            lines.append("  - Geen events")
    return "\n".join(lines)


def generate_halftime_report(df: pd.DataFrame) -> str:
    if df.empty:
        return "Nog geen data voor rustanalyse."
    kpi = build_kpi_summary(df)
    strong, risk, action = [], [], []
    if kpi["team_entry_to_shot_pct"] >= 50:
        strong.append("Entries worden goed omgezet in doelpogingen.")
    if kpi["team_on_goal_pct"] >= 50 and kpi["team_total_attempts"] > 0:
        strong.append("Een groot deel van de pogingen is op goal.")
    if kpi["team_high_wins"] >= 3:
        strong.append("De press levert bruikbare balwinsten op.")
    if kpi["team_turnover_to_counter_pct"] >= 50 and kpi["team_turnovers_own"] > 0:
        risk.append("Balverlies eigen helft leidt tot counters tegen.")
        action.append("Veiliger opbouwen in eigen helft.")
    if kpi["opp_entry_to_shot_pct"] > 50:
        risk.append("Tegenstander komt te makkelijk van entry naar doelpoging.")
        action.append("Eerder druk op bal zetten bij entry tegen.")
    if kpi["team_on_goal_pct"] < 40 and kpi["team_total_attempts"] > 0:
        risk.append("Te weinig pogingen eindigen tussen de palen.")
        action.append("Meer rust in de afronding en betere shotselectie.")
    if not strong:
        strong.append("Wedstrijdbeeld is nog vrij gebalanceerd.")
    if not risk:
        risk.append("Nog geen groot dominant risico zichtbaar.")
    if not action:
        action.append("Huidige afspraken vasthouden en details blijven monitoren.")
    txt = "RUSTANALYSE\n\n"
    txt += "Sterk:\n" + "\n".join(f"- {x}" for x in strong)
    txt += "\n\nRisico:\n" + "\n".join(f"- {x}" for x in risk)
    txt += "\n\nActie:\n" + "\n".join(f"- {x}" for x in action)
    return txt


def generate_video_analysis_summary(clips_df: pd.DataFrame) -> str:
    if clips_df.empty:
        return "Nog geen clips geregistreerd."
    lines = []
    lines.append(f"Beeldanalyse • {st.session_state.team_name} - {st.session_state.opponent_name}")
    lines.append(f"Wedstrijd-ID: {st.session_state.match_id}")
    lines.append(f"Aantal clips: {len(clips_df)}")
    lines.append("")
    lines.append("Verdeling per thema:")
    for tag, count in clips_df["tag"].value_counts().items():
        lines.append(f"- {tag}: {count}")
    lines.append("")
    lines.append("Clips:")
    ordered = clips_df.sort_values(["quarter", "start_sec", "created_at"])
    for _, row in ordered.iterrows():
        lines.append(
            f"- {row['quarter']} • {row['start_time']} - {row['end_time']} • {row['clip_title']} • {row['tag']}"
        )
        if str(row["tactical_note"]).strip():
            lines.append(f"  Analyse: {row['tactical_note']}")
        if str(row["coaching_action"]).strip():
            lines.append(f"  Coachactie: {row['coaching_action']}")
    return "\n".join(lines)


def refresh_derived_state() -> None:
    recalc_score()
    df = build_df()
    st.session_state.auto_notes = generate_auto_notes(df)
    st.session_state.last_sync_count = len(df)


# ==================================================
# RECONSTRUCTED RENDER HELPERS
# ==================================================
def get_insight_cards(df: pd.DataFrame) -> list[dict]:
    """Vier compacte inzicht-kaartjes voor in de LIVE-tab."""
    if df.empty:
        return [
            {"title": "Cirkelentries", "value": "0", "subtitle": "Nog geen data"},
            {"title": "Schoten op goal", "value": "0", "subtitle": "Nog geen data"},
            {"title": "Press succes", "value": "0", "subtitle": "Nog geen data"},
            {"title": "Turnovers eigen helft", "value": "0", "subtitle": "Nog geen data"},
        ]
    kpi = build_kpi_summary(df)
    return [
        {
            "title": "Cirkelentries",
            "value": str(kpi["team_entries"]),
            "subtitle": f"Tegen: {kpi['opp_entries']}",
        },
        {
            "title": "Schoten op goal",
            "value": str(kpi["team_shots_on_goal"]),
            "subtitle": f"Tegen: {kpi['opp_shots_on_goal']}",
        },
        {
            "title": "Press succes",
            "value": str(kpi["team_press_success"]),
            "subtitle": f"Hoge balwinst: {kpi['team_high_wins']}",
        },
        {
            "title": "Turnovers eigen helft",
            "value": str(kpi["team_turnovers_own"]),
            "subtitle": f"Counters tegen: {kpi['team_counters_against']}",
        },
    ]


def render_event_feed(df: pd.DataFrame, max_items: int = 10) -> None:
    """Toon laatste events als rustige rijen."""
    if df.empty:
        st.markdown(
            f"<div class='mini-feed' style='color:{TEXT_MUTED};text-align:center;'>Nog geen events</div>",
            unsafe_allow_html=True,
        )
        return
    last = df.tail(max_items).iloc[::-1]
    for _, row in last.iterrows():
        team = str(row.get("team", ""))
        is_own = team == st.session_state.team_name
        accent = TEAM_BLUE if is_own else OPP_RED
        zone_txt = f" · {row['zone']}" if str(row.get("zone", "")).strip() else ""
        source = str(row.get("source", "live"))
        source_pill = "pill-gray" if source == "live" else "pill-green"
        # Alles op één regel — zonder inspringen — anders rendert markdown het als codeblok
        html = (
            f'<div class="mini-feed" style="border-left:3px solid {accent};display:flex;align-items:center;justify-content:space-between;">'
            f'<div>'
            f'<span style="color:{TEXT_MUTED};font-size:12px;font-weight:500;margin-right:10px;">{row["quarter"]} · {row["time"]}</span>'
            f'<strong>{row["event"]}</strong><span style="color:{TEXT_SUB};">{zone_txt}</span>'
            f'</div>'
            f'<div>'
            f'<span class="pill {source_pill}">{source}</span>'
            f'</div>'
            f'</div>'
        )
        st.markdown(html, unsafe_allow_html=True)


def render_timeline(df: pd.DataFrame) -> None:
    """Premium tijdlijn: elk event als bolletje op een horizontale as per kwart."""
    if df.empty:
        st.info("Nog geen events voor tijdlijn.")
        return
    df = df.copy()
    df["sec"] = df["time"].apply(parse_mmss)
    max_sec = max(int(df["sec"].max()), 15 * 60)
    team = st.session_state.team_name
    for quarter in QUARTERS:
        qdf = df[df["quarter"] == quarter].sort_values("sec")
        if qdf.empty:
            continue
        dots_html = ""
        for _, row in qdf.iterrows():
            pct = (row["sec"] / max_sec) * 100
            is_own = str(row["team"]) == team
            color = TEAM_BLUE if is_own else OPP_RED
            title = f"{row['time']} • {row['team']} • {row['event']}"
            dots_html += (
                f'<div title="{title}" style="position:absolute;left:{pct:.1f}%;top:50%;'
                f'transform:translate(-50%,-50%);width:11px;height:11px;border-radius:50%;'
                f'background:{color};border:2px solid #0f172a;box-shadow:0 0 0 1px {color}55, 0 2px 6px rgba(0,0,0,0.4);"></div>'
            )
        st.markdown(
            f"<div style='margin:10px 0;'>"
            f"<div style='color:#9ca3af;font-size:12px;font-weight:600;letter-spacing:0.04em;text-transform:uppercase;margin-bottom:6px;'>{quarter}</div>"
            f"<div style='position:relative;height:22px;background:linear-gradient(90deg,#1e293b,#0f172a);border:1px solid #1f2937;border-radius:11px;'>"
            f"{dots_html}</div></div>",
            unsafe_allow_html=True,
        )


def render_field_view(df: pd.DataFrame, team: str, quarter: str, layers: list[str]) -> None:
    """Realistisch half-veld in SVG met cirkel, 23m-lijn en zone-tellers."""
    if quarter == "Alles":
        tdf = df[df["team"] == team].copy()
    else:
        tdf = df[(df["team"] == team) & (df["quarter"] == quarter)].copy()

    # Counts per zone (alleen cirkelentries hebben zone-data)
    entry_counts = {
        z: int(((tdf["event"] == "Cirkelentry") & (tdf["zone"] == z)).sum())
        for z in FIELD_ZONES
    }
    total_shots = int((tdf["event"] == "Schot").sum()) + int((tdf["event"] == "Schot op goal").sum())
    total_goals = int((tdf["event"] == "Goal").sum())

    show_entries = "Cirkelentry" in layers
    show_shots = "Schot" in layers
    show_goals = "Goal" in layers

    # SVG coördinaten — aanvallende helft van boven gezien
    # Veldafmeting in viewport (55x45 units); goal bovenaan
    field_bg = "#0d5d3a"   # diepe veldgroen
    line = "#e5e7eb"       # witte lijnen
    circle_fill = "rgba(229,231,235,0.04)"

    # Zone-teksten (links / midden / rechts boven in cirkel-gebied)
    def zone_label(x, label, count):
        if not show_entries:
            return ""
        # Eén dot met getal + label eronder
        return (
            f'<g>'
            f'<circle cx="{x}" cy="12" r="4" fill="{ACCENT}" opacity="0.9"/>'
            f'<text x="{x}" y="13.2" text-anchor="middle" fill="white" font-size="3.4" font-weight="700">{count}</text>'
            f'<text x="{x}" y="20" text-anchor="middle" fill="{line}" font-size="2.6" font-weight="600" opacity="0.85">{label}</text>'
            f'</g>'
        )

    dots_svg = (
        zone_label(18, "Links", entry_counts["Linksvoor"]) +
        zone_label(27.5, "Midden", entry_counts["Middenvoor"]) +
        zone_label(37, "Rechts", entry_counts["Rechtsvoor"])
    )

    # Totalen onderaan
    totals_lines = []
    if show_shots:
        totals_lines.append(f'<span class="pill pill-gray">Schoten: {total_shots}</span>')
    if show_goals:
        totals_lines.append(f'<span class="pill pill-green">Goals: {total_goals}</span>')
    if show_entries:
        total_entries = sum(entry_counts.values())
        totals_lines.insert(0, f'<span class="pill pill-blue">Entries: {total_entries}</span>')
    totals_bar = "".join(totals_lines)

    svg = f"""
    <svg viewBox="0 0 55 45" preserveAspectRatio="xMidYMid meet" style="width:100%;max-width:560px;height:auto;display:block;margin:0 auto;">
        <!-- veld achtergrond -->
        <rect x="0" y="0" width="55" height="45" fill="{field_bg}" rx="1.5"/>
        <!-- buitenranden -->
        <rect x="1" y="1" width="53" height="43" fill="none" stroke="{line}" stroke-width="0.3" rx="0.4"/>
        <!-- 23 meter lijn (onderkant) -->
        <line x1="1" y1="30" x2="54" y2="30" stroke="{line}" stroke-width="0.25" stroke-dasharray="0.8,0.6"/>
        <text x="3" y="32.5" fill="{line}" font-size="2" opacity="0.7">23m</text>
        <!-- schietcirkel (D) -->
        <path d="M 16 1 A 12 12 0 0 0 39 1" fill="{circle_fill}" stroke="{line}" stroke-width="0.3"/>
        <!-- goal -->
        <rect x="24" y="0.3" width="7" height="1.2" fill="{line}"/>
        <!-- strafstippel -->
        <circle cx="27.5" cy="7.5" r="0.35" fill="{line}"/>
        <!-- zone-aanduidingen -->
        {dots_svg}
    </svg>
    """

    header = f"<div style='color:{TEXT_SUB};font-size:13px;font-weight:500;margin-bottom:10px;text-align:center;'>Aanvallende helft · <strong style='color:{TEXT_MAIN};'>{team}</strong> · {quarter}</div>"
    body = f"<div>{svg}</div>"
    footer = f"<div style='display:flex;gap:8px;justify-content:center;margin-top:14px;flex-wrap:wrap;'>{totals_bar}</div>"

    html = f"""
    <div style="background:{CARD_BG};border:1px solid {CARD_BORDER_SOFT};border-radius:14px;padding:22px 18px;">
        {header}
        {body}
        {footer}
    </div>
    """
    st.markdown(html, unsafe_allow_html=True)


# ==================================================
# EXPORT HELPERS
# ==================================================
def export_pdf_report(text: str) -> bytes:
    """Legacy simpele tekst-PDF. Behouden voor bestaande aanroepen."""
    if not REPORTLAB_AVAILABLE:
        return text.encode("utf-8")
    buffer = BytesIO()
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    for line in text.split("\n"):
        safe = line if line.strip() else " "
        safe = safe.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        story.append(Paragraph(safe, styles["Normal"]))
        story.append(Spacer(1, 6))
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


# ==== Uitgebreide PDF-exports (wedstrijdrapport / wisselschema / seizoen) ====

def _pdf_base_styles():
    """Centrale stijlen voor alle PDF-exports."""
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="CoachH1",
        parent=styles["Heading1"],
        fontSize=22, leading=26,
        textColor=_rl_colors.HexColor("#111827"),
        spaceAfter=10,
    ))
    styles.add(ParagraphStyle(
        name="CoachH2",
        parent=styles["Heading2"],
        fontSize=15, leading=19,
        textColor=_rl_colors.HexColor("#1f2937"),
        spaceBefore=12, spaceAfter=6,
    ))
    styles.add(ParagraphStyle(
        name="CoachBody",
        parent=styles["Normal"],
        fontSize=10, leading=14,
        textColor=_rl_colors.HexColor("#1f2937"),
    ))
    styles.add(ParagraphStyle(
        name="CoachSub",
        parent=styles["Normal"],
        fontSize=9, leading=12,
        textColor=_rl_colors.HexColor("#6b7280"),
    ))
    styles.add(ParagraphStyle(
        name="CoachKpi",
        parent=styles["Heading1"],
        fontSize=20, leading=22,
        textColor=_rl_colors.HexColor("#2563eb"),
        alignment=1,
    ))
    return styles


def _esc(x) -> str:
    return str(x).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def export_match_report_pdf(df: pd.DataFrame, team_name: str, opponent_name: str,
                            match_id: str = "") -> bytes:
    """Professioneel wedstrijdrapport: scorebord, KPI's, kwart-splitsing, notes."""
    if not REPORTLAB_AVAILABLE:
        return b"PDF-export niet beschikbaar (reportlab ontbreekt)"
    buf = BytesIO()
    styles = _pdf_base_styles()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=18 * mm, rightMargin=18 * mm,
                            topMargin=16 * mm, bottomMargin=16 * mm)
    story = []

    # Titelblok
    story.append(Paragraph(f"Wedstrijdrapport — {_esc(team_name)} vs {_esc(opponent_name)}",
                           styles["CoachH1"]))
    sub = time.strftime("%d-%m-%Y · %H:%M") + (f" · {_esc(match_id)}" if match_id else "")
    story.append(Paragraph(sub, styles["CoachSub"]))
    story.append(Spacer(1, 6))

    # Scorebord
    own_goals = count_events(df, team_name, "Goal")
    opp_goals = count_events(df, opponent_name, "Goal")
    score_table = Table(
        [[
            Paragraph(f"<b>{_esc(team_name)}</b>", styles["CoachBody"]),
            Paragraph(f"<font size=22 color='#2563eb'><b>{own_goals}</b></font>", styles["CoachBody"]),
            Paragraph("<font color='#6b7280'>–</font>", styles["CoachBody"]),
            Paragraph(f"<font size=22 color='#dc2626'><b>{opp_goals}</b></font>", styles["CoachBody"]),
            Paragraph(f"<b>{_esc(opponent_name)}</b>", styles["CoachBody"]),
        ]],
        colWidths=[60 * mm, 20 * mm, 10 * mm, 20 * mm, 60 * mm],
    )
    score_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _rl_colors.HexColor("#f3f4f6")),
        ("BOX", (0, 0), (-1, -1), 0.6, _rl_colors.HexColor("#d1d5db")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(score_table)
    story.append(Spacer(1, 14))

    # KPI's
    kpi = build_kpi_summary(df)
    story.append(Paragraph("Kerngetallen", styles["CoachH2"]))
    kpi_rows = [["Kengetal", team_name, opponent_name]]
    labels = [
        ("Goals", "goals"),
        ("Cirkelentries", "entries"),
        ("Strafcorners", "sc"),
        ("Long corners", "lc"),
        ("Balverlies", "bv"),
    ]
    for nice, key in labels:
        own_val = kpi.get(f"own_{key}") if isinstance(kpi, dict) else None
        opp_val = kpi.get(f"opp_{key}") if isinstance(kpi, dict) else None
        kpi_rows.append([nice, str(own_val if own_val is not None else "—"),
                         str(opp_val if opp_val is not None else "—")])
    kpi_t = Table(kpi_rows, colWidths=[70 * mm, 50 * mm, 50 * mm])
    kpi_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _rl_colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), _rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, _rl_colors.HexColor("#d1d5db")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [_rl_colors.white, _rl_colors.HexColor("#f9fafb")]),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(kpi_t)
    story.append(Spacer(1, 12))

    # Per-kwart tabel
    try:
        q_df = build_quarter_stats_df(df)
    except Exception:
        q_df = None
    if q_df is not None and not q_df.empty:
        story.append(Paragraph("Per kwart", styles["CoachH2"]))
        data = [list(q_df.columns)] + q_df.astype(str).values.tolist()
        qt = Table(data, repeatRows=1)
        qt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _rl_colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), _rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.4, _rl_colors.HexColor("#d1d5db")),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ]))
        story.append(qt)
        story.append(Spacer(1, 12))

    # Auto-notes / tactische patronen
    try:
        notes = generate_auto_notes(df) or ""
    except Exception:
        notes = ""
    if notes.strip():
        story.append(Paragraph("Auto-notities", styles["CoachH2"]))
        for line in notes.split("\n"):
            if line.strip():
                story.append(Paragraph(_esc(line), styles["CoachBody"]))
                story.append(Spacer(1, 2))
        story.append(Spacer(1, 6))

    try:
        patterns = generate_tactical_patterns(df) or []
    except Exception:
        patterns = []
    if patterns:
        story.append(Paragraph("Tactische patronen", styles["CoachH2"]))
        for p in patterns:
            story.append(Paragraph(f"• {_esc(p)}", styles["CoachBody"]))
            story.append(Spacer(1, 2))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def export_lineup_pdf(schema: dict, settings: dict, players: list,
                      team_name: str, opponent_name: str) -> bytes:
    """Printbaar wisselschema-PDF: tabel met 1-en per minuut per speler, per linie."""
    if not REPORTLAB_AVAILABLE:
        return b"PDF-export niet beschikbaar (reportlab ontbreekt)"
    buf = BytesIO()
    styles = _pdf_base_styles()
    # Landscape zodat alle minuten op één pagina passen
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm)
    story = []

    story.append(Paragraph(f"Wisselschema — {_esc(team_name)} vs {_esc(opponent_name)}",
                           styles["CoachH1"]))
    date = settings.get("match_date") or time.strftime("%Y-%m-%d")
    form = settings.get("formation", "?")
    halves = int(settings.get("halves", 2))
    halflen = int(settings.get("half_length", 17))
    story.append(Paragraph(f"{_esc(date)} · formatie {_esc(form)} · {halves}×{halflen} min",
                           styles["CoachSub"]))
    story.append(Spacer(1, 8))

    cells = schema.get("cells") or {}
    minutes_per_half = int(schema.get("minutes_per_half", halflen))
    total_minutes = int(schema.get("total_minutes", minutes_per_half * halves))

    # Header: Speler, 1..N minuten, Tot
    header = ["Speler"] + [str(i + 1) for i in range(total_minutes)] + ["Tot"]
    data = [header]

    # Groepeer per linie
    by_line = {"K": [], "V": [], "M": [], "A": []}
    for p in players:
        by_line.setdefault(p.get("line", "M"), []).append(p)

    body_style = TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, _rl_colors.HexColor("#d1d5db")),
        ("BACKGROUND", (0, 0), (-1, 0), _rl_colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), _rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ])

    # Data-rijen bouwen
    line_colors = {"K": "#fef3c7", "V": "#dbeafe", "M": "#d1fae5", "A": "#fee2e2"}
    row_offset = 1
    for L in ("K", "V", "M", "A"):
        lp = sorted(by_line.get(L, []), key=lambda x: x.get("name", ""))
        if not lp:
            continue
        for p in lp:
            pc = cells.get(p["id"], [0] * total_minutes)
            vals = [("●" if (idx < len(pc) and pc[idx]) else "") for idx in range(total_minutes)]
            total = sum(1 for v in vals if v)
            data.append([p.get("name", "?")] + vals + [str(total)])
            # Rijkleur per linie
            body_style.add("BACKGROUND", (0, row_offset), (0, row_offset),
                           _rl_colors.HexColor(line_colors[L]))
            row_offset += 1

    col_widths = [40 * mm] + [7 * mm] * total_minutes + [10 * mm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(body_style)
    story.append(t)

    story.append(Spacer(1, 10))
    story.append(Paragraph("<i>● = op het veld</i>", styles["CoachSub"]))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def export_season_pdf(summary: dict, minutes_info: dict,
                      team_name: str) -> bytes:
    """Seizoensoverzicht-PDF voor evaluatiegesprek / ouderavond."""
    if not REPORTLAB_AVAILABLE:
        return b"PDF-export niet beschikbaar (reportlab ontbreekt)"
    buf = BytesIO()
    styles = _pdf_base_styles()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=18 * mm, rightMargin=18 * mm,
                            topMargin=16 * mm, bottomMargin=16 * mm)
    story = []

    story.append(Paragraph(f"Seizoensoverzicht — {_esc(team_name)}", styles["CoachH1"]))
    story.append(Paragraph(time.strftime("Gegenereerd op %d-%m-%Y"), styles["CoachSub"]))
    story.append(Spacer(1, 8))

    # KPI kaartjes in tabel
    played = summary.get("played", 0)
    win_pct = summary.get("win_pct", 0.0)
    diff = summary.get("goal_diff", 0)
    diff_str = f"+{diff}" if diff > 0 else str(diff)
    kpi_row = [[
        Paragraph(f"<b>Wedstrijden</b><br/><font size=20 color='#2563eb'>{played}</font>"
                  f"<br/><font size=8 color='#6b7280'>W {summary.get('wins', 0)} · "
                  f"G {summary.get('draws', 0)} · V {summary.get('losses', 0)}</font>",
                  styles["CoachBody"]),
        Paragraph(f"<b>Win %</b><br/><font size=20 color='#10b981'>{win_pct:.0f}%</font>",
                  styles["CoachBody"]),
        Paragraph(f"<b>Doelsaldo</b><br/><font size=20 color='#f59e0b'>{diff_str}</font>"
                  f"<br/><font size=8 color='#6b7280'>"
                  f"{summary.get('goals_for', 0)} voor · {summary.get('goals_against', 0)} tegen</font>",
                  styles["CoachBody"]),
        Paragraph(f"<b>Cirkelentries</b><br/><font size=20 color='#8b5cf6'>"
                  f"{summary.get('circle_entries_for', 0)}</font>"
                  f"<br/><font size=8 color='#6b7280'>{summary.get('shots_for', 0)} shots</font>",
                  styles["CoachBody"]),
    ]]
    kt = Table(kpi_row, colWidths=[45 * mm, 45 * mm, 45 * mm, 45 * mm])
    kt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _rl_colors.HexColor("#f9fafb")),
        ("BOX", (0, 0), (-1, -1), 0.4, _rl_colors.HexColor("#d1d5db")),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, _rl_colors.HexColor("#d1d5db")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(kt)
    story.append(Spacer(1, 14))

    # Wedstrijdlijst
    per_match = summary.get("per_match", [])
    if per_match:
        story.append(Paragraph("Wedstrijdlijst", styles["CoachH2"]))
        rows = [["Datum", "Tegenstander", "Res.", "Goals", "Cirkelentries"]]
        for p in per_match:
            rows.append([
                p.get("date", ""),
                p.get("opponent", ""),
                p.get("result", ""),
                f"{p.get('goals_for', 0)}-{p.get('goals_against', 0)}",
                str(p.get("circle_entries", 0)),
            ])
        mt = Table(rows, colWidths=[25 * mm, 60 * mm, 18 * mm, 25 * mm, 30 * mm], repeatRows=1)
        mt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _rl_colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), _rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.3, _rl_colors.HexColor("#d1d5db")),
            ("ALIGN", (2, 0), (-1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [_rl_colors.white, _rl_colors.HexColor("#f9fafb")]),
        ]))
        story.append(mt)
        story.append(Spacer(1, 14))

    # Speelminuten
    players_stats = (minutes_info or {}).get("players", {})
    if players_stats:
        story.append(Paragraph("Speelminuten per speler", styles["CoachH2"]))
        rows = [["Speler", "Minuten", "Wedstrijden"]]
        sorted_pl = sorted(players_stats.values(), key=lambda p: -p["minutes"])
        for p in sorted_pl:
            rows.append([p["name"], str(p["minutes"]), str(p["matches"])])
        mt = Table(rows, colWidths=[80 * mm, 35 * mm, 35 * mm], repeatRows=1)
        mt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _rl_colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), _rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.3, _rl_colors.HexColor("#d1d5db")),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [_rl_colors.white, _rl_colors.HexColor("#f9fafb")]),
        ]))
        story.append(mt)

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def style_excel_worksheet(ws) -> None:
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                value_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, value_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 35)


def add_section_title(ws, row: int, start_col: int, end_col: int, title: str, fill_color: str) -> None:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col)
    cell.value = title
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_header_row(ws, row: int, cols: list[int], fill_color: str) -> None:
    for col in cols:
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def format_kpi_sheet(ws) -> None:
    team_fill = "2563EB"
    opp_fill = "DC2626"
    neutral_fill = "0F172A"
    ws.insert_rows(1, 3)
    add_section_title(ws, 1, 1, 2, "Wedstrijd KPI-overzicht", neutral_fill)
    ws["A2"] = "Eigen team"
    ws["B2"] = "Tegenstander"
    style_header_row(ws, 2, [1], team_fill)
    style_header_row(ws, 2, [2], opp_fill)
    style_excel_worksheet(ws)


def format_per_quarter_sheet(ws) -> None:
    team_fill = "2563EB"
    opp_fill = "DC2626"
    neutral_fill = "0F172A"
    max_col = max(ws.max_column, 21)
    ws.insert_rows(1, 2)
    add_section_title(ws, 1, 1, max_col, "Statistieken per kwart", neutral_fill)
    add_section_title(ws, 2, 1, 13, "Eigen team", team_fill)
    add_section_title(ws, 2, 14, max_col, "Tegenstander", opp_fill)
    style_header_row(ws, 3, list(range(1, max_col + 1)), neutral_fill)
    style_excel_worksheet(ws)


def format_eventlog_sheet(ws, team_name: str, opponent_name: str) -> None:
    neutral_fill = "0F172A"
    team_fill = "DBEAFE"
    opp_fill = "FEE2E2"
    style_header_row(ws, 1, list(range(1, ws.max_column + 1)), neutral_fill)
    team_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if str(cell.value).lower() == "team":
            team_col = idx
            break
    if team_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=team_col)
            if cell.value == team_name:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=team_fill)
            elif cell.value == opponent_name:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=opp_fill)
    style_excel_worksheet(ws)


def export_excel(df: pd.DataFrame) -> bytes:
    buffer = BytesIO()
    eventlog_df = df.copy()
    kpi = build_kpi_summary(df) if not df.empty else {}
    quarter_df = build_quarter_stats_df(df) if not df.empty else pd.DataFrame()
    heatmap_df = build_entry_heatmap(df) if not df.empty else pd.DataFrame()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        eventlog_df.to_excel(writer, sheet_name="Eventlog", index=False)
        if not df.empty:
            kpi_export_df = pd.DataFrame([
                {"Eigen team": st.session_state.team_name, "Tegenstander": st.session_state.opponent_name},
                {"Eigen team": f"Cirkelentries: {kpi['team_entries']}", "Tegenstander": f"Cirkelentries: {kpi['opp_entries']}"},
                {"Eigen team": f"Schoten: {kpi['team_shots']}", "Tegenstander": f"Schoten: {kpi['opp_shots']}"},
                {"Eigen team": f"Schoten op goal: {kpi['team_shots_on_goal']}", "Tegenstander": f"Schoten op goal: {kpi['opp_shots_on_goal']}"},
                {"Eigen team": f"Goals: {kpi['team_goals']}", "Tegenstander": f"Goals: {kpi['opp_goals']}"},
                {"Eigen team": f"Entry->poging: {kpi['team_entry_to_shot_pct']:.1f}%", "Tegenstander": f"Entry->poging: {kpi['opp_entry_to_shot_pct']:.1f}%"},
                {"Eigen team": f"Op goal %: {kpi['team_on_goal_pct']:.1f}%", "Tegenstander": f"Op goal %: {kpi['opp_on_goal_pct']:.1f}%"},
                {"Eigen team": f"Shot on goal->goal: {kpi['team_shot_to_goal_pct']:.1f}%", "Tegenstander": f"Shot on goal->goal: {kpi['opp_shot_to_goal_pct']:.1f}%"},
                {"Eigen team": f"Press succes: {kpi['team_press_success']}", "Tegenstander": ""},
                {"Eigen team": f"Hoge balverovering: {kpi['team_high_wins']}", "Tegenstander": ""},
                {"Eigen team": f"Turnover eigen helft: {kpi['team_turnovers_own']}", "Tegenstander": ""},
                {"Eigen team": f"Counter tegen: {kpi['team_counters_against']}", "Tegenstander": ""},
            ])
            kpi_export_df.to_excel(writer, sheet_name="KPI", index=False)
            quarter_df.to_excel(writer, sheet_name="Per kwart", index=False)
            heatmap_df.to_excel(writer, sheet_name="Heatmap", index=False)
        workbook = writer.book
        format_eventlog_sheet(workbook["Eventlog"], st.session_state.team_name, st.session_state.opponent_name)
        if not df.empty:
            format_kpi_sheet(workbook["KPI"])
            format_per_quarter_sheet(workbook["Per kwart"])
            style_excel_worksheet(workbook["Heatmap"])
    buffer.seek(0)
    return buffer.getvalue()


def export_video_analysis_excel(clips_df: pd.DataFrame) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        clips_df.to_excel(writer, sheet_name="Beeldanalyse", index=False)
        if not clips_df.empty:
            summary_df = (
                clips_df.groupby("tag", dropna=False)
                .size()
                .reset_index(name="clips")
                .sort_values("clips", ascending=False)
            )
            summary_df.to_excel(writer, sheet_name="Samenvatting", index=False)
    buffer.seek(0)
    return buffer.getvalue()


# ==================================================
# SUPABASE
# ==================================================
@st.cache_resource(show_spinner=False)
def get_supabase_client():
    if create_client is None:
        return None
    try:
        # Flexibele lookup: probeer meerdere namen en secties.
        url, _ = _find_secret_value(["SUPABASE_URL", "supabase_url", "URL", "url"])
        key, _ = _find_secret_value([
            "SUPABASE_KEY", "supabase_key",
            "SUPABASE_ANON_KEY", "supabase_anon_key",
            "ANON_KEY", "anon_key", "KEY", "key",
        ])
        if not url or not key:
            return None
        return create_client(url, key)
    except Exception:
        return None


def cloud_enabled() -> bool:
    return get_supabase_client() is not None


def _find_secret_value(names: list[str]) -> tuple[str, str]:
    """Zoek in st.secrets naar de eerste match uit `names`, case-insensitive,
    ook onder [sections] zoals [supabase].

    Retourneert (gevonden_waarde, naam_zoals_in_secrets).
    """
    if not hasattr(st, "secrets"):
        return "", ""
    try:
        secrets = st.secrets
    except Exception:
        return "", ""

    target = {n.lower(): n for n in names}

    # 1) Top-level keys
    try:
        for k in secrets:
            if isinstance(k, str) and k.lower() in target:
                v = secrets.get(k, "")
                if v:
                    return str(v), k
    except Exception:
        pass

    # 2) Nested sections (bv. [supabase])
    try:
        for k in secrets:
            try:
                section = secrets[k]
            except Exception:
                continue
            if hasattr(section, "keys"):
                try:
                    for sk in section.keys():
                        if isinstance(sk, str) and sk.lower() in target:
                            v = section.get(sk, "")
                            if v:
                                return str(v), f"{k}.{sk}"
                except Exception:
                    continue
    except Exception:
        pass

    return "", ""


def _list_available_secret_keys() -> list[str]:
    """Geef een lijst van de top-level keys die Streamlit daadwerkelijk ziet."""
    if not hasattr(st, "secrets"):
        return []
    try:
        keys = []
        for k in st.secrets:
            if isinstance(k, str):
                keys.append(k)
        return keys
    except Exception:
        return []


def diagnose_cloud_connection() -> dict:
    """Diagnose de cloud-verbinding: waarom werkt het wel/niet?"""
    info = {
        "has_library": create_client is not None,
        "has_secrets": False,
        "url_key_used": "",
        "key_key_used": "",
        "url_short": "",
        "project_ref": "",
        "client_ok": False,
        "can_query": False,
        "teams_count": 0,
        "error": "",
        "source": "lokaal",
        "available_keys": _list_available_secret_keys(),
    }

    # Probeer SUPABASE_URL onder verschillende variaties
    url, url_key = _find_secret_value(["SUPABASE_URL", "supabase_url", "URL", "url"])
    key, key_key = _find_secret_value(["SUPABASE_KEY", "supabase_key", "SUPABASE_ANON_KEY",
                                       "supabase_anon_key", "ANON_KEY", "anon_key", "KEY", "key"])

    info["url_key_used"] = url_key
    info["key_key_used"] = key_key
    info["has_secrets"] = bool(url and key)
    if url:
        info["url_short"] = url[:60]
        m = re.match(r"https?://([a-z0-9]+)\.supabase\.co", url)
        if m:
            info["project_ref"] = m.group(1)

    if not info["has_library"]:
        info["error"] = "supabase-py library niet geïnstalleerd"
        return info
    if not info["has_secrets"]:
        missing = []
        if not url:
            missing.append("SUPABASE_URL")
        if not key:
            missing.append("SUPABASE_KEY")
        info["error"] = f"Niet gevonden in Streamlit secrets: {', '.join(missing)}"
        return info

    try:
        client = create_client(url, key)
        info["client_ok"] = client is not None
    except Exception as err:
        info["error"] = f"create_client faalt: {type(err).__name__}: {err}"
        return info

    if not info["client_ok"]:
        info["error"] = "create_client gaf None terug"
        return info

    try:
        response = client.table("teams").select("id", count="exact").execute()
        info["can_query"] = True
        info["source"] = "cloud"
        # Supabase-py retourneert count in .count als count='exact' gebruikt is
        count_val = getattr(response, "count", None)
        if count_val is None:
            count_val = len(response.data or [])
        info["teams_count"] = int(count_val or 0)
    except Exception as err:
        info["error"] = f"query teams-tabel faalt: {type(err).__name__}: {err}"
        return info

    return info


def render_login_diagnostics() -> None:
    """Laat diagnose-info zien op het inlogscherm — ingeklapt maar direct zichtbaar
    als er iets mis is."""
    info = diagnose_cloud_connection()
    all_ok = (
        info["has_library"]
        and info["has_secrets"]
        and info["client_ok"]
        and info["can_query"]
    )

    if all_ok:
        title = f"🟢 Cloud OK — {info['teams_count']} team(s) in Supabase (project `{info['project_ref'] or '?'}`)"
        expanded = False
    else:
        title = "🔴 Cloud probleem — klik open om te zien waarom teams niet verschijnen"
        expanded = True

    with st.expander(title, expanded=expanded):
        row = lambda label, val: st.markdown(
            f"<div style='display:flex; justify-content:space-between; "
            f"padding:4px 0; border-bottom:1px solid #2a3448;'>"
            f"<span style='color:#9ca3af'>{label}</span>"
            f"<span style='color:#f9fafb; font-family:monospace'>{val}</span></div>",
            unsafe_allow_html=True,
        )
        row("supabase-py geïnstalleerd", "✅ ja" if info["has_library"] else "❌ nee")
        row("Secrets aanwezig", "✅ ja" if info["has_secrets"] else "❌ nee")
        row("URL gevonden onder key", f"`{info['url_key_used']}`" if info["url_key_used"] else "—")
        row("KEY gevonden onder key", f"`{info['key_key_used']}`" if info["key_key_used"] else "—")
        row("URL (eerste 60 chars)", info["url_short"] or "—")
        row("Project-ref", info["project_ref"] or "—")
        row("Client aangemaakt", "✅ ja" if info["client_ok"] else "❌ nee")
        row("Query op 'teams'", "✅ ja" if info["can_query"] else "❌ nee")
        row("Aantal teams gevonden", str(info["teams_count"]))
        row("Bron van de teamlijst", info["source"])

        # Laat expliciet zien wat Streamlit wél in de secrets ziet
        avail = info.get("available_keys", [])
        if avail:
            st.markdown(
                f"<div style='margin-top:10px; padding:8px 12px; "
                f"background:#1f2937; border:1px solid #2a3448; border-radius:6px;'>"
                f"<div style='color:#9ca3af; font-size:12px; text-transform:uppercase; "
                f"letter-spacing:.05em;'>Keys die Streamlit ziet in secrets</div>"
                f"<div style='color:#f9fafb; font-family:monospace; margin-top:4px;'>"
                f"{', '.join(avail) if avail else '(leeg)'}</div></div>",
                unsafe_allow_html=True,
            )
        else:
            st.warning("Streamlit ziet **géén enkele** key in secrets — de secrets-file is leeg of niet opgeslagen.")

        if info["error"]:
            st.error(f"Laatste fout: **{info['error']}**")

        if not info["has_secrets"]:
            st.markdown(
                "**Oplossing:** Ga naar Streamlit Cloud → jouw app → `⋯` → **Settings** → "
                "**Secrets**, en voeg (exact deze namen, hoofdlettergevoelig!):\n\n"
                "```toml\nSUPABASE_URL = \"https://xxxx.supabase.co\"\n"
                "SUPABASE_KEY = \"eyJ...anon-key...\"\n```\n\n"
                "Klik daarna op **Save**, dan op **Reboot app** (soms moet je de app "
                "expliciet herstarten voordat nieuwe secrets worden gelezen)."
            )
        elif info["can_query"] and info["teams_count"] == 0:
            st.warning(
                "Cloud werkt, maar er staan **0 teams** in de `teams`-tabel van dit Supabase-project. "
                "Mogelijke oorzaken:\n"
                f"- Je secrets wijzen naar een ander project dan waar je teams hebt aangemaakt (huidige project: `{info['project_ref']}`).\n"
                "- De `teams`-tabel bestaat wel maar is leeg — maak rechts in de tab 'Nieuw team aanmaken' een team aan.\n"
                "- Row Level Security blokkeert het lezen — draai de `supabase_full_schema.sql` opnieuw."
            )


def _team_match_prefix() -> str:
    """Prefix voor match_id's van het huidige team, zodat data gescheiden blijft."""
    tid = st.session_state.get("active_team_id")
    if not tid:
        return ""
    return f"T-{tid[:8]}__"


def scope_match_id(match_id: str) -> str:
    """Zorg dat een match_id de huidige-team-prefix heeft."""
    prefix = _team_match_prefix()
    if not prefix:
        return match_id
    if match_id.startswith(prefix):
        return match_id
    return prefix + match_id


def unscope_match_id(match_id: str) -> str:
    """Toon-naam van een match_id zonder team-prefix."""
    prefix = _team_match_prefix()
    if prefix and match_id.startswith(prefix):
        return match_id[len(prefix):]
    return match_id


@st.cache_data(ttl=30, show_spinner=False)
def _fetch_match_ids(prefix: str, limit: int) -> list:
    """Gecachte query voor unieke match_id's."""
    client = get_supabase_client()
    if client is None:
        return []
    try:
        query = client.table("match_events").select("match_id,created_at").order(
            "created_at", desc=True
        ).limit(500)
        if prefix:
            query = query.ilike("match_id", f"{prefix}%")
        response = query.execute()
        rows = response.data or []
        seen = []
        seen_set = set()
        for r in rows:
            mid = r.get("match_id")
            if mid and mid not in seen_set:
                seen.append(mid)
                seen_set.add(mid)
            if len(seen) >= limit:
                break
        return seen
    except Exception:
        return []


def list_match_ids_from_cloud(limit: int = 50) -> list:
    """Haal unieke match_id's op uit de cloud, alleen voor het huidige team."""
    prefix = _team_match_prefix()
    result = _fetch_match_ids(prefix or "", limit)
    if result is not None:
        mark_cloud_ok()
    return result or []


def switch_to_match(match_id: str) -> None:
    """Wissel actieve wedstrijd en laad events uit cloud.

    Belangrijk: leidt ook de team-namen af uit de geladen events. Zonder dit
    blijft team_name op de default ("Ons team") staan en tellen goals niet mee
    in het scorebord / de analyse.
    """
    st.session_state.match_id = match_id
    events = load_events_from_cloud(match_id)
    st.session_state.events = events

    # Leid team-namen af uit de events: de 2 meest voorkomende teamnamen
    # worden eigen team / tegenstander. Eerste = eigen team (meeste events),
    # tweede = tegenstander. Zo werken scorebord en analyse direct.
    team_counts: dict[str, int] = {}
    for e in events:
        t = str(e.get("team", "")).strip()
        if t:
            team_counts[t] = team_counts.get(t, 0) + 1
    sorted_teams = sorted(team_counts.items(), key=lambda kv: kv[1], reverse=True)

    # Probeer eerst te bewaren welk team jouw eigen team is door de huidige
    # team_name te checken; als die nog in de events voorkomt, houd m.
    current_own = st.session_state.get("team_name", "")
    current_opp = st.session_state.get("opponent_name", "")

    names = [n for n, _ in sorted_teams]
    if current_own and current_own in names:
        new_own = current_own
        new_opp = next((n for n in names if n != new_own), current_opp or "Tegenstander")
    elif len(names) >= 2:
        new_own, new_opp = names[0], names[1]
    elif len(names) == 1:
        new_own = names[0]
        new_opp = current_opp or "Tegenstander"
    else:
        new_own = current_own or "Ons team"
        new_opp = current_opp or "Tegenstander"

    st.session_state.team_name = new_own
    st.session_state.opponent_name = new_opp
    # Widget-keys van de setup-balk opruimen zodat ze op de volgende rerun
    # opnieuw worden geïnitialiseerd vanuit team_name / opponent_name
    for wk in ("w_team_name", "w_opponent_name"):
        if wk in st.session_state:
            del st.session_state[wk]

    # Push-off-kalibraties uit cloud laden voor deze wedstrijd
    try:
        st.session_state.pushoff_offsets = cloud_load_pushoff_offsets(match_id)
    except Exception:
        st.session_state.pushoff_offsets = {q: None for q in QUARTERS}

    # Trigger wisselschema-herlaad voor deze wedstrijd op volgende _subs_init_state()
    st.session_state["subs_linked_match_id"] = None
    # Schema uit session wissen zodat de nieuwe wedstrijd leeg begint als er
    # nog geen lineup in de cloud staat
    st.session_state["subs_schema"] = None

    refresh_derived_state()


@st.cache_data(ttl=5, show_spinner=False)
def _fetch_events_from_cloud(match_id: str) -> list:
    """Gecachte eventlijst — TTL 5s zodat live wedstrijd snel ververst maar niet elke rerun."""
    client = get_supabase_client()
    if client is None:
        return []
    try:
        response = client.table("match_events").select("*").eq("match_id", match_id).order("created_at").execute()
        return [normalize_event_row(r) for r in (response.data or [])]
    except Exception:
        return []


def load_events_from_cloud(match_id: str) -> list:
    result = _fetch_events_from_cloud(match_id)
    if result is not None:
        mark_cloud_ok()
    return result or []


def save_event_to_cloud(event_row: dict) -> None:
    client = get_supabase_client()
    if client is None:
        return
    # Fouten worden bewust doorgegeven zodat de aanroeper ze kan loggen.
    safe_row = {k: v for k, v in event_row.items() if v is not None}
    client.table("match_events").insert(safe_row).execute()
    _fetch_events_from_cloud.clear()


def delete_last_event_cloud() -> None:
    client = get_supabase_client()
    if client is None or not st.session_state.events:
        return
    client.table("match_events").delete().eq("id", st.session_state.events[-1]["id"]).execute()
    _fetch_events_from_cloud.clear()


def update_event_player_cloud(event_id: str, player_id: str | None) -> None:
    """Update de player_id van een bestaand event in Supabase.

    Bij None: scorer wordt ontkoppeld.
    """
    client = get_supabase_client()
    if client is None or not event_id:
        return
    try:
        client.table("match_events").update({"player_id": player_id}).eq("id", event_id).execute()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("event speler bijwerken", err)


def set_event_player(event_id: str, player_id: str | None) -> None:
    """Koppel/ontkoppel scorer voor een event — lokaal + cloud."""
    for ev in st.session_state.get("events", []):
        if ev.get("id") == event_id:
            ev["player_id"] = player_id
            break
    update_event_player_cloud(event_id, player_id)
    refresh_derived_state()


def reset_match_cloud() -> None:
    client = get_supabase_client()
    if client is None:
        return
    client.table("match_events").delete().eq("match_id", st.session_state.match_id).execute()


def sync_from_cloud() -> None:
    if not cloud_enabled():
        return
    st.session_state.events = load_events_from_cloud(st.session_state.match_id)
    refresh_derived_state()
    st.session_state.last_sync_time = time.strftime("%H:%M:%S")


# ==================================================
# TEAM-SCOPED CLOUD: spelers, match-settings, aanwezigheid, schemas
# Alle functies doen een no-op als er geen active_team_id of geen Supabase-client is.
# ==================================================
def _active_team_id() -> str | None:
    return st.session_state.get("active_team_id")


@st.cache_data(ttl=60, show_spinner=False)
def _active_team_roster() -> list:
    """Geef het roster van het actieve team terug, bv. voor scorer-picker.

    Valt terug op session-state (subs_players) en laadt desnoods uit cloud.
    """
    roster = st.session_state.get("subs_players") or []
    if roster:
        return roster
    tid = _active_team_id()
    if not tid:
        return []
    # Lazy load: alleen als er nog niets in session staat
    try:
        cloud_roster = cloud_load_team_players()
    except Exception:
        cloud_roster = []
    if cloud_roster:
        st.session_state.subs_players = cloud_roster
    return cloud_roster


@st.cache_data(ttl=60, show_spinner=False)
def _fetch_team_players(team_id: str) -> list:
    """Gecachte spelerslijst per team."""
    client = get_supabase_client()
    if not team_id or client is None:
        return []
    try:
        response = client.table("team_players").select(
            "id,name,line,can_keep,priority"
        ).eq("team_id", team_id).order("name").execute()
        rows = response.data or []
        return [
            {
                "id": r["id"],
                "name": r.get("name", ""),
                "line": r.get("line", "M"),
                "can_keep": bool(r.get("can_keep", False)),
                "priority": r.get("priority", "normal"),
            }
            for r in rows
        ]
    except Exception:
        return []


def cloud_load_team_players() -> list:
    """Haal spelerslijst van actieve team op."""
    tid = _active_team_id()
    result = _fetch_team_players(tid or "")
    if result is not None:
        mark_cloud_ok()
    return result or []


def cloud_upsert_player(player: dict) -> None:
    """Voeg speler toe of update hem in de cloud."""
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return
    try:
        row = {
            "id": player["id"],
            "team_id": tid,
            "name": player.get("name", ""),
            "line": player.get("line", "M"),
            "can_keep": bool(player.get("can_keep", False)),
            "priority": player.get("priority", "normal"),
        }
        client.table("team_players").upsert(row).execute()
        _fetch_team_players.clear()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("speler opslaan", err)


def cloud_delete_player(player_id: str) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return
    try:
        client.table("team_players").delete().eq("id", player_id).eq("team_id", tid).execute()
        _fetch_team_players.clear()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("speler verwijderen", err)


def cloud_clear_players() -> None:
    """Verwijder alle spelers van het actieve team."""
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return
    try:
        client.table("team_players").delete().eq("team_id", tid).execute()
        client.table("team_attendance").delete().eq("team_id", tid).execute()
        _fetch_team_players.clear()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("team leegmaken", err)


def cloud_load_match_settings() -> dict | None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return None
    try:
        response = client.table("team_match_settings").select("*").eq("team_id", tid).limit(1).execute()
        rows = response.data or []
        mark_cloud_ok()
        if not rows:
            return None
        r = rows[0]
        return {
            "opponent": r.get("opponent", ""),
            "match_date": r.get("match_date", ""),
            "half_length": int(r.get("half_length", 17)),
            "halves": int(r.get("halves", 2)),
            "formation": r.get("formation", "4-3-3"),
            "fixed_keeper_id": r.get("fixed_keeper_id", "") or "",
            "block_size": int(r.get("block_size", 4)),
        }
    except Exception as err:
        log_cloud_error("match-settings laden", err)
        return None


def cloud_save_match_settings(settings: dict) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return
    try:
        row = {
            "team_id": tid,
            "opponent": settings.get("opponent", ""),
            "match_date": settings.get("match_date", ""),
            "half_length": int(settings.get("half_length", 17)),
            "halves": int(settings.get("halves", 2)),
            "formation": settings.get("formation", "4-3-3"),
            "fixed_keeper_id": settings.get("fixed_keeper_id", "") or None,
            "block_size": int(settings.get("block_size", 4)),
        }
        client.table("team_match_settings").upsert(row).execute()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("match-settings opslaan", err)


def cloud_load_attendance() -> dict:
    """Haal laatste bekende aanwezigheid op per player_id."""
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return {}
    try:
        response = client.table("team_attendance").select("player_id,present").eq("team_id", tid).execute()
        rows = response.data or []
        mark_cloud_ok()
        return {r["player_id"]: bool(r.get("present", True)) for r in rows if r.get("player_id")}
    except Exception as err:
        log_cloud_error("aanwezigheid laden", err)
        return {}


def cloud_save_attendance(player_id: str, present: bool) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return
    try:
        client.table("team_attendance").upsert({
            "team_id": tid,
            "player_id": player_id,
            "present": bool(present),
        }).execute()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("aanwezigheid opslaan", err)


def cloud_bulk_save_attendance(attendance: dict) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None or not attendance:
        return
    try:
        rows = [
            {"team_id": tid, "player_id": pid, "present": bool(val)}
            for pid, val in attendance.items()
        ]
        client.table("team_attendance").upsert(rows).execute()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("aanwezigheid bulk opslaan", err)


# --- Push-off kalibratie (per match_id + quarter) ---
def cloud_load_pushoff_offsets(match_id: str) -> dict:
    """Haal push-off-tijden op voor deze wedstrijd.
    Retourneert: {"Q1": 47.2, "Q2": None, ...}
    """
    client = get_supabase_client()
    if client is None or not match_id:
        return {q: None for q in QUARTERS}
    try:
        response = (
            client.table("match_pushoffs")
            .select("quarter,offset_sec")
            .eq("match_id", match_id)
            .execute()
        )
        rows = response.data or []
        mark_cloud_ok()
        out = {q: None for q in QUARTERS}
        for r in rows:
            q = r.get("quarter")
            if q in out:
                val = r.get("offset_sec")
                out[q] = float(val) if val is not None else None
        return out
    except Exception as err:
        log_cloud_error("push-off laden", err)
        return {q: None for q in QUARTERS}


def cloud_save_pushoff_offset(match_id: str, quarter: str, offset_sec: float | None) -> None:
    """Sla push-off-tijd op voor een specifiek kwart van een wedstrijd.
    Gebruikt upsert op (match_id, quarter) zodat herkalibratie gewoon overschrijft.
    """
    client = get_supabase_client()
    if client is None or not match_id or quarter not in QUARTERS:
        return
    try:
        client.table("match_pushoffs").upsert(
            {
                "match_id": match_id,
                "quarter": quarter,
                "offset_sec": None if offset_sec is None else float(offset_sec),
            },
            on_conflict="match_id,quarter",
        ).execute()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("push-off opslaan", err)


def cloud_save_schema(schema: dict, settings: dict) -> None:
    """Sla gegenereerd schema op met bijbehorende settings-snapshot."""
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None or not schema:
        return
    try:
        client.table("team_schemas").insert({
            "id": str(uuid.uuid4()),
            "team_id": tid,
            "opponent": settings.get("opponent", ""),
            "match_date": settings.get("match_date", time.strftime("%Y-%m-%d")),
            "schema_json": schema,
            "settings_json": settings,
        }).execute()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("schema opslaan", err)


def cloud_list_schemas(limit: int = 10) -> list:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return []
    try:
        response = client.table("team_schemas").select(
            "id,opponent,match_date,created_at,settings_json"
        ).eq("team_id", tid).order("created_at", desc=True).limit(limit).execute()
        mark_cloud_ok()
        return response.data or []
    except Exception as err:
        log_cloud_error("schema-lijst laden", err)
        return []


def cloud_load_schema(schema_id: str) -> dict | None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return None
    try:
        response = client.table("team_schemas").select("*").eq("id", schema_id).eq("team_id", tid).limit(1).execute()
        rows = response.data or []
        mark_cloud_ok()
        return rows[0] if rows else None
    except Exception as err:
        log_cloud_error("schema laden", err)
        return None


# --- Formatie-beheer per team ---
# Tabel team_formations: eigen formaties per team.
# Kolommen: id (uuid), team_id, name (text), slots_json (jsonb {V,M,A}),
#           is_default (bool), created_at (timestamptz).

def cloud_load_formations() -> list:
    """Haal alle formaties van het actieve team op."""
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return []
    try:
        response = client.table("team_formations").select(
            "id,name,slots_json,is_default"
        ).eq("team_id", tid).order("name").execute()
        rows = response.data or []
        mark_cloud_ok()
        out = []
        for r in rows:
            slots = r.get("slots_json") or {}
            out.append({
                "id": r["id"],
                "name": r.get("name", ""),
                "slots": {
                    "V": int(slots.get("V", 0)),
                    "M": int(slots.get("M", 0)),
                    "A": int(slots.get("A", 0)),
                },
                "is_default": bool(r.get("is_default", False)),
            })
        return out
    except Exception as err:
        log_cloud_error("formaties laden", err)
        return []


def cloud_upsert_formation(formation: dict) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return
    try:
        row = {
            "id": formation["id"],
            "team_id": tid,
            "name": (formation.get("name") or "").strip() or "Naamloos",
            "slots_json": {
                "V": int(formation.get("slots", {}).get("V", 0)),
                "M": int(formation.get("slots", {}).get("M", 0)),
                "A": int(formation.get("slots", {}).get("A", 0)),
            },
            "is_default": bool(formation.get("is_default", False)),
        }
        client.table("team_formations").upsert(row).execute()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("formatie opslaan", err)


def cloud_delete_formation(formation_id: str) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return
    try:
        client.table("team_formations").delete().eq("id", formation_id).eq("team_id", tid).execute()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("formatie verwijderen", err)


def cloud_set_default_formation(formation_id: str) -> None:
    """Zet één formatie als default; reset alle anderen."""
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return
    try:
        client.table("team_formations").update({"is_default": False}).eq("team_id", tid).execute()
        if formation_id:
            client.table("team_formations").update({"is_default": True}).eq(
                "id", formation_id
            ).eq("team_id", tid).execute()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("default-formatie zetten", err)


# --- Wisselschema per wedstrijd (gekoppeld aan match_id) ---
# Tabel match_lineups: match_id (text PK), team_id, schema_json, settings_json, updated_at.

def cloud_save_match_lineup(match_id: str, schema: dict, settings: dict) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None or not match_id or not schema:
        return
    try:
        client.table("match_lineups").upsert(
            {
                "match_id": match_id,
                "team_id": tid,
                "schema_json": schema,
                "settings_json": settings,
            },
            on_conflict="match_id",
        ).execute()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("wedstrijd-wisselschema opslaan", err)


def cloud_load_match_lineup(match_id: str) -> dict | None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None or not match_id:
        return None
    try:
        response = (
            client.table("match_lineups")
            .select("schema_json,settings_json")
            .eq("match_id", match_id)
            .eq("team_id", tid)
            .limit(1)
            .execute()
        )
        rows = response.data or []
        mark_cloud_ok()
        return rows[0] if rows else None
    except Exception as err:
        log_cloud_error("wedstrijd-wisselschema laden", err)
        return None


# ==================================================
# UI CALLBACKS / TIMER
# ==================================================
def sync_team_name_from_ui() -> None:
    # Val terug op "Ons team" als leeg, zodat KPI-tellingen nooit stuk gaan.
    # Let op: binnen on_change MAG je de widget-key nog wel lezen,
    # maar niet meer schrijven. Schrijven naar de master-key mag wel.
    val = (st.session_state.get("w_team_name") or "").strip()
    st.session_state.team_name = val if val else "Ons team"


def sync_opponent_name_from_ui() -> None:
    val = (st.session_state.get("w_opponent_name") or "").strip()
    st.session_state.opponent_name = val if val else "Tegenstander"


def sync_quarter_from_ui() -> None:
    st.session_state.quarter = st.session_state.get("w_quarter", "Q1")


def sync_match_id_from_ui() -> None:
    val = (st.session_state.get("w_match_id") or "").strip()
    if val:
        # Zorg dat handmatig ingevoerde match-id's bij het huidige team horen
        st.session_state.match_id = scope_match_id(val) if st.session_state.get("active_team_id") else val


def sync_device_mode_from_ui() -> None:
    st.session_state.device_mode = st.session_state.get("w_device_mode", "iPad")


def start_timer() -> None:
    if not st.session_state.timer_running:
        st.session_state.start_time = time.time()
        st.session_state.timer_running = True


def stop_timer() -> None:
    if st.session_state.timer_running:
        st.session_state.elapsed_before_run = current_elapsed_seconds()
        st.session_state.start_time = None
        st.session_state.timer_running = False


def reset_timer() -> None:
    st.session_state.timer_running = False
    st.session_state.start_time = None
    st.session_state.elapsed_before_run = 0


@st.fragment(run_every="1s" if st.session_state.timer_running else None)
def render_live_clock_bar() -> None:
    c1, c2, c3, c4 = st.columns([1.2, 1, 1, 1])
    c1.metric("Live klok", current_time_str())
    c2.button("Start", use_container_width=True, on_click=start_timer, key="clock_start_btn")
    c3.button("Stop", use_container_width=True, on_click=stop_timer, key="clock_stop_btn")
    c4.button("Reset klok", use_container_width=True, on_click=reset_timer, key="clock_reset_btn")


# ==================================================
# STYLING / HEADER
# ==================================================
def inject_custom_css() -> None:
    css = dedent(f"""
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:ital,wght@0,400;0,500;0,600;0,700;0,800;0,900;1,400&display=swap');

        /* ── Scrollbar ── */
        ::-webkit-scrollbar {{ width: 5px; height: 5px; }}
        ::-webkit-scrollbar-track {{ background: transparent; }}
        ::-webkit-scrollbar-thumb {{ background: {CARD_BORDER_SOFT}; border-radius: 10px; }}
        ::-webkit-scrollbar-thumb:hover {{ background: #2a3a5c; }}

        /* ── Achtergrond ── */
        .stApp {{
            background: {PAGE_BG_1};
            color: {TEXT_MAIN};
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
        }}
        .block-container {{
            padding-top: 1.2rem;
            padding-bottom: 5rem;
            max-width: 1400px;
        }}

        /* ── Typografie ── */
        h1, h2, h3, h4, h5, h6, p, span, label, div {{
            font-family: 'Inter', sans-serif;
        }}
        h1 {{ color: {TEXT_MAIN}; font-weight: 800; letter-spacing: -0.03em; font-size: 26px; }}
        h2 {{ color: {TEXT_MAIN}; font-weight: 700; letter-spacing: -0.02em; font-size: 21px; }}
        h3 {{ color: {TEXT_MAIN}; font-weight: 600; font-size: 17px; letter-spacing: -0.01em; }}
        h4, h5 {{ color: {TEXT_MAIN}; font-weight: 600; }}
        .stMarkdown p, .stMarkdown li {{ color: {TEXT_SUB}; line-height: 1.65; }}
        .stCaption, [data-testid="stCaptionContainer"] {{ color: {TEXT_MUTED} !important; font-size: 12px !important; }}

        /* ── Inputs ── */
        .stTextInput input,
        .stNumberInput input,
        .stTextArea textarea,
        .stSelectbox div[data-baseweb="select"] > div {{
            background: {CARD_BG_ELEVATED} !important;
            color: {TEXT_MAIN} !important;
            border: 1px solid {CARD_BORDER_SOFT} !important;
            border-radius: 10px !important;
            transition: border-color 0.2s ease, box-shadow 0.2s ease;
            font-size: 14px !important;
        }}
        .stTextInput input:focus,
        .stNumberInput input:focus,
        .stTextArea textarea:focus {{
            border-color: {ACCENT} !important;
            box-shadow: 0 0 0 3px {ACCENT_GLOW} !important;
            outline: none !important;
        }}
        .stTextInput label, .stNumberInput label, .stTextArea label,
        .stSelectbox label, .stRadio label, .stMultiSelect label,
        .stSlider label, .stFileUploader label, .stDateInput label {{
            color: {TEXT_SUB} !important;
            font-weight: 600 !important;
            font-size: 12px !important;
            text-transform: uppercase !important;
            letter-spacing: 0.05em !important;
        }}
        .stSelectbox div[data-baseweb="select"] > div {{ min-height: 42px; }}

        /* ── Knoppen ── */
        div.stButton > button {{
            border-radius: 10px;
            border: 1px solid {CARD_BORDER_SOFT};
            background: {CARD_BG_ELEVATED};
            color: {TEXT_MAIN};
            font-weight: 600;
            font-size: 13px;
            min-height: 42px;
            transition: all 0.18s ease;
            letter-spacing: 0.01em;
        }}
        div.stButton > button:hover {{
            background: #1a2540;
            border-color: {ACCENT};
            color: {TEXT_MAIN};
            transform: translateY(-1px);
            box-shadow: 0 4px 12px rgba(0,0,0,0.25);
        }}
        div.stButton > button[kind="primary"] {{
            background: linear-gradient(135deg, {ACCENT} 0%, #2563eb 100%);
            border-color: transparent;
            color: white;
            font-weight: 700;
            box-shadow: 0 2px 12px rgba(59,130,246,0.35);
        }}
        div.stButton > button[kind="primary"]:hover {{
            background: linear-gradient(135deg, {ACCENT_SOFT} 0%, {ACCENT} 100%);
            box-shadow: 0 6px 20px rgba(59,130,246,0.5);
            transform: translateY(-2px);
        }}
        div.stDownloadButton > button {{
            border-radius: 10px;
            border: 1px solid {CARD_BORDER_SOFT};
            background: {CARD_BG_ELEVATED};
            color: {TEXT_MAIN};
            font-weight: 600;
            min-height: 42px;
            transition: all 0.18s ease;
        }}
        div.stDownloadButton > button:hover {{
            border-color: {ACCENT};
            background: #1a2540;
            transform: translateY(-1px);
        }}

        /* ── Metric widgets ── */
        [data-testid="stMetricValue"] {{
            color: {TEXT_MAIN} !important;
            font-weight: 800 !important;
            font-size: 30px !important;
            letter-spacing: -0.03em;
            line-height: 1.1;
        }}
        [data-testid="stMetricLabel"] {{
            color: {TEXT_MUTED} !important;
            font-weight: 700 !important;
            font-size: 10px !important;
            text-transform: uppercase;
            letter-spacing: 0.1em;
        }}
        [data-testid="stMetricDelta"] {{ font-size: 12px !important; font-weight: 600 !important; }}
        [data-testid="stMetric"] {{
            background: {CARD_BG};
            border: 1px solid {CARD_BORDER};
            border-radius: 14px;
            padding: 18px 22px;
            transition: all 0.2s ease;
        }}
        [data-testid="stMetric"]:hover {{
            border-color: rgba(59,130,246,0.4);
            box-shadow: 0 6px 20px rgba(0,0,0,0.3);
            transform: translateY(-1px);
        }}

        /* ── Data-tabellen ── */
        .stDataFrame {{
            background: {CARD_BG};
            border-radius: 14px;
            border: 1px solid {CARD_BORDER};
            overflow: hidden;
            box-shadow: 0 2px 12px rgba(0,0,0,0.2);
        }}
        .stDataFrame [data-testid="stDataFrameResizable"] {{
            border-radius: 14px !important;
        }}

        /* ── Container met border ── */
        [data-testid="stVerticalBlockBorderWrapper"] {{
            background: {CARD_BG} !important;
            border: 1px solid {CARD_BORDER} !important;
            border-radius: 14px !important;
        }}

        /* ── Expanders ── */
        [data-testid="stExpander"] {{
            background: {CARD_BG} !important;
            border: 1px solid {CARD_BORDER} !important;
            border-radius: 14px !important;
            overflow: hidden;
            transition: border-color 0.2s ease;
        }}
        [data-testid="stExpander"]:hover {{ border-color: {CARD_BORDER_SOFT} !important; }}
        [data-testid="stExpander"] summary {{
            background: {CARD_BG} !important;
            padding: 16px 20px !important;
            font-weight: 600 !important;
            font-size: 14px !important;
            color: {TEXT_MAIN} !important;
        }}
        [data-testid="stExpander"] summary:hover {{ background: {CARD_BG_ELEVATED} !important; }}
        [data-testid="stExpander"] > div > div {{ padding: 0 20px 18px 20px !important; }}

        /* ── st.tabs() ── */
        [data-testid="stTabs"] [role="tablist"] {{
            gap: 2px;
            border-bottom: 1px solid {CARD_BORDER};
        }}
        [data-testid="stTabs"] [role="tab"] {{
            color: {TEXT_MUTED};
            font-weight: 600;
            font-size: 13px;
            padding: 10px 18px;
            border-radius: 8px 8px 0 0;
            border: none;
            background: transparent;
            transition: all 0.15s ease;
        }}
        [data-testid="stTabs"] [role="tab"]:hover {{
            color: {TEXT_SUB};
            background: rgba(59,130,246,0.06);
        }}
        [data-testid="stTabs"] [role="tab"][aria-selected="true"] {{
            color: {ACCENT_SOFT};
            font-weight: 700;
            background: rgba(59,130,246,0.1);
        }}
        [data-testid="stTabs"] [data-baseweb="tab-highlight"] {{
            background-color: {ACCENT} !important;
            height: 2px !important;
            border-radius: 2px 2px 0 0;
        }}

        /* ── Kaart (algemeen) ── */
        .safe-card {{
            background: {CARD_BG};
            border: 1px solid {CARD_BORDER};
            border-radius: 14px;
            padding: 20px 22px;
            height: 100%;
            min-height: 130px;
            transition: all 0.2s ease;
        }}
        .safe-card:hover {{
            border-color: rgba(59,130,246,0.4);
            box-shadow: 0 8px 24px rgba(0,0,0,0.3);
            transform: translateY(-2px);
        }}
        .safe-card-title {{
            font-size: 10px; color: {TEXT_MUTED}; font-weight: 700;
            text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 10px;
        }}
        .safe-card-value {{
            font-size: 30px; font-weight: 800; color: {TEXT_MAIN};
            line-height: 1.1; margin-bottom: 6px; letter-spacing: -0.03em;
        }}
        .safe-card-sub {{ color: {TEXT_SUB}; font-size: 13px; line-height: 1.5; }}
        .accent-blue {{ border-left: 3px solid {TEAM_BLUE}; }}
        .accent-red {{ border-left: 3px solid {OPP_RED}; }}
        .accent-green {{ border-left: 3px solid {SUCCESS_GREEN}; }}
        .accent-orange {{ border-left: 3px solid {WARNING_ORANGE}; }}

        /* ── Tool-kaarten op homescherm ── */
        .tool-card {{
            background: {CARD_BG};
            border: 1px solid {CARD_BORDER};
            border-radius: 16px;
            padding: 20px 20px 14px 20px;
            min-height: 200px;
            transition: all 0.22s cubic-bezier(0.4, 0, 0.2, 1);
            display: flex;
            flex-direction: column;
            gap: 6px;
            position: relative;
            overflow: hidden;
            cursor: pointer;
        }}
        .tool-card::before {{
            content: '';
            position: absolute;
            top: 0; left: 0; right: 0;
            height: 2px;
            background: linear-gradient(90deg, {ACCENT}, {ACCENT_SOFT});
            opacity: 0;
            transition: opacity 0.22s ease;
        }}
        .tool-card::after {{
            content: '';
            position: absolute;
            inset: 0;
            background: radial-gradient(ellipse at top left, rgba(59,130,246,0.06) 0%, transparent 60%);
            opacity: 0;
            transition: opacity 0.22s ease;
        }}
        .tool-card:hover {{
            border-color: rgba(59,130,246,0.5);
            background: {CARD_BG_ELEVATED};
            box-shadow: 0 16px 48px rgba(0,0,0,0.4), 0 0 0 1px rgba(59,130,246,0.1);
            transform: translateY(-4px);
        }}
        .tool-card:hover::before {{ opacity: 1; }}
        .tool-card:hover::after {{ opacity: 1; }}
        .tool-card-icon-wrap {{
            width: 48px; height: 48px; border-radius: 13px;
            background: linear-gradient(135deg, rgba(59,130,246,0.18) 0%, rgba(59,130,246,0.06) 100%);
            border: 1px solid rgba(59,130,246,0.22);
            display: flex; align-items: center; justify-content: center;
            font-size: 24px; margin-bottom: 8px;
            box-shadow: 0 2px 8px rgba(59,130,246,0.15);
            transition: box-shadow 0.22s ease;
        }}
        .tool-card:hover .tool-card-icon-wrap {{
            box-shadow: 0 4px 16px rgba(59,130,246,0.3);
        }}
        .tool-card-title {{
            font-size: 15px; font-weight: 700; color: {TEXT_MAIN};
            letter-spacing: -0.02em; line-height: 1.2;
        }}
        .tool-card-desc {{
            color: {TEXT_SUB}; font-size: 12.5px; line-height: 1.55; flex: 1;
        }}
        .tool-card-tabs {{
            color: {TEXT_MUTED}; font-size: 10px; font-weight: 600;
            letter-spacing: 0.05em; text-transform: uppercase;
            padding-top: 10px; border-top: 1px solid {CARD_BORDER};
        }}
        .tool-card-new {{
            position: absolute; top: 12px; right: 12px;
            background: linear-gradient(135deg, rgba(16,185,129,0.2), rgba(16,185,129,0.08));
            border: 1px solid rgba(16,185,129,0.3);
            color: #34d399; font-size: 9px; font-weight: 700;
            letter-spacing: 0.08em; text-transform: uppercase;
            padding: 2px 7px; border-radius: 6px;
        }}

        /* ── Mini event-feed ── */
        .mini-feed {{
            background: {CARD_BG};
            border: 1px solid {CARD_BORDER};
            border-radius: 10px;
            padding: 10px 14px;
            margin-bottom: 6px;
            color: {TEXT_MAIN};
            font-size: 14px;
            transition: border-color 0.15s ease;
        }}
        .mini-feed:hover {{ border-color: {CARD_BORDER_SOFT}; }}
        .mini-feed strong {{ color: {TEXT_MAIN}; font-weight: 600; }}

        /* ── Pillen / badges ── */
        .pill {{
            display: inline-block; padding: 3px 9px; border-radius: 6px;
            font-size: 11px; font-weight: 600; margin-right: 6px;
            letter-spacing: 0.02em;
        }}
        .pill-blue {{ background: rgba(59,130,246,0.15); color: #93c5fd; }}
        .pill-red {{ background: rgba(244,63,94,0.15); color: #fda4af; }}
        .pill-green {{ background: rgba(16,185,129,0.15); color: #6ee7b7; }}
        .pill-gray {{ background: rgba(156,163,175,0.12); color: {TEXT_SUB}; }}

        /* ── Hero navbar ── */
        .hero {{
            background: linear-gradient(135deg, {CARD_BG} 0%, #0b1528 100%);
            border: 1px solid {CARD_BORDER};
            border-radius: 16px;
            padding: 14px 24px;
            margin-bottom: 8px;
            box-shadow: 0 4px 24px rgba(0,0,0,0.3), inset 0 1px 0 rgba(59,130,246,0.08);
            position: relative;
            overflow: hidden;
        }}
        .hero::after {{
            content: '';
            position: absolute;
            top: 0; left: 0; right: 0;
            height: 1px;
            background: linear-gradient(90deg, transparent, rgba(59,130,246,0.5), transparent);
        }}
        .hero-top {{
            display: flex; justify-content: space-between;
            align-items: center; gap: 16px; flex-wrap: wrap;
        }}
        .hero-brand {{ display: flex; align-items: center; gap: 14px; }}
        .hero-logo {{
            width: 40px; height: 40px; border-radius: 11px;
            background: linear-gradient(135deg, {ACCENT} 0%, #1d4ed8 100%);
            display: flex; align-items: center; justify-content: center;
            color: white; font-weight: 900; font-size: 15px;
            box-shadow: 0 4px 16px rgba(59,130,246,0.4);
            flex-shrink: 0; letter-spacing: -0.03em;
        }}
        .hero-title {{
            font-size: 17px; font-weight: 800; color: {TEXT_MAIN};
            letter-spacing: -0.03em; line-height: 1.2;
        }}
        .hero-sub {{ color: {TEXT_MUTED}; font-size: 12px; margin-top: 2px; font-weight: 500; }}
        .status-chip {{
            display: inline-flex; align-items: center; gap: 5px;
            background: rgba(255,255,255,0.04);
            border: 1px solid {CARD_BORDER_SOFT};
            border-radius: 20px; padding: 5px 12px;
            font-size: 11px; font-weight: 600; color: {TEXT_MUTED};
            white-space: nowrap; letter-spacing: 0.02em;
        }}
        .status-chip.ok {{
            color: #34d399; border-color: rgba(16,185,129,0.3);
            background: rgba(16,185,129,0.06);
        }}
        .status-chip.live {{
            color: #fb7185; border-color: rgba(244,63,94,0.35);
            background: rgba(244,63,94,0.08);
        }}
        .status-chip.live .dot {{
            width: 6px; height: 6px; border-radius: 50%;
            background: {OPP_RED}; box-shadow: 0 0 8px {OPP_RED};
            animation: pulse 1.8s infinite;
        }}
        @keyframes pulse {{
            0%, 100% {{ opacity: 1; transform: scale(1); }}
            50% {{ opacity: 0.4; transform: scale(0.85); }}
        }}

        /* ── Topbar ── */
        .cs-topbar {{
            display: flex; align-items: center;
            justify-content: space-between;
            padding: 6px 4px 12px 4px; gap: 8px;
        }}
        .cs-topbar-team {{ font-size: 13px; color: {TEXT_SUB}; font-weight: 500; }}
        .cs-topbar-team strong {{ color: {TEXT_MAIN}; font-weight: 600; }}

        /* ── Score-bar ── */
        .scorebar {{
            background: {CARD_BG}; border: 1px solid {CARD_BORDER};
            border-radius: 14px; padding: 18px 24px;
            display: flex; align-items: center;
            justify-content: space-between; margin-bottom: 16px;
        }}
        .scorebar-team {{ flex: 1; text-align: center; }}
        .scorebar-team-name {{
            font-size: 13px; font-weight: 600; color: {TEXT_SUB};
            text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 4px;
        }}
        .scorebar-team-tag {{ font-size: 11px; color: {TEXT_MUTED}; }}
        .scorebar-score {{
            font-size: 48px; font-weight: 800; color: {TEXT_MAIN};
            line-height: 1; letter-spacing: -0.03em; padding: 0 32px;
            display: flex; align-items: center; gap: 18px;
        }}
        .scorebar-score-dash {{ color: {TEXT_MUTED}; font-weight: 300; }}
        .scorebar-score-team-own {{ color: {TEAM_BLUE}; }}
        .scorebar-score-team-opp {{ color: {OPP_RED}; }}

        /* ── Team header ── */
        .team-header {{
            padding: 12px 18px; border-radius: 10px;
            font-weight: 600; font-size: 15px; color: white;
            margin-bottom: 12px; text-align: center; letter-spacing: -0.01em;
        }}

        /* ── Navigatie-tab balk ── */
        .cs-nav-wrap {{
            background: {CARD_BG}; border: 1px solid {CARD_BORDER};
            border-radius: 14px; padding: 5px; margin-bottom: 18px;
            display: flex; gap: 4px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.2);
        }}

        /* ── Inlogscherm ── */
        .login-shell {{
            max-width: 460px; margin: 56px auto 0 auto;
            background: linear-gradient(160deg, {CARD_BG} 0%, #0b1220 100%);
            border: 1px solid {CARD_BORDER_SOFT}; border-radius: 20px;
            padding: 48px 40px;
            box-shadow: 0 24px 60px rgba(0,0,0,0.5), 0 0 0 1px rgba(59,130,246,0.06);
        }}
        .login-logo {{
            width: 64px; height: 64px; border-radius: 18px;
            background: linear-gradient(135deg, {ACCENT} 0%, #1d4ed8 100%);
            display: flex; align-items: center; justify-content: center;
            color: white; font-weight: 900; font-size: 26px;
            margin: 0 auto 24px auto;
            box-shadow: 0 10px 30px rgba(59,130,246,0.45);
            letter-spacing: -0.04em;
        }}
        .login-title {{
            text-align: center; font-size: 26px; font-weight: 800;
            color: {TEXT_MAIN}; letter-spacing: -0.03em; margin-bottom: 6px;
        }}
        .login-sub {{
            text-align: center; color: {TEXT_MUTED}; font-size: 14px;
            margin-bottom: 32px; line-height: 1.55;
        }}

        /* ── Back-to-home knop ── */
        [data-testid="stButton"] button[kind="secondary"][data-testid*="btn_back_home_hero"] {{
            background: transparent !important;
            border: 1px solid {CARD_BORDER_SOFT} !important;
            color: {TEXT_MUTED} !important;
            font-size: 12px !important;
            min-height: 32px !important;
            padding: 4px 12px !important;
            border-radius: 8px !important;
            margin-bottom: 6px;
        }}
        [data-testid="stButton"] button[kind="secondary"][data-testid*="btn_back_home_hero"]:hover {{
            border-color: {ACCENT} !important;
            color: {ACCENT_SOFT} !important;
            background: rgba(59,130,246,0.06) !important;
        }}

        /* ── Verberg Streamlit-menu & footer ── */
        #MainMenu, footer, [data-testid="stToolbar"] {{ visibility: hidden; }}
        [data-testid="stDecoration"] {{ display: none; }}

        /* ── Radio ── */
        [data-baseweb="radio"] {{ color: {TEXT_MAIN}; }}

        /* ── Alerts ── */
        [data-testid="stAlert"] {{
            border-radius: 12px; border-width: 1px;
            font-size: 13px; font-weight: 500;
        }}

        /* ── Divider ── */
        hr {{ border-color: {CARD_BORDER} !important; margin: 20px 0 !important; opacity: 0.7; }}

        /* ── Form containers ── */
        [data-testid="stForm"] {{
            background: {CARD_BG};
            border: 1px solid {CARD_BORDER} !important;
            border-radius: 14px !important;
            padding: 4px !important;
        }}

        /* ── Welcome sectie op homescherm ── */
        .cs-welcome {{ padding: 8px 2px 20px 2px; }}
        .cs-welcome-title {{
            font-size: 26px; font-weight: 800; color: {TEXT_MAIN};
            letter-spacing: -0.03em; line-height: 1.2;
        }}
        .cs-welcome-title span {{
            background: linear-gradient(90deg, {ACCENT_SOFT}, {ACCENT});
            -webkit-background-clip: text; -webkit-text-fill-color: transparent;
            background-clip: text;
        }}
        .cs-welcome-sub {{
            color: {TEXT_SUB}; font-size: 14px; margin-top: 6px;
        }}
        .cs-section-label {{
            font-size: 10px; font-weight: 800; text-transform: uppercase;
            letter-spacing: 0.12em; color: {TEXT_MUTED}; margin: 24px 0 12px 2px;
        }}

        </style>
    """).strip()
    st.markdown(css, unsafe_allow_html=True)


def render_info_card(title: str, value: str, subtitle: str, accent: str) -> None:
    accent_class = {"blue": "accent-blue", "red": "accent-red", "green": "accent-green", "orange": "accent-orange"}.get(accent, "accent-blue")
    html = (
        f'<div class="safe-card {accent_class}">'
        f'<div class="safe-card-title">{title}</div>'
        f'<div class="safe-card-value">{value}</div>'
        f'<div class="safe-card-sub">{subtitle}</div>'
        f'</div>'
    )
    st.markdown(html, unsafe_allow_html=True)


def render_hero_header() -> None:
    active_tool = st.session_state.get("active_tool")
    team_name = st.session_state.get("active_team_name") or "—"
    _tool_labels = {
        "MATCH_ANALYSIS": "Wedstrijd analyse",
        "VIDEO_ANALYSIS": "Video analyse",
        "SUBSTITUTION": "Wisselschema",
        "SEASON": "Seizoensoverzicht",
        "PLAYER_PROFILE": "Spelersprofiel",
        "MATCH_MGMT": "Wedstrijden & uitslagen",
        "SELECTION": "Selectietool",
        "TRAINING": "Trainingsplanning",
        "INJURIES": "Blessure tracker",
        "SCOUTING": "Tegenstander scouting",
        "GOALS": "Seizoensdoelen",
    }
    if active_tool:
        subtitle = _tool_labels.get(active_tool, active_tool)
    else:
        subtitle = "Kies een tool om te beginnen"

    is_live = st.session_state.timer_running
    cloud_ok = cloud_enabled()
    chips = []
    if active_tool == "MATCH_ANALYSIS":
        # Toon kwart en live-indicator alleen in de wedstrijdtool
        chips.append(f'<span class="status-chip">{st.session_state.quarter}</span>')
        if is_live:
            chips.append('<span class="status-chip live"><span class="dot"></span>LIVE</span>')
    if cloud_ok:
        chips.append('<span class="status-chip ok">☁ Cloud</span>')
    else:
        chips.append(f'<span class="status-chip">Lokaal</span>')
    chips_html = "".join(chips)

    html = (
        f'<div class="hero">'
        f'<div class="hero-top">'
        f'<div class="hero-brand">'
        f'<div class="hero-logo">CS</div>'
        f'<div>'
        f'<div class="hero-title">Coach Studio</div>'
        f'<div class="hero-sub">{team_name} · {subtitle}</div>'
        f'</div>'
        f'</div>'
        f'<div style="display:flex;gap:6px;flex-wrap:wrap;align-items:center;">'
        f'{chips_html}'
        f'</div>'
        f'</div>'
        f'</div>'
    )
    st.markdown(html, unsafe_allow_html=True)

    # ── Terug-naar-home knop ──
    if active_tool:
        col_back, _ = st.columns([1, 7])
        with col_back:
            if st.button("← Home", key="btn_back_home_hero", use_container_width=True):
                st.session_state.active_tool = None
                st.session_state.active_screen = None
                st.rerun()


def render_match_scorebar() -> None:
    # Alles op één regel zodat Streamlit/markdown geen inspringen als codeblok interpreteert
    html = (
        f'<div class="scorebar">'
        f'<div class="scorebar-team">'
        f'<div class="scorebar-team-name">{st.session_state.team_name}</div>'
        f'<div class="scorebar-team-tag">Eigen team</div>'
        f'</div>'
        f'<div class="scorebar-score">'
        f'<span class="scorebar-score-team-own">{st.session_state.score_team}</span>'
        f'<span class="scorebar-score-dash">–</span>'
        f'<span class="scorebar-score-team-opp">{st.session_state.score_opponent}</span>'
        f'</div>'
        f'<div class="scorebar-team">'
        f'<div class="scorebar-team-name">{st.session_state.opponent_name}</div>'
        f'<div class="scorebar-team-tag">Tegenstander</div>'
        f'</div>'
        f'</div>'
    )
    st.markdown(html, unsafe_allow_html=True)


def render_navigation() -> None:
    # De zichtbare tabs hangen af van welke tool actief is.
    active_tool = st.session_state.get("active_tool", "MATCH_ANALYSIS")
    if active_tool == "MATCH_ANALYSIS":
        screens = [
            ("LIVE", "Live", "●"),
            ("ANALYSE", "Analyse", "▣"),
            ("VELD", "Veld", "◉"),
            ("RAPPORT", "Rapport", "≡"),
        ]
    elif active_tool == "VIDEO_ANALYSIS":
        screens = [("BEELDANALYSE", "Video", "▶")]
    elif active_tool == "SUBSTITUTION":
        screens = [("WISSEL", "Wisselschema", "⇄")]
    elif active_tool == "SEASON":
        screens = [("SEIZOEN", "Seizoensoverzicht", "📊")]
    elif active_tool == "PLAYER_PROFILE":
        screens = [("PROFIEL", "Spelersprofiel", "👤")]
    elif active_tool == "MATCH_MGMT":
        screens = [("WEDSTRIJDEN", "Wedstrijden", "🏑")]
    elif active_tool == "SELECTION":
        screens = [("SELECTIE", "Selectietool", "👥")]
    elif active_tool == "TRAINING":
        screens = [("TRAINING", "Trainingsplanning", "📅")]
    elif active_tool == "INJURIES":
        screens = [("BLESSURES", "Blessure tracker", "🩹")]
    elif active_tool == "SCOUTING":
        screens = [("SCOUTING", "Scouting", "🔍")]
    elif active_tool == "GOALS":
        screens = [("GOALS", "Seizoensdoelen", "🎯")]
    else:
        screens = []

    # Zorg dat active_screen bij deze tool hoort
    allowed = {s[0] for s in screens}
    if screens and st.session_state.get("active_screen") not in allowed:
        st.session_state.active_screen = screens[0][0]

    if len(screens) <= 1:
        # Eén-tab tools: toon een subtiele tool-label balk
        if screens:
            _, lbl, icon = screens[0]
            st.markdown(
                f'<div style="padding:4px 0 12px 2px;color:{TEXT_MUTED};'
                f'font-size:13px;font-weight:500;">{icon} {lbl}</div>',
                unsafe_allow_html=True,
            )
        return

    # Navigatie als pill-tabs in een achtergrond-container
    st.markdown('<div class="cs-nav-wrap">', unsafe_allow_html=True)
    cols = st.columns(len(screens))
    for i, (screen_id, label, icon) in enumerate(screens):
        is_active = st.session_state.active_screen == screen_id
        if cols[i].button(
            f"{icon}  {label}",
            use_container_width=True,
            type="primary" if is_active else "secondary",
            key=f"nav_{screen_id}",
        ):
            st.session_state.active_screen = screen_id
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)


def render_setup_bar() -> None:
    top1, top2, top3, top4, top5 = st.columns([1.05, 1.05, 0.65, 0.95, 0.9])
    # Gebruik losse widget-keys (w_*) en lees hun initiële waarde uit de master-state.
    # Zo mogen we match_id / team_name etc. elders vrij overschrijven zonder
    # dat Streamlit crasht.
    with top1:
        st.text_input("Naam eigen team", value=st.session_state.team_name,
                      key="w_team_name", on_change=sync_team_name_from_ui)
    with top2:
        st.text_input("Naam tegenstander", value=st.session_state.opponent_name,
                      key="w_opponent_name", on_change=sync_opponent_name_from_ui)
    with top3:
        q_idx = QUARTERS.index(st.session_state.quarter) if st.session_state.quarter in QUARTERS else 0
        st.selectbox("Kwart", QUARTERS, index=q_idx,
                     key="w_quarter", on_change=sync_quarter_from_ui)
    with top4:
        st.text_input("Wedstrijd-ID", value=st.session_state.match_id,
                      key="w_match_id", on_change=sync_match_id_from_ui)
    with top5:
        modes = ["Cockpit", "MacBook", "iPad", "iPhone"]
        m_idx = modes.index(st.session_state.device_mode) if st.session_state.device_mode in modes else 2
        st.selectbox("Versie", modes, index=m_idx,
                     key="w_device_mode", on_change=sync_device_mode_from_ui)
    # Alleen Sync blijft hier — wedstrijd openen en nieuwe wedstrijd staan nu op het tool-overzicht.
    sb1, sb2 = st.columns([3, 1])
    with sb1:
        st.caption(
            "💡 Wil je een andere wedstrijd openen of een nieuwe starten? "
            "Ga terug naar het tool-overzicht (logo linksboven) en kies daar."
        )
    with sb2:
        st.button(
            "🔄  Sync",
            use_container_width=True,
            on_click=sync_from_cloud,
            key="setup_sync_btn",
            help="Haal de laatste events uit de cloud",
        )
    # Wedstrijd-kiezer paneel (opent na klik)
    if st.session_state.get("show_match_picker"):
        with st.container(border=True):
            st.markdown("#### Kies een eerder opgeslagen wedstrijd")
            match_ids = list_match_ids_from_cloud(limit=30)
            if not match_ids:
                st.info("Geen eerdere wedstrijden gevonden in de cloud.")
            else:
                picked = st.selectbox("Wedstrijd", match_ids, key="picker_match_id")
                pc1, pc2 = st.columns([1, 1])
                if pc1.button("Open deze wedstrijd", use_container_width=True,
                              type="primary", key="picker_open_btn"):
                    switch_to_match(picked)
                    st.session_state.show_match_picker = False
                    st.success(f"Geladen: {picked} ({len(st.session_state.events)} events).")
                    st.rerun()
                if pc2.button("Annuleer", use_container_width=True, key="picker_cancel_btn"):
                    st.session_state.show_match_picker = False
                    st.rerun()
    # Nieuwe-wedstrijd formulier (opent na klik)
    if st.session_state.get("show_new_match_form"):
        with st.container(border=True):
            st.markdown("#### Nieuwe wedstrijd starten")
            nm1, nm2 = st.columns(2)
            with nm1:
                nm_team = st.text_input("Eigen team", value=st.session_state.team_name or "Ons team",
                                        key="nm_team_input")
                nm_opp = st.text_input("Tegenstander", value="", placeholder="Bijv. Kampong D1",
                                       key="nm_opp_input")
            with nm2:
                nm_date = st.text_input("Datum", value=time.strftime("%Y-%m-%d"), key="nm_date_input")
                nm_label = st.text_input("Korte omschrijving (optioneel)", value="",
                                         placeholder="Bijv. thuis / competitie",
                                         key="nm_label_input")
            sc1, sc2 = st.columns([1, 1])
            if sc1.button("Start wedstrijd", use_container_width=True, type="primary",
                          key="nm_start_btn"):
                # Bouw een leesbare match_id
                clean = lambda s: re.sub(r"\W+", "-", s.strip()).strip("-") or "match"
                parts = [clean(nm_team), clean(nm_opp) if nm_opp else "vs", clean(nm_date)]
                if nm_label:
                    parts.append(clean(nm_label))
                new_id = "-".join(parts)
                # Alleen master-keys zetten — widgets lezen die op de volgende rerun
                st.session_state.match_id = new_id
                st.session_state.team_name = nm_team or "Ons team"
                if nm_opp:
                    st.session_state.opponent_name = nm_opp
                st.session_state.events = []
                st.session_state.video_clips = []
                st.session_state.pushoff_offsets = {"Q1": None, "Q2": None, "Q3": None, "Q4": None}
                st.session_state.quarter = "Q1"
                st.session_state.score_team = 0
                st.session_state.score_opponent = 0
                st.session_state.auto_notes = ""
                st.session_state.halftime_report = ""
                st.session_state.show_new_match_form = False
                st.success(f"Nieuwe wedstrijd: {new_id}")
                st.rerun()
            if sc2.button("Annuleer", use_container_width=True, key="nm_cancel_btn"):
                st.session_state.show_new_match_form = False
                st.rerun()

    render_live_clock_bar()


def start_new_match_flow() -> None:
    """Open het nieuwe-wedstrijd-formulier."""
    st.session_state.show_new_match_form = True


# ==================================================
# SMART TAG PANEL + LIVE SCREEN
# ==================================================
def render_scorer_picker() -> None:
    """Toon een inline scorer-picker voor het laatst gemaakte eigen-team goal.

    Verschijnt zodra _maybe_mark_goal_scorer_pending last_goal_event_id heeft gezet.
    Gebruiker kan kiezen uit roster, 'onbekend' kiezen (overslaan), of later
    aanpassen in het event-log.
    """
    pending_id = st.session_state.get("last_goal_event_id")
    if not pending_id:
        return
    # Zoek het bijbehorende event
    ev = next((e for e in st.session_state.get("events", []) if e.get("id") == pending_id), None)
    if ev is None:
        st.session_state["last_goal_event_id"] = None
        return
    roster = _active_team_roster()
    # Sorteer: A-lijn eerst (aanvallers scoren vaakst), dan M, dan V, dan K
    line_order = {"A": 0, "M": 1, "V": 2, "K": 3}
    roster_sorted = sorted(roster, key=lambda p: (line_order.get(p.get("line", "M"), 9), p.get("name", "")))

    st.markdown("---")
    st.markdown(
        f"<div style='background:#ecfdf5; border:1px solid #10b981; border-radius:10px; "
        f"padding:10px 14px;'>"
        f"<b>⚽ Wie heeft gescoord?</b><br>"
        f"<span style='color:#047857;'>Goal in {ev.get('quarter', '?')} · {ev.get('time', '?')}</span>"
        f"</div>",
        unsafe_allow_html=True,
    )
    if not roster_sorted:
        st.caption("Geen spelers in roster — voeg eerst spelers toe in de Wisselschema-tool.")
        if st.button("Overslaan", key=f"scorer_skip_empty_{pending_id}", use_container_width=True):
            st.session_state["last_goal_event_id"] = None
            st.rerun()
        return

    # Grid van 4 knoppen per rij
    cols_per_row = 4
    for i in range(0, len(roster_sorted), cols_per_row):
        row = roster_sorted[i : i + cols_per_row]
        cols = st.columns(cols_per_row)
        for j, p in enumerate(row):
            label = f"{p['name']} ({p.get('line', '?')})"
            if cols[j].button(label, key=f"scorer_pick_{pending_id}_{p['id']}", use_container_width=True):
                set_event_player(pending_id, p["id"])
                st.session_state["last_goal_event_id"] = None
                st.session_state.last_undo_msg = f"Scorer: {p['name']}"
                st.rerun()
    skip_col1, skip_col2 = st.columns(2)
    if skip_col1.button("❓ Onbekend / overslaan", key=f"scorer_skip_{pending_id}", use_container_width=True):
        st.session_state["last_goal_event_id"] = None
        st.rerun()
    if skip_col2.button("↶ Maak goal ongedaan", key=f"scorer_undo_{pending_id}", use_container_width=True):
        remove_last_event()
        st.session_state["last_goal_event_id"] = None
        st.rerun()


def render_smart_tag_panel(team_name: str, prefix: str, color: str) -> None:
    st.markdown(
        f"<div class='team-header' style='background:{color};'>{team_name}</div>",
        unsafe_allow_html=True,
    )
    if st.session_state.pending_event and st.session_state.pending_team == team_name:
        st.markdown(f"**Kies zone voor: {st.session_state.pending_event}**")
        z1, z2, z3 = st.columns(3)
        if z1.button("Links", key=f"{prefix}_zone_left", use_container_width=True):
            add_smart_event(team_name, st.session_state.pending_event, "Linksvoor")
            st.rerun()
        if z2.button("Midden", key=f"{prefix}_zone_mid", use_container_width=True):
            add_smart_event(team_name, st.session_state.pending_event, "Middenvoor")
            st.rerun()
        if z3.button("Rechts", key=f"{prefix}_zone_right", use_container_width=True):
            add_smart_event(team_name, st.session_state.pending_event, "Rechtsvoor")
            st.rerun()
        if st.button("Annuleer", key=f"{prefix}_zone_cancel", use_container_width=True):
            clear_pending_tag()
            st.rerun()
        return
    st.markdown("**Kies actie**")
    rows = [
        ["Cirkelentry", "Schot", "Goal"],
        ["Schot op goal", "Press succes", "Hoge balverovering"],
        ["Turnover", "Turnover eigen helft", "Counter tegen na balverlies"],
        ["Opbouw mislukt", "Strafcorner"],
    ]
    for row_i, row in enumerate(rows):
        cols = st.columns(len(row))
        for col_i, event_name in enumerate(row):
            if cols[col_i].button(event_name, key=f"{prefix}_{row_i}_{col_i}", use_container_width=True):
                start_smart_tag(team_name, event_name)
                st.rerun()
    st.caption("Cirkelentry vraagt zonekeuze. Andere events worden direct opgeslagen.")


def render_cockpit_grid(df: pd.DataFrame) -> None:
    """Compacte cockpit-weergave: alles op één scherm voor tijdens de wedstrijd.

    Vier panelen in vaste grid:
        ┌─────────────┬──────────────────────────┐
        │  Scorebord  │   Event-knoppen (grid)   │
        │  + Kwart    │   (eigen team /          │
        │  + KPI's    │    tegenstander)         │
        ├─────────────┴──────────────────────────┤
        │  Scorer-picker (inline, na goal)       │
        ├────────────────────────────────────────┤
        │  Recente events-feed met undo          │
        └────────────────────────────────────────┘
    """
    # Top rij: scorebord
    render_match_scorebar()

    # Scorer-picker meteen er onder zodat hij zichtbaar is na goals
    render_scorer_picker()

    # Hoofdgrid: links KPI's + controls, rechts event-knoppen per team
    left, right = st.columns([0.9, 2.1])

    with left:
        # Kwart + tijd + snelle controls
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Kwart", st.session_state.quarter)
        with col2:
            st.metric("Tijd", current_time_str())

        # Mini-KPI's
        if df.empty:
            st.caption("Nog geen events — tag de eerste actie →")
        else:
            try:
                kpi = build_kpi_summary(df)
                st.markdown(
                    f"<div style='background:#f9fafb; border:1px solid #e5e7eb; border-radius:8px; "
                    f"padding:8px 12px; margin-top:6px;'>"
                    f"<div style='font-size:11px; color:#6b7280; text-transform:uppercase;'>Eigen team</div>"
                    f"<div style='font-size:13px; color:#111827;'>"
                    f"🅴 {kpi.get('team_entries', 0)} entries · "
                    f"🎯 {kpi.get('team_shots_on_goal', 0)}/{kpi.get('team_shots', 0)} op goal · "
                    f"🟨 {kpi.get('team_sc', 0)} SC"
                    f"</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
                st.markdown(
                    f"<div style='background:#fef2f2; border:1px solid #fecaca; border-radius:8px; "
                    f"padding:8px 12px; margin-top:6px;'>"
                    f"<div style='font-size:11px; color:#991b1b; text-transform:uppercase;'>Tegenstander</div>"
                    f"<div style='font-size:13px; color:#111827;'>"
                    f"🅴 {kpi.get('opp_entries', 0)} entries · "
                    f"🎯 {kpi.get('opp_shots_on_goal', 0)}/{kpi.get('opp_shots', 0)} op goal · "
                    f"🟨 {kpi.get('opp_sc', 0)} SC"
                    f"</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            except Exception:
                pass

        # Snelle controls
        st.markdown(" ")
        ctrl1, ctrl2 = st.columns(2)
        if ctrl1.button("↶ Undo", use_container_width=True, key="cockpit_undo"):
            remove_last_event()
            st.rerun()
        if ctrl2.button("↷ Herstel", use_container_width=True, key="cockpit_redo",
                        disabled=not st.session_state.get("undo_stack")):
            redo_last_event()
            st.rerun()
        if st.button("🕐 Reset klok", use_container_width=True, key="cockpit_reset_clock"):
            reset_timer()
            st.rerun()

        if st.session_state.get("last_undo_msg"):
            st.caption(f"💬 {st.session_state.last_undo_msg}")

    with right:
        st.markdown("#### Tag events")
        # Twee kolommen: eigen team / tegenstander
        own_col, opp_col = st.columns(2)
        with own_col:
            st.markdown(
                f"<div style='background:#1d4ed8; color:white; padding:6px 10px; "
                f"border-radius:6px; font-weight:600; text-align:center;'>"
                f"{st.session_state.team_name}</div>",
                unsafe_allow_html=True,
            )
            render_smart_tag_panel(st.session_state.team_name, "cockpit_team", TEAM_BLUE)
        with opp_col:
            st.markdown(
                f"<div style='background:#b91c1c; color:white; padding:6px 10px; "
                f"border-radius:6px; font-weight:600; text-align:center;'>"
                f"{st.session_state.opponent_name}</div>",
                unsafe_allow_html=True,
            )
            render_smart_tag_panel(st.session_state.opponent_name, "cockpit_opp", OPP_RED)

    # Onderste rij: events-feed
    st.markdown("---")
    feed_col, info_col = st.columns([1.4, 1])
    with feed_col:
        st.markdown("#### Laatste events")
        render_event_feed(df, max_items=8)
    with info_col:
        st.markdown("#### Coachfocus")
        cards = get_insight_cards(df)
        for i, card in enumerate(cards[:3]):
            render_info_card(card["title"], card["value"], card["subtitle"],
                             ["green", "orange", "blue", "red"][i % 4])


def render_live_screen(df: pd.DataFrame) -> None:
    # Cockpit-modus: radicale grid-weergave
    if st.session_state.get("device_mode") == "Cockpit":
        if is_viewer():
            render_match_scorebar()
            st.info("Viewer-modus: je kunt meekijken, maar geen events toevoegen of aanpassen.")
            render_event_feed(df, max_items=10)
            return
        render_cockpit_grid(df)
        return
    render_match_scorebar()
    render_scorer_picker()
    if is_viewer():
        st.info("Viewer-modus: je kunt meekijken, maar geen events toevoegen of aanpassen.")
        st.markdown("### Laatste events")
        render_event_feed(df, max_items=10)
        return
    a1, a2, a3, a4, a5 = st.columns(5)
    if a1.button("↶ Undo", use_container_width=True, key="live_undo_btn",
                 help="Laatst getagde event ongedaan maken"):
        remove_last_event()
        st.rerun()
    if a2.button("↷ Herstel", use_container_width=True, key="live_redo_btn",
                 disabled=not st.session_state.get("undo_stack"),
                 help="Laatst verwijderde event terughalen"):
        redo_last_event()
        st.rerun()
    if a3.button("Ververs", use_container_width=True, key="live_refresh_btn",
                 help="Analyse opnieuw berekenen"):
        refresh_derived_state()
        st.rerun()
    if a4.button("Reset klok", use_container_width=True, key="live_reset_clock_btn"):
        reset_timer()
        st.rerun()
    # Alleen coach mag de hele wedstrijd resetten
    reset_disabled = not is_coach()
    if a5.button("Reset wedstrijd", use_container_width=True, key="live_reset_match_btn",
                 disabled=reset_disabled,
                 help="Alleen coach mag dit" if reset_disabled else "Alle events verwijderen"):
        st.session_state.confirm_reset = not st.session_state.confirm_reset
        st.rerun()
    # Undo-feedback tonen
    if st.session_state.get("last_undo_msg"):
        st.caption(f"💬 {st.session_state.last_undo_msg}")
    if st.session_state.confirm_reset:
        st.warning("Weet je zeker dat je de wedstrijd wilt resetten? Dit verwijdert ALLE events.")
        r1, r2 = st.columns(2)
        if r1.button("Ja, reset alles", use_container_width=True, key="live_confirm_reset_yes"):
            reset_all()
            st.session_state.last_undo_msg = ""
            st.rerun()
        if r2.button("Annuleer reset", use_container_width=True, key="live_confirm_reset_no"):
            st.session_state.confirm_reset = False
            st.rerun()
    mode = st.session_state.device_mode
    if mode == "iPhone":
        st.markdown("### 📱 iPhone coachmodus")
        p1, p2 = st.columns(2)
        with p1:
            st.metric("Tijd", current_time_str())
        with p2:
            st.metric("Kwart", st.session_state.quarter)
        st.markdown("#### Eigen team")
        render_smart_tag_panel(st.session_state.team_name, "iphone_team", TEAM_BLUE)
        st.markdown("#### Tegenstander")
        render_smart_tag_panel(st.session_state.opponent_name, "iphone_opp", OPP_RED)
        st.markdown("### Laatste events")
        render_event_feed(df, max_items=5)
        return
    if mode == "MacBook":
        left, mid, right = st.columns([1.05, 1.05, 0.9])
        with left:
            render_smart_tag_panel(st.session_state.team_name, "team", TEAM_BLUE)
        with mid:
            render_smart_tag_panel(st.session_state.opponent_name, "opp", OPP_RED)
        with right:
            st.markdown("### Live inzichten")
            for i, card in enumerate(get_insight_cards(df)):
                render_info_card(card["title"], card["value"], card["subtitle"], ["green", "orange", "blue", "red"][i])
        b1, b2 = st.columns([1.1, 1.2])
        with b1:
            st.markdown("### Laatste events")
            render_event_feed(df, max_items=10)
        with b2:
            st.markdown("### Match timeline")
            render_timeline(df)
        return
    left, right = st.columns(2)
    with left:
        render_smart_tag_panel(st.session_state.team_name, "team", TEAM_BLUE)
    with right:
        render_smart_tag_panel(st.session_state.opponent_name, "opp", OPP_RED)
    l1, l2 = st.columns([1.2, 1.1])
    with l1:
        st.markdown("### Laatste events")
        render_event_feed(df, max_items=10)
    with l2:
        st.markdown("### Coachfocus")
        for i, card in enumerate(get_insight_cards(df)):
            render_info_card(card["title"], card["value"], card["subtitle"], ["green", "orange", "blue", "red"][i])


def render_analysis_screen(df: pd.DataFrame) -> None:
    if df.empty:
        st.info("Nog geen events toegevoegd.")
        return
    kpi = build_kpi_summary(df)
    st.markdown("### Kernstatistieken")
    row1 = st.columns(4)
    with row1[0]:
        render_info_card("Cirkelentries", str(kpi["team_entries"]), "Entries eigen team", "blue")
    with row1[1]:
        render_info_card("Schoten", str(kpi["team_shots"]), "Pogingen naast of geblokt", "orange")
    with row1[2]:
        render_info_card("Schoten op goal", str(kpi["team_shots_on_goal"]), "Doelpogingen op goal", "green")
    with row1[3]:
        render_info_card("Shot on goal → goal", f"{kpi['team_shot_to_goal_pct']:.0f}%", "Afwerking op doelpogingen", "blue")
    st.markdown("### Statistieken per kwart")
    quarter_df = build_quarter_stats_df(df)
    st.dataframe(quarter_df, use_container_width=True, hide_index=True)
    st.markdown("### Momentum analyse")
    moments = detect_momentum(df)
    if moments:
        for m in moments:
            st.success(m)
    else:
        st.info("Nog geen duidelijke momentumfase herkend.")
    st.markdown("### Cirkelentry heatmap")
    st.dataframe(build_entry_heatmap(df), use_container_width=True, hide_index=True)
    st.markdown("### Match timeline")
    render_timeline(df)


def render_field_screen(df: pd.DataFrame) -> None:
    if df.empty:
        st.info("Nog geen data voor veldvisualisatie.")
        return
    teams = [st.session_state.team_name, st.session_state.opponent_name]
    # Default team als er nog niks gekozen is of als de eerder gekozen naam niet meer bestaat
    current_team = st.session_state.get("field_team") or ""
    if current_team not in teams:
        current_team = st.session_state.team_name
    team_idx = teams.index(current_team)
    c1, c2 = st.columns(2)
    with c1:
        st.selectbox("Kies team", teams, index=team_idx, key="field_team")
    with c2:
        quarters_list = ["Alles"] + QUARTERS
        cur_q = st.session_state.get("field_quarter") or "Alles"
        q_idx = quarters_list.index(cur_q) if cur_q in quarters_list else 0
        st.selectbox("Kies kwart", quarters_list, index=q_idx, key="field_quarter")
    st.multiselect("Toon lagen", ["Cirkelentry", "Schot", "Goal"], key="field_layers")
    st.caption("Alleen cirkelentries worden per links / midden / rechts-zone opgeslagen. Schoten en goals tellen mee als event, maar niet meer per zone.")
    selected_layers = st.session_state.field_layers or ["Cirkelentry"]
    render_field_view(df, st.session_state.field_team, st.session_state.field_quarter, selected_layers)


def render_scorers_editor(df: pd.DataFrame) -> None:
    """Toon goals van het eigen team en laat de scorer achteraf koppelen/wijzigen."""
    if df.empty:
        return
    own = st.session_state.get("team_name", "")
    if not own:
        return
    own_goals = df[(df["team"] == own) & (df["event"] == "Goal")]
    if own_goals.empty:
        return

    st.markdown("### ⚽ Scorers")
    roster = _active_team_roster()
    id_to_name = {p["id"]: p["name"] for p in roster}

    # Topscorers overzicht
    scored = own_goals[own_goals["player_id"].notna() & (own_goals["player_id"] != "")]
    if not scored.empty:
        counts = scored["player_id"].value_counts()
        top_lines = []
        for pid, cnt in counts.items():
            nm = id_to_name.get(pid, "onbekend")
            top_lines.append(f"<span style='background:#eff6ff; color:#1d4ed8; "
                             f"padding:4px 8px; border-radius:999px; margin-right:6px; "
                             f"display:inline-block; font-size:13px;'>{nm} · {int(cnt)}</span>")
        st.markdown(
            "<div style='margin-bottom:8px;'>" + "".join(top_lines) + "</div>",
            unsafe_allow_html=True,
        )

    # Per-goal editor
    opt_ids = [""] + [p["id"] for p in roster]
    opt_labels = ["— Onbekend —"] + [f"{p['name']} ({p.get('line','?')})" for p in roster]

    with st.expander(f"Bewerk scorer per goal ({len(own_goals)} goals)", expanded=False):
        if not roster:
            st.caption("Geen spelers in roster — voeg eerst spelers toe in de Wisselschema-tool.")
        else:
            for _, row in own_goals.sort_values(["quarter", "time"]).iterrows():
                ev_id = row["id"]
                current_pid = row.get("player_id") or ""
                try:
                    idx = opt_ids.index(current_pid)
                except ValueError:
                    idx = 0
                c1, c2 = st.columns([2, 3])
                with c1:
                    st.markdown(f"**{row['quarter']} · {row['time']}**")
                with c2:
                    new_idx = st.selectbox(
                        "Scorer",
                        options=list(range(len(opt_ids))),
                        format_func=lambda i: opt_labels[i],
                        index=idx,
                        key=f"scorer_edit_{ev_id}",
                        label_visibility="collapsed",
                    )
                    new_pid = opt_ids[new_idx] or None
                    if new_pid != (current_pid or None):
                        set_event_player(ev_id, new_pid)
                        st.rerun()


def render_report_screen(df: pd.DataFrame) -> None:
    if df.empty:
        st.info("Nog geen data voor coachrapport.")
        return
    st.markdown("### Rustanalyse")
    if st.button("Genereer rustanalyse", key="report_generate_halftime"):
        st.session_state.halftime_report = generate_halftime_report(df)
    if st.session_state.halftime_report:
        st.text_area("Rustanalyse", st.session_state.halftime_report, height=200, key="report_halftime_area")
    st.markdown("### Volledig coachrapport")
    report_text = st.session_state.auto_notes
    st.text_area("Rapport", report_text, height=420, key="report_full_area")
    st.markdown("### Statistieken per kwart")
    quarter_df = build_quarter_stats_df(df)
    st.dataframe(quarter_df, use_container_width=True, hide_index=True)
    st.markdown("### Exports")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.download_button("Download TXT rapport", data=report_text.encode("utf-8"), file_name="coachrapport.txt", key="dl_report_txt")
    with c2:
        try:
            pdf_bytes = export_match_report_pdf(
                df,
                st.session_state.get("team_name", "Ons team"),
                st.session_state.get("opponent_name", "Tegenstander"),
                st.session_state.get("match_id", ""),
            )
            pdf_mime = "application/pdf" if REPORTLAB_AVAILABLE else "text/plain"
        except Exception:
            pdf_bytes = export_pdf_report(report_text)
            pdf_mime = "application/pdf" if REPORTLAB_AVAILABLE else "text/plain"
        st.download_button(
            "📄 Download PDF-rapport",
            data=pdf_bytes,
            file_name=f"coachrapport_{st.session_state.get('match_id','match')}.pdf",
            mime=pdf_mime,
            key="dl_report_pdf",
        )
    with c3:
        st.download_button(
            "Download Excel",
            data=export_excel(df),
            file_name="wedstrijd_analyse.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_report_excel",
        )
    # ---- Scorers: goals per speler bewerken ----
    render_scorers_editor(df)

    st.markdown("### Eventlog")
    st.dataframe(
        df[["quarter", "time", "team", "event", "zone", "source", "video_time_sec", "notes"]],
        use_container_width=True,
        hide_index=True,
    )

    # ---- Backup & herstel ----
    st.markdown("### Backup & herstel")
    st.caption("Download een JSON-backup als veiligheidskopie, of herstel eerder opgeslagen data.")
    bu1, bu2 = st.columns(2)
    with bu1:
        backup_payload = {
            "match_id": st.session_state.match_id,
            "team_name": st.session_state.team_name,
            "opponent_name": st.session_state.opponent_name,
            "events": st.session_state.events,
            "video_clips": st.session_state.video_clips,
            "pushoff_offsets": st.session_state.pushoff_offsets,
            "backup_time": time.strftime("%Y-%m-%d %H:%M:%S"),
        }
        import json as _json
        backup_bytes = _json.dumps(backup_payload, ensure_ascii=False, indent=2).encode("utf-8")
        st.download_button(
            "📦 Download backup (JSON)",
            data=backup_bytes,
            file_name=f"backup_{st.session_state.match_id}_{time.strftime('%Y%m%d_%H%M')}.json",
            mime="application/json",
            use_container_width=True,
            key="dl_backup_json",
        )
    with bu2:
        restore_file = st.file_uploader(
            "Herstel backup (JSON)",
            type=["json"],
            key="restore_backup_uploader",
            help="Upload een eerder gedownloade backup om alles terug te zetten.",
        )
        if restore_file is not None:
            if st.button("✔︎ Herstel nu", use_container_width=True, key="restore_backup_btn"):
                try:
                    import json as _json
                    data = _json.loads(restore_file.read().decode("utf-8"))
                    st.session_state.events = data.get("events", [])
                    st.session_state.video_clips = data.get("video_clips", [])
                    st.session_state.pushoff_offsets = data.get(
                        "pushoff_offsets", {"Q1": None, "Q2": None, "Q3": None, "Q4": None}
                    )
                    if data.get("team_name"):
                        st.session_state.team_name = data["team_name"]
                    if data.get("opponent_name"):
                        st.session_state.opponent_name = data["opponent_name"]
                    if data.get("match_id"):
                        st.session_state.match_id = data["match_id"]
                    refresh_derived_state()
                    st.success(
                        f"Backup hersteld: {len(st.session_state.events)} events, "
                        f"{len(st.session_state.video_clips)} clips."
                    )
                    st.rerun()
                except Exception as err:
                    st.error(f"Backup niet geldig: {type(err).__name__}")


# ==================================================
# VIDEOSPELER COMPONENT (HTML + JS)
# ==================================================
def build_video_player_html(
    source_type: str,
    video_url: str,
    jump_to: float | None,
    player_height: int = 440,
) -> str:
    """Bouw een custom videospeler-component.
    source_type: 'file' (blob/url), 'url' (direct mp4), 'youtube'.
    """
    jump_js = "null" if jump_to is None else f"{float(jump_to):.2f}"
    url_js = json.dumps(video_url or "")

    if source_type == "youtube":
        yt_id = extract_youtube_id(video_url)
        yt_id_js = json.dumps(yt_id)
        html = f"""
        <div style="font-family:system-ui,Segoe UI,Roboto,sans-serif;">
          <div id="yt-player" style="width:100%;aspect-ratio:16/9;background:#000;border-radius:14px;overflow:hidden;"></div>
          <div style="display:flex;gap:6px;flex-wrap:wrap;margin-top:10px;">
            <button onclick="ytPlay()" style="padding:10px 14px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;cursor:pointer;">▶ Play</button>
            <button onclick="ytPause()" style="padding:10px 14px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;cursor:pointer;">⏸ Pauze</button>
            <button onclick="ytSeek(-5)" style="padding:10px 14px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;cursor:pointer;">⏪ -5s</button>
            <button onclick="ytSeek(5)" style="padding:10px 14px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;cursor:pointer;">⏩ +5s</button>
            <select id="yt-speed" onchange="ytSetSpeed(this.value)" style="padding:10px 12px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;">
              <option value="0.25">0.25x</option>
              <option value="0.5">0.5x</option>
              <option value="1" selected>1x</option>
              <option value="1.5">1.5x</option>
              <option value="2">2x</option>
            </select>
            <div style="padding:10px 14px;border-radius:10px;background:#f1f5f9;color:#0f172a;font-weight:700;">Tijd: <span id="yt-time">00:00</span></div>
            <button onclick="ytCopyTime()" style="padding:10px 14px;border-radius:10px;border:1px solid #2563eb;background:#2563eb;color:#fff;font-weight:700;cursor:pointer;">📋 Kopieer tijd</button>
          </div>
          <div id="yt-hint" style="margin-top:8px;color:#475569;font-size:13px;">Klik 'Kopieer tijd' en plak in het veld 'videotijd' hieronder om een tag te koppelen.</div>
        </div>
        <script src="https://www.youtube.com/iframe_api"></script>
        <script>
          var ytPlayer;
          var ytReady = false;
          var ytTargetSeek = {jump_js};
          function onYouTubeIframeAPIReady() {{
            ytPlayer = new YT.Player('yt-player', {{
              videoId: {yt_id_js},
              playerVars: {{ 'playsinline': 1, 'rel': 0, 'modestbranding': 1 }},
              events: {{
                'onReady': function(e) {{
                  ytReady = true;
                  if (ytTargetSeek !== null) {{ e.target.seekTo(ytTargetSeek, true); }}
                  setInterval(updateYtTime, 500);
                }}
              }}
            }});
          }}
          function updateYtTime() {{
            if (!ytReady) return;
            var t = ytPlayer.getCurrentTime();
            var mm = String(Math.floor(t/60)).padStart(2,'0');
            var ss = String(Math.floor(t%60)).padStart(2,'0');
            document.getElementById('yt-time').innerText = mm + ':' + ss;
          }}
          function ytPlay() {{ if (ytReady) ytPlayer.playVideo(); }}
          function ytPause() {{ if (ytReady) ytPlayer.pauseVideo(); }}
          function ytSeek(delta) {{ if (ytReady) ytPlayer.seekTo(ytPlayer.getCurrentTime() + delta, true); }}
          function ytSetSpeed(v) {{ if (ytReady) ytPlayer.setPlaybackRate(parseFloat(v)); }}
          function ytCopyTime() {{
            if (!ytReady) return;
            var t = ytPlayer.getCurrentTime();
            navigator.clipboard.writeText(t.toFixed(2));
            var h = document.getElementById('yt-hint');
            h.innerText = '✅ Videotijd gekopieerd: ' + t.toFixed(2) + ' sec — plak het hieronder in het veld videotijd.';
          }}
        </script>
        """
        return html

    # File (blob/data-url) of directe mp4-link: HTML5 video
    html = f"""
    <div style="font-family:system-ui,Segoe UI,Roboto,sans-serif;">
      <video id="hv" src={url_js} controls playsinline style="width:100%;aspect-ratio:16/9;background:#000;border-radius:14px;"></video>
      <div style="display:flex;gap:6px;flex-wrap:wrap;margin-top:10px;">
        <button onclick="hvPlay()" style="padding:10px 14px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;cursor:pointer;">▶ Play</button>
        <button onclick="hvPause()" style="padding:10px 14px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;cursor:pointer;">⏸ Pauze</button>
        <button onclick="hvSeek(-5)" style="padding:10px 14px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;cursor:pointer;">⏪ -5s</button>
        <button onclick="hvSeek(5)" style="padding:10px 14px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;cursor:pointer;">⏩ +5s</button>
        <button onclick="hvFrame(-1)" style="padding:10px 14px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;cursor:pointer;">⏮ frame</button>
        <button onclick="hvFrame(1)" style="padding:10px 14px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;cursor:pointer;">frame ⏭</button>
        <select id="hv-speed" onchange="hvSetSpeed(this.value)" style="padding:10px 12px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;">
          <option value="0.25">0.25x (slow mo)</option>
          <option value="0.5">0.5x</option>
          <option value="1" selected>1x</option>
          <option value="1.5">1.5x</option>
          <option value="2">2x</option>
        </select>
        <div style="padding:10px 14px;border-radius:10px;background:#f1f5f9;color:#0f172a;font-weight:700;">Tijd: <span id="hv-time">00:00.00</span></div>
        <button onclick="hvCopyTime()" style="padding:10px 14px;border-radius:10px;border:1px solid #2563eb;background:#2563eb;color:#fff;font-weight:700;cursor:pointer;">📋 Kopieer tijd</button>
      </div>
      <div id="hv-hint" style="margin-top:8px;color:#475569;font-size:13px;">
        Sneltoetsen: <b>spatie</b> = play/pauze, <b>←/→</b> = 5 sec, <b>j/l</b> = 5 sec, <b>,/.</b> = frame voor/achter.
      </div>
    </div>
    <script>
      (function(){{
        var v = document.getElementById('hv');
        var target = {jump_js};
        if (target !== null) {{ v.addEventListener('loadedmetadata', function(){{ v.currentTime = target; }}); }}
        setInterval(function(){{
          var t = v.currentTime || 0;
          var mm = String(Math.floor(t/60)).padStart(2,'0');
          var ss = String(Math.floor(t%60)).padStart(2,'0');
          var ms = String(Math.floor((t%1)*100)).padStart(2,'0');
          var el = document.getElementById('hv-time');
          if (el) el.innerText = mm + ':' + ss + '.' + ms;
        }}, 100);
        window.hvPlay = function(){{ v.play(); }};
        window.hvPause = function(){{ v.pause(); }};
        window.hvSeek = function(d){{ v.currentTime = Math.max(0, v.currentTime + d); }};
        window.hvFrame = function(d){{ v.pause(); v.currentTime = Math.max(0, v.currentTime + d*0.04); }};
        window.hvSetSpeed = function(x){{ v.playbackRate = parseFloat(x); }};
        window.hvCopyTime = function(){{
          var t = (v.currentTime || 0).toFixed(2);
          if (navigator.clipboard) navigator.clipboard.writeText(t);
          var h = document.getElementById('hv-hint');
          h.innerText = '✅ Videotijd gekopieerd: ' + t + ' sec — plak in het veld videotijd hieronder.';
        }};
        document.addEventListener('keydown', function(e){{
          if (e.target.tagName === 'INPUT' || e.target.tagName === 'TEXTAREA') return;
          if (e.code === 'Space') {{ e.preventDefault(); if (v.paused) v.play(); else v.pause(); }}
          if (e.code === 'ArrowLeft' || e.key === 'j') {{ e.preventDefault(); v.currentTime = Math.max(0, v.currentTime - 5); }}
          if (e.code === 'ArrowRight' || e.key === 'l') {{ e.preventDefault(); v.currentTime = v.currentTime + 5; }}
          if (e.key === ',') {{ e.preventDefault(); v.pause(); v.currentTime = Math.max(0, v.currentTime - 0.04); }}
          if (e.key === '.') {{ e.preventDefault(); v.pause(); v.currentTime = v.currentTime + 0.04; }}
        }});
      }})();
    </script>
    """
    return html


# ==================================================
# BEELDANALYSE-SCHERM (nieuw)
# ==================================================
def render_video_analysis_screen(df: pd.DataFrame) -> None:
    st.markdown("### 🎥 Beeldanalyse")
    st.caption(
        "Upload een wedstrijdvideo of plak een link. Markeer per kwart het push-off-moment, "
        "klik op tags om erheen te springen, en maak nieuwe tags of clips op basis van de video."
    )

    # --- Video-bron kiezen ---
    with st.expander("Grote videobestanden / limiet verhogen"):
        st.markdown("""
Voor lokaal gebruik kun je de uploadlimiet verhogen door een bestand `.streamlit/config.toml` te maken met:
```toml
[server]
maxUploadSize = 2000
```
Op Streamlit Cloud is de limiet strakker — gebruik daar liever een directe mp4-link of een YouTube-link.
        """)

    source_mode = st.radio(
        "Videobron",
        ["Bestand uploaden", "Directe videolink (mp4)", "YouTube-link"],
        horizontal=True,
        key="video_source_mode",
    )

    active_video_name = ""
    source_type = "file"
    video_url = ""

    if source_mode == "Bestand uploaden":
        video_file = st.file_uploader(
            "Upload wedstrijdvideo",
            type=["mp4", "mov", "m4v", "webm"],
            key="match_video_uploader",
        )
        if video_file is not None:
            # Zet video om naar data-URL zodat JS-speler hem kan gebruiken
            video_bytes = video_file.getvalue()
            import base64
            b64 = base64.b64encode(video_bytes).decode("ascii")
            mime = "video/mp4" if video_file.name.lower().endswith((".mp4", ".m4v")) else (
                "video/quicktime" if video_file.name.lower().endswith(".mov") else "video/webm"
            )
            video_url = f"data:{mime};base64,{b64}"
            source_type = "file"
            active_video_name = video_file.name
            st.session_state.uploaded_video_name = video_file.name
            st.session_state.video_source_type = "file"
            st.session_state.video_url = video_url
            st.success(f"Video geladen: {video_file.name}")
        else:
            st.info("Nog geen video geladen. Je kunt wel alvast clips handmatig registreren.")
    elif source_mode == "Directe videolink (mp4)":
        url = st.text_input(
            "Directe videolink",
            placeholder="https://voorbeeld.nl/wedstrijd.mp4",
            key="match_video_url_input",
        )
        if url and is_probable_video_url(url):
            video_url = url
            source_type = "url"
            active_video_name = url
            st.session_state.uploaded_video_name = url
            st.session_state.video_source_type = "url"
            st.session_state.video_url = url
            st.success("Videolink geladen.")
        elif url:
            st.warning("Gebruik een volledige link die begint met http:// of https://.")
        else:
            st.info("Nog geen videolink ingevuld.")
    else:
        url = st.text_input(
            "YouTube-link",
            placeholder="https://www.youtube.com/watch?v=...",
            key="match_video_yt_input",
        )
        if url and is_youtube_url(url) and extract_youtube_id(url):
            video_url = url
            source_type = "youtube"
            active_video_name = url
            st.session_state.uploaded_video_name = url
            st.session_state.video_source_type = "youtube"
            st.session_state.video_url = url
            st.success("YouTube-link geladen.")
        elif url:
            st.warning("Deze link herken ik niet als YouTube-link.")
        else:
            st.info("Nog geen YouTube-link ingevuld.")

    # --- Videospeler + tag-paneel naast elkaar ---
    if video_url:
        left, right = st.columns([1.4, 1])
        with left:
            st.markdown("#### Video")
            jump_to = st.session_state.jump_to_video_time
            html = build_video_player_html(source_type, video_url, jump_to)
            components.html(html, height=520)
            # Eenmaal gesprongen, reset zodat hij bij volgende rerun niet opnieuw springt
            st.session_state.jump_to_video_time = None

            # --- Push-off kalibratie per kwart (ingeklapt om het scherm rustig te houden) ---
            # Korte samenvatting altijd zichtbaar
            offsets_summary = " • ".join([
                f"{q}: {('—' if v is None else format_seconds_to_mmss(v))}"
                for q, v in st.session_state.pushoff_offsets.items()
            ])
            set_count = sum(1 for v in st.session_state.pushoff_offsets.values() if v is not None)
            with st.expander(f"⏱️ Push-off kalibratie ({set_count}/4 kwarten gezet) — {offsets_summary}", expanded=set_count == 0):
                st.caption(
                    "Zet de video op het moment van push-off van dit kwart, klik op 'Kopieer tijd' in de speler, "
                    "plak hieronder, en druk op 'Markeer push-off'. Herhaal per kwart. "
                    "Kalibraties worden automatisch in de cloud opgeslagen."
                )
                pc1, pc2, pc3 = st.columns([1, 1.3, 1])
                with pc1:
                    calib_quarter = st.selectbox("Kwart", QUARTERS, key="calib_quarter_select")
                with pc2:
                    calib_val = st.text_input(
                        "Videotijd (sec)",
                        value="",
                        placeholder="bv. 47.20",
                        key=f"calib_val_{calib_quarter}",
                    )
                with pc3:
                    st.write("")
                    if st.button("✅  Markeer push-off", use_container_width=True, key="btn_set_pushoff", type="primary"):
                        try:
                            v = float(calib_val.replace(",", "."))
                            st.session_state.pushoff_offsets[calib_quarter] = v
                            # Ook naar cloud zodat de kalibratie persistent blijft
                            try:
                                cloud_save_pushoff_offset(st.session_state.match_id, calib_quarter, v)
                            except Exception as err:
                                log_cloud_error("push-off opslaan", err)
                            st.success(f"Push-off {calib_quarter} = {v:.2f} sec opgeslagen.")
                            st.rerun()
                        except Exception:
                            st.error("Geen geldige tijd. Gebruik een getal zoals 47.2")

        with right:
            render_video_tag_panel(df)
    else:
        st.info("Zodra je een video hebt geladen verschijnt hier de speler met klikbare tags ernaast.")

    # --- Highlight-reel (auto-afspelen van events) ---
    if video_url:
        st.markdown("---")
        render_highlight_reel_section(df, video_url, source_type)

    # --- Onderste deel: clip toevoegen + cliplog ---
    st.markdown("---")
    render_clip_section(df, active_video_name or st.session_state.uploaded_video_name)

    # --- Tekenen op beeld (tactische analyse) ---
    st.markdown("---")
    render_draw_section()


def render_video_tag_panel(df: pd.DataFrame) -> None:
    """Rechts naast de video: filters + klikbare lijst van events + handmatig taggen."""
    st.markdown("#### 🏷️ Tags & sprongen")

    if df.empty:
        st.info("Nog geen tags. Voeg ze toe op het LIVE-scherm of hieronder met 'Tag op video-moment'.")
    else:
        fcol1, fcol2 = st.columns(2)
        with fcol1:
            filter_quarter = st.selectbox("Kwart filter", ["Alles"] + QUARTERS, key="vt_filter_q")
        with fcol2:
            filter_team = st.selectbox(
                "Team filter",
                ["Alles", st.session_state.team_name, st.session_state.opponent_name],
                key="vt_filter_t",
            )
        filter_event = st.multiselect(
            "Event filter",
            sorted(df["event"].dropna().unique().tolist()),
            key="vt_filter_e",
        )

        fdf = df.copy()
        if filter_quarter != "Alles":
            fdf = fdf[fdf["quarter"] == filter_quarter]
        if filter_team != "Alles":
            fdf = fdf[fdf["team"] == filter_team]
        if filter_event:
            fdf = fdf[fdf["event"].isin(filter_event)]

        st.caption(f"{len(fdf)} tags tonen. Klik ▶️ om naar het moment in de video te springen (met 3 sec pre-roll).")
        pre_roll = st.slider("Pre-roll (seconden voor tag)", 0, 15, 3, 1, key="vt_preroll")

        fdf = fdf.sort_values(["quarter", "time"])
        for _, row in fdf.iterrows():
            game_sec = parse_mmss(str(row["time"]))
            video_sec = None
            # Als de tag uit video komt, hebben we al een video_time_sec
            if pd.notna(row.get("video_time_sec", None)):
                try:
                    video_sec = float(row["video_time_sec"])
                except Exception:
                    video_sec = None
            else:
                video_sec = video_time_from_game(game_sec, str(row["quarter"]))
            team_color = TEAM_BLUE if row["team"] == st.session_state.team_name else OPP_RED
            zone_txt = f" • {row['zone']}" if str(row.get("zone", "")).strip() else ""
            label = f"{row['quarter']} {row['time']} • {row['event']}{zone_txt}"
            c1, c2 = st.columns([3, 1])
            with c1:
                st.markdown(
                    f"<div style='padding:8px 12px;border-left:4px solid {team_color};"
                    f"background:{CARD_BG_ELEVATED};border:1px solid {CARD_BORDER_SOFT};"
                    f"border-radius:8px;margin-bottom:4px;'>"
                    f"<small style='color:{TEXT_SUB};'>{row['team']}</small><br>"
                    f"<strong style='color:{TEXT_MAIN};'>{label}</strong></div>",
                    unsafe_allow_html=True,
                )
            with c2:
                disabled = video_sec is None
                if st.button(
                    "▶️ Spring naar tag",
                    key=f"jump_{row['id']}",
                    use_container_width=True,
                    disabled=disabled,
                    help="Push-off van dit kwart is nog niet gezet." if disabled else "Spring naar dit moment in de video.",
                ):
                    st.session_state.jump_to_video_time = max(0.0, float(video_sec) - pre_roll)
                    st.rerun()

    # --- Handmatig tag aanmaken op huidige video-tijd ---
    st.markdown("---")
    st.markdown("#### ➕ Tag op video-moment")
    st.caption("Klik op 'Kopieer tijd' in de videospeler, plak hieronder en kies team + actie.")

    t1, t2 = st.columns([1.2, 1])
    with t1:
        vt_time = st.text_input("Videotijd (sec)", value="", placeholder="bv. 312.4", key="new_tag_vtime")
    with t2:
        vt_quarter = st.selectbox("Kwart", QUARTERS, key="new_tag_q")

    vt_team = st.selectbox(
        "Team",
        [st.session_state.team_name, st.session_state.opponent_name],
        key="new_tag_team",
    )
    vt_event = st.selectbox(
        "Event",
        ["Cirkelentry", "Schot", "Schot op goal", "Goal", "Press succes", "Hoge balverovering",
         "Turnover", "Turnover eigen helft", "Counter tegen na balverlies", "Opbouw mislukt", "Strafcorner"],
        key="new_tag_event",
    )
    vt_zone = ""
    if vt_event in EVENT_NEEDS_ZONE:
        vt_zone = st.selectbox("Zone", FIELD_ZONES, key="new_tag_zone")

    if st.button("Tag opslaan", use_container_width=True, key="btn_save_video_tag"):
        try:
            video_sec = float(str(vt_time).replace(",", "."))
        except Exception:
            st.error("Geen geldige videotijd. Typ een getal zoals 312.4")
            return
        # Zet videotijd → wedstrijdtijd als we een push-off-offset hebben voor dit kwart
        offset = get_pushoff_offset(vt_quarter)
        if offset is None:
            st.warning(
                f"Push-off voor {vt_quarter} is nog niet gekalibreerd. De tag krijgt 00:00 als wedstrijdtijd; "
                "de videotijd wordt wel correct bewaard."
            )
            game_sec = 0
        else:
            game_sec = int(max(0, video_sec - offset))

        row = normalize_event_row({
            "id": str(uuid.uuid4()),
            "match_id": st.session_state.match_id,
            "quarter": vt_quarter,
            "time": format_seconds_to_mmss(game_sec),
            "team": vt_team,
            "event": vt_event,
            "zone": vt_zone,
            "notes": "",
            "created_at": time.time(),
            "source": "video",
            "video_time_sec": video_sec,
        })
        st.session_state.events.append(row)
        try:
            save_event_to_cloud(row)
            mark_cloud_ok()
        except Exception as err:
            log_cloud_error("video-tag opslaan", err)
        refresh_derived_state()
        st.success(f"Tag opgeslagen op videotijd {video_sec:.2f}s.")
        st.rerun()


def build_highlight_reel_html(
    video_id: str,
    fragments: list[dict],
    pre_roll: float = 3.0,
    post_roll: float = 4.0,
    player_height: int = 440,
) -> str:
    """Bouw een YouTube-speler die automatisch door een lijst fragmenten speelt.

    fragments: lijst met dicts met keys: start_sec, label (bv. "Q1 08:23 • Goal voor").
    Speelt elk fragment van (start_sec - pre_roll) tot (start_sec + post_roll) af en gaat daarna door naar het volgende.
    """
    yt_id_js = json.dumps(video_id)
    frags_js = json.dumps([
        {
            "start": max(0.0, float(f["start_sec"]) - pre_roll),
            "end": float(f["start_sec"]) + post_roll,
            "label": str(f.get("label", "")),
        }
        for f in fragments
    ])
    html = f"""
    <div style="font-family:system-ui,Segoe UI,Roboto,sans-serif;">
      <div id="reel-player" style="width:100%;aspect-ratio:16/9;background:#000;border-radius:14px;overflow:hidden;"></div>
      <div style="display:flex;gap:6px;flex-wrap:wrap;margin-top:10px;align-items:center;">
        <button onclick="reelPrev()" style="padding:10px 14px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;cursor:pointer;">⏮ Vorige</button>
        <button onclick="reelReplay()" style="padding:10px 14px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;cursor:pointer;">↻ Herhaal</button>
        <button onclick="reelNext()" style="padding:10px 14px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;cursor:pointer;">Volgende ⏭</button>
        <button onclick="reelAutoToggle()" id="reel-auto-btn" style="padding:10px 14px;border-radius:10px;border:1px solid #16a34a;background:#16a34a;color:#fff;font-weight:700;cursor:pointer;">▶ Auto aan</button>
        <select id="reel-speed" onchange="reelSpeed(this.value)" style="padding:10px 12px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;">
          <option value="0.5">0.5x</option>
          <option value="0.75">0.75x</option>
          <option value="1" selected>1x</option>
          <option value="1.25">1.25x</option>
          <option value="1.5">1.5x</option>
        </select>
        <div style="padding:10px 14px;border-radius:10px;background:#f1f5f9;color:#0f172a;font-weight:700;">
          <span id="reel-idx">1</span> / <span id="reel-total">{len(fragments)}</span>
        </div>
      </div>
      <div id="reel-label" style="margin-top:10px;padding:10px 14px;border-radius:10px;background:#eff6ff;border-left:4px solid #2563eb;font-weight:600;color:#1e3a8a;">
        Klaar om af te spelen.
      </div>
    </div>
    <script src="https://www.youtube.com/iframe_api"></script>
    <script>
      var reelPlayer;
      var reelReady = false;
      var reelFrags = {frags_js};
      var reelIdx = 0;
      var reelAuto = true;
      var reelTimer = null;

      function onYouTubeIframeAPIReady() {{
        reelPlayer = new YT.Player('reel-player', {{
          videoId: {yt_id_js},
          playerVars: {{ 'playsinline': 1, 'rel': 0, 'modestbranding': 1, 'controls': 1 }},
          events: {{
            'onReady': function(e) {{
              reelReady = true;
              playFrag(0);
            }},
            'onStateChange': function(e) {{
              if (e.data === YT.PlayerState.PLAYING) {{
                checkEndLoop();
              }}
            }}
          }}
        }});
      }}

      function playFrag(i) {{
        if (!reelReady || !reelFrags.length) return;
        if (i < 0) i = 0;
        if (i >= reelFrags.length) {{ reelIdx = reelFrags.length - 1; return; }}
        reelIdx = i;
        var f = reelFrags[i];
        document.getElementById('reel-idx').innerText = (i + 1);
        document.getElementById('reel-label').innerText = (i + 1) + '. ' + f.label;
        reelPlayer.seekTo(f.start, true);
        reelPlayer.playVideo();
      }}

      function checkEndLoop() {{
        if (reelTimer) clearInterval(reelTimer);
        reelTimer = setInterval(function() {{
          if (!reelReady) return;
          var t = reelPlayer.getCurrentTime();
          var f = reelFrags[reelIdx];
          if (!f) return;
          if (t >= f.end) {{
            clearInterval(reelTimer);
            reelTimer = null;
            if (reelAuto && reelIdx < reelFrags.length - 1) {{
              playFrag(reelIdx + 1);
            }} else {{
              reelPlayer.pauseVideo();
              document.getElementById('reel-label').innerText = 'Einde fragment. ' + (reelAuto ? 'Laatste fragment bereikt.' : 'Auto-play staat uit.');
            }}
          }}
        }}, 200);
      }}

      function reelPrev() {{ playFrag(reelIdx - 1); }}
      function reelNext() {{ playFrag(reelIdx + 1); }}
      function reelReplay() {{ playFrag(reelIdx); }}
      function reelSpeed(v) {{ if (reelReady) reelPlayer.setPlaybackRate(parseFloat(v)); }}
      function reelAutoToggle() {{
        reelAuto = !reelAuto;
        var b = document.getElementById('reel-auto-btn');
        if (reelAuto) {{
          b.innerText = '▶ Auto aan';
          b.style.background = '#16a34a';
          b.style.borderColor = '#16a34a';
        }} else {{
          b.innerText = '⏸ Auto uit';
          b.style.background = '#fff';
          b.style.color = '#0f172a';
          b.style.borderColor = '#cbd5e1';
        }}
      }}
    </script>
    """
    return html


def build_draw_canvas_html(image_data_url: str, canvas_id: str = "draw_canvas") -> str:
    """Bouw een canvas waarop je met pen/pijlen/lijnen/cirkels op een snapshot kan tekenen.

    image_data_url: data:image/...;base64,... — de afbeelding die als achtergrond gezet wordt.
    Je kunt tekenen met muis of touch, van kleur wisselen, en de afbeelding + tekening downloaden als PNG.
    """
    img_js = json.dumps(image_data_url or "")
    cid = canvas_id
    html = f"""
    <div style="font-family:system-ui,Segoe UI,Roboto,sans-serif;">
      <div style="display:flex;gap:6px;flex-wrap:wrap;margin-bottom:8px;align-items:center;">
        <select id="tool-{cid}" style="padding:8px 12px;border-radius:10px;border:1px solid #cbd5e1;font-weight:700;">
          <option value="pen">✏️ Pen</option>
          <option value="arrow" selected>➡️ Pijl</option>
          <option value="line">／ Lijn</option>
          <option value="circle">⭕ Cirkel</option>
        </select>
        <input type="color" id="color-{cid}" value="#ef4444" style="width:42px;height:38px;border:1px solid #cbd5e1;border-radius:10px;cursor:pointer;">
        <input type="range" id="size-{cid}" min="2" max="14" value="5" style="width:120px;">
        <button onclick="dcUndo_{cid}()" style="padding:8px 12px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;cursor:pointer;">↶ Ongedaan</button>
        <button onclick="dcClear_{cid}()" style="padding:8px 12px;border-radius:10px;border:1px solid #cbd5e1;background:#fff;color:#0f172a;font-weight:700;cursor:pointer;">🗑️ Leeg</button>
        <button onclick="dcDownload_{cid}()" style="padding:8px 12px;border-radius:10px;border:1px solid #2563eb;background:#2563eb;color:#fff;font-weight:700;cursor:pointer;">💾 Download PNG</button>
      </div>
      <canvas id="{cid}" style="width:100%;border:2px solid #cbd5e1;border-radius:14px;background:#0f172a;cursor:crosshair;touch-action:none;"></canvas>
      <div id="hint-{cid}" style="margin-top:8px;color:#475569;font-size:13px;">
        Tip: klik & sleep om te tekenen. Gebruik de toolbar boven om van pijl → lijn → cirkel → pen te wisselen.
      </div>
    </div>
    <script>
    (function(){{
      var canvas = document.getElementById("{cid}");
      var ctx = canvas.getContext('2d');
      var toolEl = document.getElementById("tool-{cid}");
      var colorEl = document.getElementById("color-{cid}");
      var sizeEl = document.getElementById("size-{cid}");

      var img = new Image();
      img.crossOrigin = "anonymous";
      var bg = {img_js};
      var strokes = [];       // lijst van afgeronde strokes
      var current = null;     // huidige actieve stroke
      var drawing = false;

      function fitCanvas() {{
        // Interne resolutie = natuurlijke image grootte (voor scherpe download)
        if (img.naturalWidth) {{
          canvas.width = img.naturalWidth;
          canvas.height = img.naturalHeight;
        }} else {{
          canvas.width = 1280;
          canvas.height = 720;
        }}
        redraw();
      }}

      img.onload = fitCanvas;
      img.onerror = function(){{
        canvas.width = 1280; canvas.height = 720;
        ctx.fillStyle = '#0f172a';
        ctx.fillRect(0,0,canvas.width,canvas.height);
        ctx.fillStyle = '#94a3b8';
        ctx.font = '28px sans-serif';
        ctx.fillText('Geen achtergrondbeeld geladen', 40, 60);
      }};
      if (bg) img.src = bg;
      else fitCanvas();

      function getPos(e) {{
        var rect = canvas.getBoundingClientRect();
        var scaleX = canvas.width / rect.width;
        var scaleY = canvas.height / rect.height;
        var x, y;
        if (e.touches && e.touches.length) {{
          x = (e.touches[0].clientX - rect.left) * scaleX;
          y = (e.touches[0].clientY - rect.top) * scaleY;
        }} else {{
          x = (e.clientX - rect.left) * scaleX;
          y = (e.clientY - rect.top) * scaleY;
        }}
        return {{x: x, y: y}};
      }}

      function drawStroke(s) {{
        ctx.save();
        ctx.strokeStyle = s.color;
        ctx.fillStyle = s.color;
        ctx.lineWidth = s.size;
        ctx.lineCap = 'round';
        ctx.lineJoin = 'round';
        if (s.tool === 'pen' && s.points.length > 1) {{
          ctx.beginPath();
          ctx.moveTo(s.points[0].x, s.points[0].y);
          for (var i=1; i<s.points.length; i++) ctx.lineTo(s.points[i].x, s.points[i].y);
          ctx.stroke();
        }} else if (s.tool === 'line' && s.points.length >= 2) {{
          var a = s.points[0], b = s.points[s.points.length-1];
          ctx.beginPath();
          ctx.moveTo(a.x, a.y);
          ctx.lineTo(b.x, b.y);
          ctx.stroke();
        }} else if (s.tool === 'arrow' && s.points.length >= 2) {{
          var a = s.points[0], b = s.points[s.points.length-1];
          ctx.beginPath();
          ctx.moveTo(a.x, a.y);
          ctx.lineTo(b.x, b.y);
          ctx.stroke();
          // Pijlpunt
          var ang = Math.atan2(b.y - a.y, b.x - a.x);
          var head = Math.max(12, s.size * 3);
          ctx.beginPath();
          ctx.moveTo(b.x, b.y);
          ctx.lineTo(b.x - head * Math.cos(ang - Math.PI/7), b.y - head * Math.sin(ang - Math.PI/7));
          ctx.lineTo(b.x - head * Math.cos(ang + Math.PI/7), b.y - head * Math.sin(ang + Math.PI/7));
          ctx.closePath();
          ctx.fill();
        }} else if (s.tool === 'circle' && s.points.length >= 2) {{
          var a = s.points[0], b = s.points[s.points.length-1];
          var r = Math.hypot(b.x - a.x, b.y - a.y);
          ctx.beginPath();
          ctx.arc(a.x, a.y, r, 0, Math.PI*2);
          ctx.stroke();
        }}
        ctx.restore();
      }}

      function redraw() {{
        ctx.clearRect(0,0,canvas.width,canvas.height);
        if (img && img.naturalWidth) {{
          ctx.drawImage(img, 0, 0, canvas.width, canvas.height);
        }} else {{
          ctx.fillStyle = '#0f172a'; ctx.fillRect(0,0,canvas.width,canvas.height);
        }}
        strokes.forEach(drawStroke);
        if (current) drawStroke(current);
      }}

      function onStart(e) {{
        e.preventDefault();
        drawing = true;
        var p = getPos(e);
        current = {{
          tool: toolEl.value,
          color: colorEl.value,
          size: parseInt(sizeEl.value || '5', 10),
          points: [p]
        }};
      }}
      function onMove(e) {{
        if (!drawing || !current) return;
        e.preventDefault();
        current.points.push(getPos(e));
        redraw();
      }}
      function onEnd(e) {{
        if (!drawing || !current) return;
        drawing = false;
        strokes.push(current);
        current = null;
        redraw();
      }}

      canvas.addEventListener('mousedown', onStart);
      canvas.addEventListener('mousemove', onMove);
      window.addEventListener('mouseup', onEnd);
      canvas.addEventListener('touchstart', onStart, {{passive:false}});
      canvas.addEventListener('touchmove', onMove, {{passive:false}});
      canvas.addEventListener('touchend', onEnd);

      window["dcUndo_{cid}"] = function() {{ strokes.pop(); redraw(); }};
      window["dcClear_{cid}"] = function() {{ strokes = []; redraw(); }};
      window["dcDownload_{cid}"] = function() {{
        try {{
          var data = canvas.toDataURL('image/png');
          var a = document.createElement('a');
          a.href = data;
          a.download = 'tactische-analyse.png';
          document.body.appendChild(a);
          a.click();
          document.body.removeChild(a);
        }} catch(err) {{
          var h = document.getElementById('hint-{cid}');
          h.innerText = '⚠️ Download mislukt (CORS). Maak een screenshot van je scherm of gebruik eerst Upload.';
        }}
      }};
    }})();
    </script>
    """
    return html


def render_draw_section() -> None:
    """Tekenen op een snapshot of geüploade screenshot.

    Werkt met:
    - Een geüploade foto (PNG/JPG) — altijd mogelijk
    - Een frame uit een geüploade mp4 — via <video> → <canvas> capture in de browser
    - YouTube: alleen via een geüploade screenshot (iframe kan niet gecaptured worden)
    """
    st.markdown("### ✏️ Tekenen op beeld (tactische analyse)")
    st.caption(
        "Teken pijlen, lijnen of cirkels op een wedstrijdbeeld. Upload een screenshot uit je video "
        "of capture een frame uit een geüploade mp4 via de knop hieronder."
    )

    uploaded = st.file_uploader(
        "Upload een screenshot of foto om op te tekenen",
        type=["png", "jpg", "jpeg"],
        key="draw_snapshot_uploader",
    )

    image_data_url = ""
    if uploaded is not None:
        import base64 as _b64
        b = uploaded.getvalue()
        ext = uploaded.name.lower().split(".")[-1]
        mime = "image/png" if ext == "png" else ("image/jpeg" if ext in ("jpg", "jpeg") else "image/png")
        b64 = _b64.b64encode(b).decode("ascii")
        image_data_url = f"data:{mime};base64,{b64}"
        st.success(f"Beeld geladen: {uploaded.name}")

    # Tekenen op frame uit geüploade mp4: capture via JS (alleen voor file-source)
    if not image_data_url and st.session_state.get("video_source_type") in ("file", "url"):
        st.caption(
            "💡 Geen screenshot geüpload? Je kunt een frame capturen uit de huidige video "
            "(werkt voor uploads en directe mp4-links)."
        )
        frame_time = st.number_input(
            "Frame-tijd (seconden video)",
            min_value=0.0,
            value=float(st.session_state.get("current_video_time") or 0.0),
            step=0.5,
            key="draw_frame_time",
        )
        if st.button("📸 Capture frame uit video", key="draw_capture_btn"):
            vurl_js = json.dumps(st.session_state.get("video_url", ""))
            t_js = f"{float(frame_time):.2f}"
            cap_html = f"""
            <div style="font-family:system-ui,Segoe UI,Roboto,sans-serif;">
              <video id="cap_v" src={vurl_js} crossorigin="anonymous" style="width:100%;max-height:300px;background:#000;border-radius:10px;"></video>
              <canvas id="cap_c" style="display:none;"></canvas>
              <div id="cap_out" style="margin-top:10px;">Frame wordt geladen…</div>
              <script>
              (function(){{
                var v = document.getElementById('cap_v');
                var c = document.getElementById('cap_c');
                var out = document.getElementById('cap_out');
                v.addEventListener('loadedmetadata', function(){{ v.currentTime = {t_js}; }});
                v.addEventListener('seeked', function(){{
                  try {{
                    c.width = v.videoWidth; c.height = v.videoHeight;
                    c.getContext('2d').drawImage(v, 0, 0);
                    var data = c.toDataURL('image/png');
                    var img = new Image();
                    img.src = data;
                    img.style.width = '100%';
                    img.style.borderRadius = '10px';
                    out.innerHTML = '<b>Frame gecaptured.</b> Klik rechtsboven op de afbeelding → "Opslaan als" om m lokaal op te slaan en daarna hieronder te uploaden om op te tekenen.';
                    out.appendChild(img);
                  }} catch(err) {{
                    out.innerText = '⚠️ Kan frame niet capturen (CORS-beperking). Maak een screenshot van je scherm en upload die hierboven.';
                  }}
                }});
              }})();
              </script>
            </div>
            """
            components.html(cap_html, height=480)

    if not image_data_url:
        st.info("Upload eerst een screenshot of capture een frame om te beginnen met tekenen.")
        return

    components.html(build_draw_canvas_html(image_data_url), height=760)
    st.caption("Na tekenen: klik '💾 Download PNG' om het opgeslagen beeld te downloaden. Je kunt het daarna delen of koppelen aan een clip.")


def render_highlight_reel_section(df: pd.DataFrame, video_url: str, source_type: str) -> None:
    """Bouw een reel op basis van gefilterde event-tags met videotijd en speel ze achter elkaar af."""
    st.markdown("### ⭐ Highlight-reel")
    st.caption(
        "Filter tags op event-type, team of kwart en speel ze automatisch achter elkaar af. "
        "Werkt het best met een YouTube-link."
    )

    if df.empty:
        st.info("Nog geen tags geregistreerd. Voeg eerst tags toe op het LIVE-scherm of via 'Tag op video-moment'.")
        return

    # Alleen tags met een videotijd (direct of afgeleid via push-off) zijn bruikbaar in een reel
    working = df.copy()
    def _resolve_vtime(row):
        vt = row.get("video_time_sec", None)
        try:
            if vt is not None and pd.notna(vt):
                return float(vt)
        except Exception:
            pass
        game_sec = parse_mmss(str(row.get("time", "00:00")))
        return video_time_from_game(game_sec, str(row.get("quarter", "Q1")))

    working["reel_vtime"] = working.apply(_resolve_vtime, axis=1)
    working = working[working["reel_vtime"].notna()]

    if working.empty:
        st.warning("Geen tags met videotijd gevonden. Kalibreer eerst de push-off per kwart, of tag rechtstreeks op video-moment.")
        return

    c1, c2, c3 = st.columns(3)
    with c1:
        r_event = st.multiselect(
            "Events",
            sorted(working["event"].dropna().unique().tolist()),
            key="reel_event_filter",
            help="Kies bijv. 'Goal' voor alle goals, of 'Strafcorner' voor alle SC's.",
        )
    with c2:
        teams = ["Alles", st.session_state.team_name, st.session_state.opponent_name]
        r_team = st.selectbox("Team", teams, key="reel_team_filter")
    with c3:
        r_q = st.selectbox("Kwart", ["Alles"] + QUARTERS, key="reel_q_filter")

    c4, c5 = st.columns(2)
    with c4:
        pre_roll = st.slider("Pre-roll (sec)", 0, 15, 4, 1, key="reel_pre")
    with c5:
        post_roll = st.slider("Post-roll (sec)", 0, 20, 5, 1, key="reel_post")

    fdf = working.copy()
    if r_event:
        fdf = fdf[fdf["event"].isin(r_event)]
    if r_team != "Alles":
        fdf = fdf[fdf["team"] == r_team]
    if r_q != "Alles":
        fdf = fdf[fdf["quarter"] == r_q]

    fdf = fdf.sort_values(["quarter", "reel_vtime"])
    st.caption(f"🎯 {len(fdf)} fragmenten geselecteerd.")

    if fdf.empty:
        st.info("Geen tags matchen de filters. Pas de filters aan of tag meer events.")
        return

    # Preview-lijst
    with st.expander("Bekijk de fragmentenlijst"):
        for _, row in fdf.iterrows():
            zone_txt = f" • {row['zone']}" if str(row.get("zone", "")).strip() else ""
            st.markdown(
                f"- **{row['quarter']} {row['time']}** • {row['team']} • {row['event']}{zone_txt} "
                f"<small style='color:{TEXT_SUB};'>(video {float(row['reel_vtime']):.1f}s)</small>",
                unsafe_allow_html=True,
            )

    # Speel-knop
    play_col, share_col = st.columns([1, 1])
    with play_col:
        play_reel = st.button("▶ Speel highlight-reel", use_container_width=True, key="reel_play_btn", type="primary")
    with share_col:
        if source_type == "youtube" and video_url:
            yt_id = extract_youtube_id(video_url)
            if yt_id and st.button("🔗 Deel-links genereren", use_container_width=True, key="reel_share_btn"):
                st.session_state.reel_share_links = [
                    {
                        "label": f"{row['quarter']} {row['time']} • {row['team']} • {row['event']}",
                        "url": build_youtube_share_url(
                            yt_id,
                            max(0.0, float(row["reel_vtime"]) - pre_roll),
                            float(row["reel_vtime"]) + post_roll,
                        ),
                    }
                    for _, row in fdf.iterrows()
                ]

    # Toon reel-speler zodra gevraagd
    if play_reel:
        st.session_state.reel_active = True
        st.session_state.reel_fragments = [
            {
                "start_sec": float(row["reel_vtime"]),
                "label": f"{row['quarter']} {row['time']} • {row['team']} • {row['event']}",
            }
            for _, row in fdf.iterrows()
        ]

    if st.session_state.get("reel_active") and st.session_state.get("reel_fragments"):
        if source_type == "youtube" and video_url:
            yt_id = extract_youtube_id(video_url)
            if yt_id:
                html = build_highlight_reel_html(
                    yt_id,
                    st.session_state.reel_fragments,
                    pre_roll=float(pre_roll),
                    post_roll=float(post_roll),
                )
                components.html(html, height=560)
                if st.button("Sluit reel-speler", key="reel_close_btn"):
                    st.session_state.reel_active = False
                    st.rerun()
            else:
                st.error("YouTube-ID kon niet worden gelezen uit de link.")
        else:
            st.warning(
                "De auto-reel werkt alleen met YouTube-links. Voor uploads / mp4-links: gebruik de 'Spring'-knoppen "
                "naast de tags om fragmenten handmatig te bekijken."
            )

    # Deel-links lijst
    if st.session_state.get("reel_share_links"):
        st.markdown("#### 🔗 Deelbare fragmentlinks")
        st.caption("Kopieer een link en stuur hem naar een speler — de link opent YouTube op het juiste moment en stopt automatisch.")
        for i, item in enumerate(st.session_state.reel_share_links):
            lc1, lc2 = st.columns([3, 1])
            with lc1:
                st.markdown(f"**{item['label']}**")
                st.code(item["url"], language=None)
            with lc2:
                st.write("")
        if st.button("Verberg deel-links", key="reel_share_clear"):
            st.session_state.reel_share_links = []
            st.rerun()


def render_clip_section(df: pd.DataFrame, active_video_name: str) -> None:
    """Clips toevoegen, tonen, exporteren."""
    st.markdown("### 🎬 Clips")

    f1, f2, f3 = st.columns(3)
    with f1:
        clip_title = st.text_input("Clipnaam", placeholder="Bijv. pressmoment links", key="clip_title_input")
    with f2:
        clip_tag = st.selectbox("Thema", VIDEO_TAGS, key="clip_tag_select")
    with f3:
        clip_team_focus = st.selectbox(
            "Focus",
            [st.session_state.team_name, st.session_state.opponent_name, "Algemeen"],
            key="clip_team_focus_select",
        )
    t1, t2, t3 = st.columns(3)
    with t1:
        clip_quarter = st.selectbox("Kwart", QUARTERS, key="clip_quarter_select")
    with t2:
        clip_start = st.number_input("Start (seconden video)", min_value=0, value=0, step=1, key="clip_start_input")
    with t3:
        clip_end = st.number_input("Einde (seconden video)", min_value=0, value=10, step=1, key="clip_end_input")
    st.caption(f"Clipbereik: {format_seconds_to_mmss(clip_start)} - {format_seconds_to_mmss(clip_end)}")

    snapshot_file = st.file_uploader(
        "Upload eventueel een screenshot van dit moment",
        type=["png", "jpg", "jpeg"],
        key="clip_snapshot_uploader",
    )
    tactical_note = st.text_area(
        "Tactische analyse",
        placeholder="Wat gebeurt hier tactisch?",
        height=120,
        key="clip_tactical_note",
    )
    coaching_action = st.text_area(
        "Coachactie / leerpunt",
        placeholder="Wat wil je hier coachen of meenemen naar training?",
        height=100,
        key="clip_coaching_action",
    )
    b1, b2, b3 = st.columns(3)
    with b1:
        if st.button("💾  Clip opslaan", use_container_width=True, key="clip_save_btn", type="primary"):
            if clip_end < clip_start:
                st.error("Eindtijd moet gelijk aan of later zijn dan starttijd.")
            else:
                add_video_clip(
                    video_name=active_video_name or "geen video",
                    clip_title=clip_title,
                    tag=clip_tag,
                    team_focus=clip_team_focus,
                    quarter=clip_quarter,
                    start_sec=clip_start,
                    end_sec=clip_end,
                    tactical_note=tactical_note,
                    coaching_action=coaching_action,
                    snapshot_name=snapshot_file.name if snapshot_file else "",
                )
                st.success("Clip opgeslagen.")
                st.rerun()
    with b2:
        if st.button("🗑️  Verwijder laatste clip", use_container_width=True, key="clip_remove_last_btn"):
            remove_last_clip()
            st.rerun()
    with b3:
        if st.button("⏩  Naar clipstart", use_container_width=True, key="clip_jump_start_btn"):
            st.session_state.jump_to_video_time = float(clip_start)
            st.rerun()

    if snapshot_file is not None:
        st.markdown("#### Screenshot")
        st.image(snapshot_file, use_container_width=True)

    clips_df = build_clips_df()
    st.markdown("#### Overzicht")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Aantal clips", len(clips_df))
    c2.metric("Press clips", int((clips_df["tag"] == "Press").sum()) if not clips_df.empty else 0)
    c3.metric("Leerclips", int((clips_df["tag"] == "Leerclip").sum()) if not clips_df.empty else 0)
    c4.metric("Positieve clips", int((clips_df["tag"] == "Positief voorbeeld").sum()) if not clips_df.empty else 0)

    st.markdown("#### Cliplog (klik ▶️ om te bekijken)")
    if clips_df.empty:
        st.info("Nog geen clips toegevoegd.")
        return

    # YouTube-ID voor deel-links (alleen bruikbaar als de huidige bron een YouTube-link is)
    current_yt_id = ""
    if st.session_state.get("video_source_type") == "youtube":
        current_yt_id = extract_youtube_id(st.session_state.get("video_url", ""))

    ordered = clips_df.sort_values(["quarter", "start_sec"])
    for _, row in ordered.iterrows():
        col1, col2, col3 = st.columns([5, 1, 1])
        with col1:
            note = row.get('tactical_note','') or ''
            st.markdown(
                f"<div style='padding:10px 14px;background:{CARD_BG_ELEVATED};"
                f"border:1px solid {CARD_BORDER_SOFT};border-left:4px solid {SUCCESS_GREEN};"
                f"border-radius:8px;margin-bottom:6px;'>"
                f"<strong style='color:{TEXT_MAIN};'>{row['quarter']} • {row['start_time']} - {row['end_time']}</strong> "
                f"<span style='color:{TEXT_SUB};'>•</span> "
                f"<span style='color:{TEXT_MAIN};'>{row['clip_title']}</span> "
                f"<span style='color:{TEXT_SUB};'>—</span> <em style='color:{ACCENT_SOFT};'>{row['tag']}</em>"
                + (f"<br><small style='color:{TEXT_SUB};'>{note}</small>" if note else "")
                + "</div>",
                unsafe_allow_html=True,
            )
        with col2:
            if st.button("▶️ Bekijk clip", key=f"clip_jump_{row['id']}", use_container_width=True):
                st.session_state.jump_to_video_time = float(row["start_sec"])
                st.rerun()
        with col3:
            share_disabled = not bool(current_yt_id)
            if st.button(
                "🔗 Deel clip",
                key=f"clip_share_{row['id']}",
                use_container_width=True,
                disabled=share_disabled,
                help="Beschikbaar zodra de wedstrijdvideo een YouTube-link is." if share_disabled else "Genereer deel-link voor deze clip.",
            ):
                url = build_youtube_share_url(
                    current_yt_id,
                    float(row["start_sec"]),
                    float(row["end_sec"]),
                )
                links = dict(st.session_state.get("clip_share_links") or {})
                links[row["id"]] = url
                st.session_state.clip_share_links = links
                st.rerun()

        # Toon deel-link als gegenereerd
        existing_link = (st.session_state.get("clip_share_links") or {}).get(row["id"])
        if existing_link:
            lc1, lc2 = st.columns([6, 1])
            with lc1:
                st.code(existing_link, language=None)
            with lc2:
                if st.button("✖", key=f"clip_share_clear_{row['id']}", help="Verberg link"):
                    links = dict(st.session_state.get("clip_share_links") or {})
                    links.pop(row["id"], None)
                    st.session_state.clip_share_links = links
                    st.rerun()

    st.markdown("#### Automatische samenvatting")
    summary_text = generate_video_analysis_summary(clips_df)
    st.text_area("Samenvatting beeldanalyse", summary_text, height=220, key="video_summary_area")
    e1, e2, e3 = st.columns(3)
    with e1:
        st.download_button("Download TXT beeldanalyse", data=summary_text.encode("utf-8"), file_name="beeldanalyse.txt", key="dl_video_txt")
    with e2:
        st.download_button(
            "Download PDF beeldanalyse",
            data=export_pdf_report(summary_text),
            file_name="beeldanalyse.pdf",
            mime="application/pdf" if REPORTLAB_AVAILABLE else "text/plain",
            key="dl_video_pdf",
        )
    with e3:
        st.download_button(
            "Download Excel beeldanalyse",
            data=export_video_analysis_excel(clips_df),
            file_name="beeldanalyse.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_video_excel",
        )


# ==================================================
# AUTO SYNC
# ==================================================
@st.fragment(run_every="2s" if cloud_enabled() else None)
def auto_sync_cloud():
    if cloud_enabled() and st.session_state.match_id:
        fresh = load_events_from_cloud(st.session_state.match_id)
        if len(fresh) != st.session_state.last_sync_count:
            st.session_state.events = fresh
            refresh_derived_state()
        st.session_state.last_sync_count = len(fresh)
        st.session_state.last_sync_time = time.strftime("%H:%M:%S")


# ==================================================
# WISSELSCHEMA MODULE
# ==================================================
SUBS_LINE_NAMES = {"V": "Achter (Verdediging)", "M": "Midden", "A": "Voor (Aanval)", "K": "Keeper"}
SUBS_LINE_ORDER = ["A", "M", "V", "K"]

# Defaults: worden gebruikt als het team nog geen eigen formaties heeft.
# Zodra het team zelf formaties aanmaakt in de cloud, worden deze vervangen.
SUBS_FORMATIONS_DEFAULT = {
    "3-3-3-1": {"V": 3, "M": 3, "A": 3},   # Zaalhockey-achtig (keeper apart)
    "3-3-3":   {"V": 3, "M": 3, "A": 3},
    "2-3-3-2": {"V": 2, "M": 3, "A": 3},
    "4-3-3":   {"V": 4, "M": 3, "A": 3},
    "3-4-3":   {"V": 3, "M": 4, "A": 3},
    "4-4-2":   {"V": 4, "M": 4, "A": 2},
}


def _subs_active_formations() -> dict:
    """Retourneer dict {naam: {V,M,A}} van de formaties voor het actieve team.

    Voorkeur: team_formations uit de cloud (custom per team).
    Fallback: SUBS_FORMATIONS_DEFAULT.
    """
    formations = st.session_state.get("subs_formations", [])
    if formations:
        return {f["name"]: f["slots"] for f in formations}
    return dict(SUBS_FORMATIONS_DEFAULT)


def _subs_formation_slots(name: str) -> dict:
    """Slots (V,M,A) voor een formatie-naam; valt terug op 4-3-3 als hij onbekend is."""
    active = _subs_active_formations()
    if name in active:
        return active[name]
    # Fallback: eerste beschikbare formatie, anders 4-3-3
    if active:
        first = next(iter(active.values()))
        return first
    return {"V": 4, "M": 3, "A": 3}


def _subs_init_state() -> None:
    """Initialize session state voor wisselschema.
    Laadt spelers, match-settings en aanwezigheid automatisch uit Supabase
    voor het actieve team, zodat je na inloggen meteen je team-data ziet.
    """
    defaults = {
        "subs_players": [],  # [{id, name, line, can_keep, priority}]
        "subs_attendance": {},  # {player_id: bool}
        "subs_match": {
            "opponent": "",
            "match_date": time.strftime("%Y-%m-%d"),
            "half_length": 17,
            "halves": 2,
            "formation": "3-3-3-1",
            "fixed_keeper_id": "",
            "block_size": 4,
        },
        "subs_schema": None,  # {cells: {player_id: [0,1,...]}, minutes_per_half, halves}
        "subs_warnings": [],
        "subs_formations": [],  # [{id, name, slots, is_default}]
        "subs_cloud_loaded_for": None,  # team_id waarvoor laatst is geladen
        "subs_linked_match_id": None,   # match_id waarvoor lineup is geladen
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

    # Laad cloud-data als we nog niets hebben geladen voor dit team
    active_tid = _active_team_id()
    if active_tid and st.session_state.get("subs_cloud_loaded_for") != active_tid:
        cloud_players = cloud_load_team_players()
        cloud_settings = cloud_load_match_settings()
        cloud_att = cloud_load_attendance()
        cloud_forms = cloud_load_formations()
        if cloud_players:
            st.session_state.subs_players = cloud_players
        if cloud_forms:
            st.session_state.subs_formations = cloud_forms
            # Als team een default-formatie heeft en huidig subs_match.formation is niet geldig, switch
            default_f = next((f for f in cloud_forms if f.get("is_default")), None)
            names = {f["name"] for f in cloud_forms}
            if st.session_state.subs_match.get("formation") not in names:
                if default_f:
                    st.session_state.subs_match["formation"] = default_f["name"]
                elif cloud_forms:
                    st.session_state.subs_match["formation"] = cloud_forms[0]["name"]
        if cloud_settings:
            # Merge: behoud bestaande defaults, overschrijf met cloud-waarden
            st.session_state.subs_match.update({
                k: v for k, v in cloud_settings.items() if v not in (None, "")
            })
            # Behoud lege strings expliciet voor bv. fixed_keeper_id
            if "fixed_keeper_id" in cloud_settings:
                st.session_state.subs_match["fixed_keeper_id"] = cloud_settings["fixed_keeper_id"] or ""
        # Aanwezigheid: initialiseer alle geladen spelers op True, dan overschrijf met cloud-waarden
        att = {p["id"]: True for p in st.session_state.subs_players}
        att.update(cloud_att)
        st.session_state.subs_attendance = att
        st.session_state.subs_cloud_loaded_for = active_tid

    # Synchroniseer met huidige geladen wedstrijd: laad lineup als match_id is gewisseld
    current_mid = st.session_state.get("match_id")
    if current_mid and st.session_state.get("subs_linked_match_id") != current_mid:
        lineup = cloud_load_match_lineup(current_mid)
        if lineup:
            if lineup.get("schema_json"):
                st.session_state.subs_schema = lineup["schema_json"]
            if lineup.get("settings_json"):
                # Merge alleen ingevulde velden, maar overschrijf formatie/block_size/etc.
                for k, v in lineup["settings_json"].items():
                    if v is not None:
                        st.session_state.subs_match[k] = v
        st.session_state.subs_linked_match_id = current_mid


def _subs_prio_weight(priority: str) -> float:
    return {"high": 1.3, "low": 0.7}.get(priority or "normal", 1.0)


def _subs_generate_schema() -> None:
    """Genereer schema op basis van spelers, aanwezigheid en match-instellingen."""
    players = [p for p in st.session_state.subs_players
               if st.session_state.subs_attendance.get(p["id"], True)]
    m = st.session_state.subs_match
    slots = _subs_formation_slots(m["formation"])
    minutes_per_half = int(m["half_length"])
    halves = int(m["halves"])
    total_minutes = minutes_per_half * halves
    block_size = int(m.get("block_size", 4))
    warnings = []

    if not players:
        st.session_state.subs_warnings = ["Geen aanwezige spelers."]
        return

    # Groepeer per linie
    by_line = {"V": [], "M": [], "A": [], "K": []}
    for p in players:
        by_line[p.get("line", "M")].append(p)

    # Vaste keeper
    fixed_keeper = None
    if m.get("fixed_keeper_id"):
        fixed_keeper = next((p for p in players if p["id"] == m["fixed_keeper_id"]), None)
    if not fixed_keeper:
        candidates = [p for p in players if p.get("can_keep") or p.get("line") == "K"]
        if not candidates:
            st.session_state.subs_warnings = ["Geen keeper aanwezig. Markeer minstens één speler als keeper."]
            return
        fixed_keeper = candidates[0]
        warnings.append(f'Geen vaste keeper gekozen — "{fixed_keeper["name"]}" wordt gebruikt.')

    # Initialiseer cellen
    cells = {p["id"]: [0] * total_minutes for p in players}
    for t in range(total_minutes):
        cells[fixed_keeper["id"]][t] = 1

    # Vul veldlinies
    for L in ("V", "M", "A"):
        line_slots = slots[L]
        candidates = [p for p in by_line[L] if p["id"] != fixed_keeper["id"]]
        if not candidates:
            warnings.append(f"Geen spelers in linie {SUBS_LINE_NAMES[L]}.")
            continue
        if len(candidates) < line_slots:
            warnings.append(
                f"Te weinig spelers in {SUBS_LINE_NAMES[L]}: {len(candidates)} beschikbaar voor {line_slots} plekken."
            )

        total_line_min = line_slots * total_minutes
        sum_w = sum(_subs_prio_weight(p.get("priority")) for p in candidates)
        targets = [{
            "id": p["id"],
            "name": p["name"],
            "target": (_subs_prio_weight(p.get("priority")) * total_line_min) / sum_w,
            "floor": 0,
            "target_min": 0,
        } for p in candidates]

        for t in targets:
            t["floor"] = int(t["target"])
        assigned = sum(t["floor"] for t in targets)
        rem = total_line_min - assigned
        targets.sort(key=lambda t: (t["target"] - t["floor"]), reverse=True)
        for i in range(len(targets)):
            if rem <= 0:
                break
            targets[i]["floor"] += 1
            rem -= 1
        for t in targets:
            t["target_min"] = min(t["floor"], total_minutes)
        overflow = sum(t["floor"] for t in targets) - sum(t["target_min"] for t in targets)
        targets.sort(key=lambda t: t["target_min"])
        for i in range(len(targets)):
            if overflow <= 0:
                break
            room = total_minutes - targets[i]["target_min"]
            add = min(room, overflow)
            targets[i]["target_min"] += add
            overflow -= add

        rem_map = {t["id"]: t["target_min"] for t in targets}

        # Blok-gebaseerd invullen
        for h in range(halves):
            half_start = h * minutes_per_half
            half_end = half_start + minutes_per_half
            mm = half_start
            while mm < half_end:
                block_len = min(block_size, half_end - mm)
                last_block_players = set()
                if mm > half_start:
                    for t in targets:
                        if cells[t["id"]][mm - 1]:
                            last_block_players.add(t["id"])
                scored = [(t["id"], rem_map[t["id"]], t["id"] in last_block_players) for t in targets]
                # Sorteer: hoogste remaining eerst, bij gelijk wie vorig blok NIET speelde eerst
                scored.sort(key=lambda s: (-s[1], s[2]))
                chosen = [s[0] for s in scored if rem_map[s[0]] > 0][:line_slots]
                if len(chosen) < line_slots:
                    extras = [s[0] for s in scored if s[0] not in chosen][:line_slots - len(chosen)]
                    chosen += extras
                for t_idx in range(mm, mm + block_len):
                    for pid in chosen:
                        cells[pid][t_idx] = 1
                        rem_map[pid] = max(0, rem_map[pid] - 1)
                mm += block_len

    st.session_state.subs_schema = {
        "cells": cells,
        "minutes_per_half": minutes_per_half,
        "halves": halves,
        "total_minutes": total_minutes,
    }
    st.session_state.subs_warnings = warnings


def _subs_render_schema_table(line: str, players_in_line: list, cells: dict,
                              minutes_per_half: int, halves: int, required: int) -> None:
    """Render één linie-tabel met Plotly-like HTML table."""
    total_minutes = minutes_per_half * halves
    # HTML bouwen
    html_parts = []
    colors = {"V": "#3b82f6", "M": "#10b981", "A": "#ef4444", "K": "#f59e0b"}
    header_color = colors.get(line, "#6366f1")
    html_parts.append(
        f'<div style="margin-bottom:18px; background:#1e2639; border:1px solid #2a3448; border-radius:10px; overflow:hidden;">'
    )
    html_parts.append(
        f'<div style="padding:10px 14px; font-weight:700; color:white; font-size:14px; '
        f'text-transform:uppercase; letter-spacing:1px; background:{header_color};">'
        f'{SUBS_LINE_NAMES[line]} — {len(players_in_line)} spelers, {required} in het veld</div>'
    )
    html_parts.append(
        '<div style="overflow-x:auto;"><table style="width:100%; border-collapse:collapse; font-size:11px;">'
    )
    # Header
    html_parts.append('<thead><tr>')
    html_parts.append(
        '<th style="border:1px solid #2a3448; padding:6px 10px; background:#161c2b; color:#e7ecf5; text-align:left; min-width:120px;">Speler</th>'
    )
    for h in range(halves):
        for m in range(minutes_per_half):
            divider = "border-left:3px solid #6366f1;" if (h > 0 and m == 0) else ""
            html_parts.append(
                f'<th style="border:1px solid #2a3448; padding:4px; background:#161c2b; color:#8a96ad; min-width:24px; {divider}">{m + 1}</th>'
            )
    html_parts.append(
        '<th style="border:1px solid #2a3448; padding:4px 8px; background:#6366f1; color:white;">Half</th>'
    )
    html_parts.append(
        '<th style="border:1px solid #2a3448; padding:4px 8px; background:#6366f1; color:white;">Vol</th>'
    )
    html_parts.append('</tr></thead><tbody>')

    # Spelerrijen
    for p in players_in_line:
        highlight = p.get("priority") == "high"
        player_cells = cells.get(p["id"], [0] * total_minutes)
        html_parts.append('<tr>')
        star = " ⭐" if highlight else ""
        html_parts.append(
            f'<td style="border:1px solid #2a3448; padding:6px 10px; background:#161c2b; '
            f'color:#e7ecf5; font-weight:600;">{p["name"]}{star}</td>'
        )
        h1 = 0
        full = 0
        for h in range(halves):
            for m in range(minutes_per_half):
                idx = h * minutes_per_half + m
                val = player_cells[idx] if idx < len(player_cells) else 0
                divider = "border-left:3px solid #6366f1;" if (h > 0 and m == 0) else ""
                if val:
                    full += 1
                    if idx < minutes_per_half:
                        h1 += 1
                    if highlight:
                        cell_style = "background:#facc15; color:#111;"
                    else:
                        cell_style = "background:#10b981; color:white;"
                    cell_content = "1"
                else:
                    cell_style = "background:#1e2639; color:#8a96ad;"
                    cell_content = ""
                html_parts.append(
                    f'<td style="border:1px solid #2a3448; text-align:center; '
                    f'font-weight:700; height:28px; {cell_style} {divider}">{cell_content}</td>'
                )
        html_parts.append(
            f'<td style="border:1px solid #2a3448; padding:4px 8px; background:#161c2b; '
            f'color:#e7ecf5; text-align:center; font-weight:700;">{h1}</td>'
        )
        html_parts.append(
            f'<td style="border:1px solid #2a3448; padding:4px 8px; background:#161c2b; '
            f'color:#e7ecf5; text-align:center; font-weight:700;">{full}</td>'
        )
        html_parts.append('</tr>')

    # Capaciteit-rij
    html_parts.append('<tr>')
    html_parts.append(
        '<td style="border:1px solid #2a3448; padding:6px 10px; background:#161c2b; '
        'color:#8a96ad; font-weight:700; font-size:10px; text-transform:uppercase;">In veld</td>'
    )
    for h in range(halves):
        for m in range(minutes_per_half):
            idx = h * minutes_per_half + m
            count = sum(1 for p in players_in_line
                        if idx < len(cells.get(p["id"], [])) and cells[p["id"]][idx])
            divider = "border-left:3px solid #6366f1;" if (h > 0 and m == 0) else ""
            if count > required:
                bg = "#ef4444"; col = "white"
            elif count < required:
                bg = "#f59e0b"; col = "#111"
            else:
                bg = "#10b981"; col = "white"
            html_parts.append(
                f'<td style="border:1px solid #2a3448; text-align:center; font-weight:700; '
                f'font-size:10px; height:24px; background:{bg}; color:{col}; {divider}">{count}</td>'
            )
    html_parts.append(
        '<td style="border:1px solid #2a3448; background:#161c2b; color:#8a96ad;">—</td>'
    )
    html_parts.append(
        '<td style="border:1px solid #2a3448; background:#161c2b; color:#8a96ad;">—</td>'
    )
    html_parts.append('</tr>')
    html_parts.append('</tbody></table></div></div>')
    st.markdown("".join(html_parts), unsafe_allow_html=True)


def render_substitution_screen() -> None:
    """Hoofdscreen voor wisselschema."""
    _subs_init_state()
    st.markdown("### ⇄ Wisselschema")
    st.caption(
        "Maak een wisselschema per minuut met automatische verdeling per linie. "
        "Markeer 'meer speeltijd' voor A-spelers; het schema rekent de minuten evenredig."
    )

    # Laat zien aan welke wedstrijd dit schema gekoppeld is (belangrijk voor cloud-opslag)
    linked_mid = st.session_state.get("match_id") or "—"
    team_label = st.session_state.get("team_name", "Ons team")
    if linked_mid != "—":
        opp = st.session_state.get("opponent_name") or "Tegenstander"
        st.caption(f"Gekoppeld aan **{team_label} – {opp}** (`{linked_mid}`). "
                   f"Het schema wordt automatisch per wedstrijd opgeslagen.")
    else:
        st.caption("Nog geen wedstrijd geladen — kies eerst een wedstrijd via het startscherm "
                   "zodat het schema bij de juiste wedstrijd wordt opgeslagen.")

    tabs = st.tabs(["1. Team", "2. Formaties", "3. Wedstrijd", "4. Schema"])

    # ----- TAB 1: TEAM -----
    with tabs[0]:
        tc_left, tc_right = st.columns([3, 1])
        with tc_right:
            if st.button("🗑 Team leegmaken", use_container_width=True, key="subs_clear_team",
                         help="Verwijder alle spelers van dit team (kan niet ongedaan worden)"):
                cloud_clear_players()
                st.session_state.subs_players = []
                st.session_state.subs_attendance = {}
                st.session_state.subs_schema = None
                st.rerun()

        st.markdown("#### Speler toevoegen")
        fc1, fc2, fc3, fc4, fc5 = st.columns([2, 1.2, 1, 1.2, 0.8])
        with fc1:
            new_name = st.text_input("Naam", key="subs_new_name", placeholder="Bijv. Puk")
        with fc2:
            new_line = st.selectbox("Linie", ["A", "M", "V", "K"],
                                    format_func=lambda x: SUBS_LINE_NAMES[x], key="subs_new_line")
        with fc3:
            new_keep = st.selectbox("Keeper?", [False, True],
                                    format_func=lambda x: "Ja" if x else "Nee", key="subs_new_keep")
        with fc4:
            new_prio = st.selectbox("Speeltijd", ["normal", "high", "low"],
                                    format_func=lambda x: {"normal": "Normaal",
                                                           "high": "Meer (A-speler) ⭐",
                                                           "low": "Minder"}[x],
                                    key="subs_new_prio")
        with fc5:
            st.write("")
            if st.button("+ Voeg toe", use_container_width=True, key="subs_add_player"):
                if new_name.strip():
                    pid = f"p_{uuid.uuid4().hex[:10]}"
                    player = {
                        "id": pid, "name": new_name.strip(), "line": new_line,
                        "can_keep": bool(new_keep) or new_line == "K",
                        "priority": new_prio,
                    }
                    st.session_state.subs_players.append(player)
                    st.session_state.subs_attendance[pid] = True
                    cloud_upsert_player(player)
                    cloud_save_attendance(pid, True)
                    st.rerun()

        st.markdown(f"#### Huidige selectie ({len(st.session_state.subs_players)} spelers)")
        if not st.session_state.subs_players:
            st.info("Nog geen spelers toegevoegd. Voeg hierboven spelers toe; ze worden automatisch "
                    "opgeslagen bij jouw team en blijven de volgende keer beschikbaar.")
        else:
            sorted_players = sorted(st.session_state.subs_players,
                                    key=lambda p: (SUBS_LINE_ORDER.index(p.get("line", "M")), p["name"]))
            for p in sorted_players:
                lc1, lc2, lc3, lc4, lc5 = st.columns([2, 1.2, 1, 1.2, 0.5])
                with lc1:
                    star = " ⭐" if p.get("priority") == "high" else ""
                    down = " ↓" if p.get("priority") == "low" else ""
                    st.markdown(f"**{p['name']}**{star}{down}  `{SUBS_LINE_NAMES[p.get('line', 'M')]}`")
                with lc2:
                    new_line_val = st.selectbox("Linie", ["A", "M", "V", "K"],
                                                index=["A", "M", "V", "K"].index(p.get("line", "M")),
                                                format_func=lambda x: SUBS_LINE_NAMES[x],
                                                key=f"subs_line_{p['id']}", label_visibility="collapsed")
                    if new_line_val != p.get("line"):
                        p["line"] = new_line_val
                        if new_line_val == "K":
                            p["can_keep"] = True
                        cloud_upsert_player(p)
                        st.rerun()
                with lc3:
                    new_keep_val = st.selectbox("Keeper", [False, True],
                                                index=1 if p.get("can_keep") else 0,
                                                format_func=lambda x: "Keeper" if x else "—",
                                                key=f"subs_keep_{p['id']}", label_visibility="collapsed")
                    if new_keep_val != p.get("can_keep"):
                        p["can_keep"] = new_keep_val
                        cloud_upsert_player(p)
                        st.rerun()
                with lc4:
                    new_prio_val = st.selectbox("Speeltijd", ["normal", "high", "low"],
                                                index=["normal", "high", "low"].index(p.get("priority", "normal")),
                                                format_func=lambda x: {"normal": "Normaal",
                                                                       "high": "Meer ⭐",
                                                                       "low": "Minder"}[x],
                                                key=f"subs_prio_{p['id']}", label_visibility="collapsed")
                    if new_prio_val != p.get("priority"):
                        p["priority"] = new_prio_val
                        cloud_upsert_player(p)
                        st.rerun()
                with lc5:
                    if st.button("🗑", key=f"subs_del_{p['id']}", use_container_width=True):
                        cloud_delete_player(p["id"])
                        st.session_state.subs_players = [
                            x for x in st.session_state.subs_players if x["id"] != p["id"]
                        ]
                        st.session_state.subs_attendance.pop(p["id"], None)
                        st.rerun()

    # ----- TAB 2: FORMATIES -----
    with tabs[1]:
        st.markdown("#### Formaties van dit team")
        st.caption(
            "Voeg je eigen formaties toe: kies een naam en geef per linie aan hoeveel spelers "
            "er in het veld staan. Het totaal zonder keeper moet 10 zijn (11 inclusief keeper). "
            "Je kunt er één als standaard markeren — die wordt voorgesteld bij een nieuwe wedstrijd."
        )

        forms = st.session_state.get("subs_formations", [])

        # ---- Nieuwe formatie toevoegen ----
        with st.expander("➕ Nieuwe formatie toevoegen", expanded=not forms):
            fn_cols = st.columns([2, 1, 1, 1, 1])
            with fn_cols[0]:
                new_f_name = st.text_input("Naam", key="subs_new_form_name",
                                           placeholder="Bijv. 3-3-3-1 thuis")
            with fn_cols[1]:
                new_f_v = st.number_input("Achter (V)", min_value=0, max_value=10, value=3,
                                          step=1, key="subs_new_form_v")
            with fn_cols[2]:
                new_f_m = st.number_input("Midden (M)", min_value=0, max_value=10, value=3,
                                          step=1, key="subs_new_form_m")
            with fn_cols[3]:
                new_f_a = st.number_input("Voor (A)", min_value=0, max_value=10, value=3,
                                          step=1, key="subs_new_form_a")
            with fn_cols[4]:
                st.write("")
                if st.button("Toevoegen", key="subs_add_formation", use_container_width=True):
                    total = int(new_f_v) + int(new_f_m) + int(new_f_a)
                    name = (new_f_name or "").strip()
                    if not name:
                        st.warning("Vul een naam in.")
                    elif any(f["name"].lower() == name.lower() for f in forms):
                        st.warning("Er bestaat al een formatie met die naam.")
                    else:
                        new_formation = {
                            "id": str(uuid.uuid4()),
                            "name": name,
                            "slots": {"V": int(new_f_v), "M": int(new_f_m), "A": int(new_f_a)},
                            "is_default": len(forms) == 0,
                        }
                        cloud_upsert_formation(new_formation)
                        if new_formation["is_default"]:
                            cloud_set_default_formation(new_formation["id"])
                        st.session_state.subs_formations = cloud_load_formations()
                        if total + 1 != 11:
                            st.info(f"Let op: veldspelers + keeper = {total + 1} (normaal is dit 11).")
                        st.rerun()

        # ---- Bestaande formaties tonen + bewerken ----
        if not forms:
            st.info("Nog geen formaties. Voeg er hierboven eentje toe.")
        else:
            st.markdown("##### Jouw formaties")
            for f in forms:
                fc = st.columns([2, 1, 1, 1, 1, 1, 0.6])
                total_field = f["slots"]["V"] + f["slots"]["M"] + f["slots"]["A"]
                with fc[0]:
                    star = " ⭐" if f.get("is_default") else ""
                    st.markdown(f"**{f['name']}**{star}  \n"
                                f"<span style='color:#9ca3af; font-size:12px'>"
                                f"{total_field} veld + 1 keeper = {total_field + 1} totaal</span>",
                                unsafe_allow_html=True)
                with fc[1]:
                    v_val = st.number_input("V", min_value=0, max_value=10,
                                            value=int(f["slots"]["V"]), step=1,
                                            key=f"subs_f_v_{f['id']}", label_visibility="collapsed")
                with fc[2]:
                    m_val = st.number_input("M", min_value=0, max_value=10,
                                            value=int(f["slots"]["M"]), step=1,
                                            key=f"subs_f_m_{f['id']}", label_visibility="collapsed")
                with fc[3]:
                    a_val = st.number_input("A", min_value=0, max_value=10,
                                            value=int(f["slots"]["A"]), step=1,
                                            key=f"subs_f_a_{f['id']}", label_visibility="collapsed")
                with fc[4]:
                    if (v_val, m_val, a_val) != (f["slots"]["V"], f["slots"]["M"], f["slots"]["A"]):
                        if st.button("💾 Opslaan", key=f"subs_f_save_{f['id']}",
                                     use_container_width=True):
                            f["slots"] = {"V": int(v_val), "M": int(m_val), "A": int(a_val)}
                            cloud_upsert_formation(f)
                            st.session_state.subs_formations = cloud_load_formations()
                            st.rerun()
                    else:
                        st.write("")
                with fc[5]:
                    if f.get("is_default"):
                        st.markdown("<div style='color:#10b981; font-weight:700; padding-top:6px'>Standaard</div>",
                                    unsafe_allow_html=True)
                    else:
                        if st.button("Als standaard", key=f"subs_f_def_{f['id']}",
                                     use_container_width=True):
                            cloud_set_default_formation(f["id"])
                            st.session_state.subs_formations = cloud_load_formations()
                            st.rerun()
                with fc[6]:
                    if st.button("🗑", key=f"subs_f_del_{f['id']}", use_container_width=True):
                        cloud_delete_formation(f["id"])
                        st.session_state.subs_formations = cloud_load_formations()
                        st.rerun()

    # ----- TAB 3: WEDSTRIJD -----
    with tabs[2]:
        st.markdown("#### Wedstrijdinstellingen")
        # Onthoud de oude settings zodat we weten of er iets is veranderd
        old_settings_snapshot = json.dumps(st.session_state.subs_match, sort_keys=True)

        oc1, oc2 = st.columns([2, 1])
        with oc1:
            new_opp = st.text_input("Tegenstander",
                                    value=st.session_state.subs_match.get("opponent", ""),
                                    key="subs_opponent", placeholder="Bijv. HC Rotterdam")
            st.session_state.subs_match["opponent"] = new_opp
        with oc2:
            new_date = st.text_input("Datum",
                                     value=st.session_state.subs_match.get("match_date",
                                                                           time.strftime("%Y-%m-%d")),
                                     key="subs_match_date", placeholder="YYYY-MM-DD")
            st.session_state.subs_match["match_date"] = new_date

        mc1, mc2, mc3, mc4 = st.columns(4)
        with mc1:
            new_half_length = st.selectbox(
                "Lengte per helft",
                [17, 15, 12, 10],
                format_func=lambda x: {17: "17,5 min (17)", 15: "15 min", 12: "12,5 min (12)", 10: "10 min"}[x],
                index=[17, 15, 12, 10].index(st.session_state.subs_match.get("half_length", 17)),
                key="subs_half_length",
            )
            st.session_state.subs_match["half_length"] = new_half_length
        with mc2:
            new_halves = st.selectbox(
                "Aantal helften", [2, 1],
                index=[2, 1].index(st.session_state.subs_match.get("halves", 2)),
                key="subs_halves",
            )
            st.session_state.subs_match["halves"] = new_halves
        with mc3:
            # Formaties uit cloud (als beschikbaar); anders fallback-defaults.
            form_names = list(_subs_active_formations().keys())
            if not form_names:
                form_names = list(SUBS_FORMATIONS_DEFAULT.keys())
            current_form = st.session_state.subs_match.get("formation", form_names[0])
            if current_form not in form_names:
                current_form = form_names[0]
            new_form = st.selectbox(
                "Formatie", form_names,
                index=form_names.index(current_form),
                key="subs_formation",
                help="Beheer je formaties in tab '2. Formaties'.",
            )
            st.session_state.subs_match["formation"] = new_form
        with mc4:
            new_block = st.selectbox(
                "Wissel elke…", [2, 3, 4, 5, 6],
                format_func=lambda x: f"{x} min",
                index=[2, 3, 4, 5, 6].index(st.session_state.subs_match.get("block_size", 4)),
                key="subs_block_size",
            )
            st.session_state.subs_match["block_size"] = new_block

        keepers = [p for p in st.session_state.subs_players
                   if (p.get("can_keep") or p.get("line") == "K")
                   and st.session_state.subs_attendance.get(p["id"], True)]
        if keepers:
            keeper_ids = [""] + [k["id"] for k in keepers]
            keeper_labels = ["— kies keeper —"] + [k["name"] for k in keepers]
            current = st.session_state.subs_match.get("fixed_keeper_id", "")
            idx = keeper_ids.index(current) if current in keeper_ids else 0
            chosen_keeper = st.selectbox("Vaste keeper", keeper_ids,
                                         format_func=lambda x: keeper_labels[keeper_ids.index(x)],
                                         index=idx, key="subs_fixed_keeper")
            st.session_state.subs_match["fixed_keeper_id"] = chosen_keeper
        else:
            st.warning("Nog geen keeper aangewezen. Zet minstens één speler op linie 'Keeper' of vink 'Kan keepen' aan.")

        # Als er iets is veranderd, meteen opslaan in cloud
        new_settings_snapshot = json.dumps(st.session_state.subs_match, sort_keys=True)
        if new_settings_snapshot != old_settings_snapshot:
            cloud_save_match_settings(st.session_state.subs_match)

        st.markdown("#### Aanwezigheid")
        if not st.session_state.subs_players:
            st.info("Voeg eerst spelers toe in tab 1.")
        else:
            ac1, ac2 = st.columns(2)
            with ac1:
                if st.button("Allemaal aanwezig", key="subs_all_present", use_container_width=True):
                    for p in st.session_state.subs_players:
                        st.session_state.subs_attendance[p["id"]] = True
                    cloud_bulk_save_attendance(st.session_state.subs_attendance)
                    st.rerun()
            with ac2:
                if st.button("Allemaal afwezig", key="subs_none_present", use_container_width=True):
                    for p in st.session_state.subs_players:
                        st.session_state.subs_attendance[p["id"]] = False
                    cloud_bulk_save_attendance(st.session_state.subs_attendance)
                    st.rerun()

            sorted_players = sorted(st.session_state.subs_players,
                                    key=lambda p: (SUBS_LINE_ORDER.index(p.get("line", "M")), p["name"]))
            cols = st.columns(4)
            for i, p in enumerate(sorted_players):
                with cols[i % 4]:
                    prev = st.session_state.subs_attendance.get(p["id"], True)
                    present = st.checkbox(
                        f"{p['name']} ({p.get('line', 'M')})",
                        value=prev,
                        key=f"subs_att_{p['id']}",
                    )
                    st.session_state.subs_attendance[p["id"]] = present
                    if present != prev:
                        cloud_save_attendance(p["id"], present)

    # ----- TAB 4: SCHEMA -----
    with tabs[3]:
        current_match_id = st.session_state.get("match_id")
        sc1, sc2, sc3 = st.columns([1.5, 1, 3])
        with sc1:
            if st.button("⚡ Automatisch genereren", key="subs_generate",
                         use_container_width=True, type="primary"):
                _subs_generate_schema()
                # Sla op: zowel per-wedstrijd (match_lineups) als generiek (team_schemas historie).
                if st.session_state.get("subs_schema"):
                    if current_match_id:
                        cloud_save_match_lineup(current_match_id, st.session_state.subs_schema,
                                                st.session_state.subs_match)
                    cloud_save_schema(st.session_state.subs_schema, st.session_state.subs_match)
                st.rerun()
        with sc2:
            if st.button("Leegmaken", key="subs_clear_schema", use_container_width=True):
                st.session_state.subs_schema = None
                st.session_state.subs_warnings = []
                st.rerun()
        with sc3:
            # Dropdown met eerder opgeslagen schema's
            saved = cloud_list_schemas(limit=15)
            if saved:
                options = [""] + [s["id"] for s in saved]
                def _fmt_saved(sid):
                    if not sid:
                        return "— eerder opgeslagen schema openen —"
                    item = next((s for s in saved if s["id"] == sid), None)
                    if not item:
                        return sid
                    when = (item.get("match_date") or "")[:10]
                    opp = item.get("opponent") or "zonder tegenstander"
                    return f"{when} • {opp}"
                picked = st.selectbox("Eerder opgeslagen", options,
                                      format_func=_fmt_saved,
                                      key="subs_saved_pick",
                                      label_visibility="collapsed")
                if picked:
                    loaded = cloud_load_schema(picked)
                    if loaded and loaded.get("schema_json"):
                        st.session_state.subs_schema = loaded["schema_json"]
                        if loaded.get("settings_json"):
                            st.session_state.subs_match.update(loaded["settings_json"])
                        # Koppel dit schema direct ook aan de huidige wedstrijd
                        if current_match_id:
                            cloud_save_match_lineup(
                                current_match_id,
                                st.session_state.subs_schema,
                                st.session_state.subs_match,
                            )
                        st.success("Eerder schema geladen.")
                        st.rerun()

        for w in st.session_state.get("subs_warnings", []):
            st.warning(f"⚠ {w}")

        schema = st.session_state.subs_schema
        if not schema:
            st.info("Nog geen schema. Klik op 'Automatisch genereren'.")
        else:
            st.caption("🟩 In het veld  •  ⬛ Bank  •  🟨 Speler met voorkeur (meer speeltijd)  •  🟥 Te veel in veld  •  🟧 Te weinig")
            players = [p for p in st.session_state.subs_players
                       if st.session_state.subs_attendance.get(p["id"], True)]
            by_line = {"V": [], "M": [], "A": [], "K": []}
            for p in players:
                by_line[p.get("line", "M")].append(p)
            for L in SUBS_LINE_ORDER:
                line_players = sorted(by_line[L], key=lambda x: x["name"])
                if not line_players:
                    continue
                slots = _subs_formation_slots(st.session_state.subs_match["formation"])
                required = 1 if L == "K" else slots.get(L, 0)
                _subs_render_schema_table(L, line_players, schema["cells"],
                                          schema["minutes_per_half"], schema["halves"], required)

            # Download CSV
            st.markdown("#### Exporteren")
            lines = []
            for L in SUBS_LINE_ORDER:
                lp = sorted(by_line[L], key=lambda x: x["name"])
                if not lp:
                    continue
                lines.append(SUBS_LINE_NAMES[L])
                header = "Min>>," + ",".join(str(i + 1) for i in range(schema["total_minutes"])) + ",Half,Vol"
                lines.append(header)
                for p in lp:
                    pc = schema["cells"].get(p["id"], [])
                    h1 = sum(pc[:schema["minutes_per_half"]])
                    full = sum(pc)
                    row = p["name"] + "," + ",".join(str(v) for v in pc) + f",{h1},{full}"
                    lines.append(row)
                lines.append("")
            csv_content = "\n".join(lines)
            exp_c1, exp_c2 = st.columns(2)
            with exp_c1:
                st.download_button(
                    "📊 Download als CSV",
                    data=csv_content,
                    file_name=f"wisselschema_{st.session_state.get('opponent_name', 'wedstrijd')}.csv",
                    mime="text/csv",
                    key="subs_download_csv",
                )
            with exp_c2:
                try:
                    pdf_bytes = export_lineup_pdf(
                        schema,
                        st.session_state.subs_match,
                        players,
                        st.session_state.get("team_name", "Ons team"),
                        st.session_state.get("opponent_name", "Tegenstander"),
                    )
                    pdf_mime = "application/pdf" if REPORTLAB_AVAILABLE else "text/plain"
                except Exception:
                    pdf_bytes = b""
                    pdf_mime = "text/plain"
                st.download_button(
                    "📄 Download wisselschema PDF",
                    data=pdf_bytes,
                    file_name=f"wisselschema_{st.session_state.get('opponent_name', 'wedstrijd')}.pdf",
                    mime=pdf_mime,
                    key="subs_download_pdf",
                    disabled=(not pdf_bytes),
                )


# ==================================================
# SEIZOENS-AGGREGATIE — dashboard-stats over alle wedstrijden
# ==================================================
def _load_all_team_matches(include_unscoped: bool = True) -> list[dict]:
    """Haal alle wedstrijden + events van het actieve team op.

    Retourneert: [{"match_id": ..., "events": [...], "date": ...}, ...]
    Wedstrijden zonder events worden overgeslagen.

    Combineert twee bronnen zodat álle wedstrijden van dit team zichtbaar zijn:
      1. Wedstrijden met de team-prefix (`T-xxxxxxxx__...`) — data van na
         de multi-team split.
      2. Optioneel óók onscoped wedstrijden (zonder prefix) — oude data van
         vóór de multi-team split. Deze worden alleen meegenomen als
         `include_unscoped=True`.
    """
    client = get_supabase_client()
    if client is None:
        return []
    prefix = _team_match_prefix()

    def _run(ilike_pattern: str | None):
        q = client.table("match_events").select(
            "match_id,team,event,quarter,zone,time_sec,source,note,created_at"
        ).order("created_at", desc=True).limit(5000)
        if ilike_pattern:
            q = q.ilike("match_id", ilike_pattern)
        return q.execute().data or []

    rows_scoped: list[dict] = []
    rows_unscoped: list[dict] = []
    try:
        if prefix:
            rows_scoped = _run(f"{prefix}%")
        if include_unscoped:
            # Haal alle events op en filter in-memory de prefix-matches eruit,
            # zodat we alleen echte onscoped wedstrijden overhouden.
            # (Supabase heeft geen NOT-LIKE-met-wildcards op dezelfde ilike call.)
            all_rows = _run(None)
            if prefix:
                rows_unscoped = [
                    r for r in all_rows
                    if not str(r.get("match_id", "")).startswith(prefix)
                    # én óók niet de prefix van een ánder team (T-xxxxxxxx__)
                    and not (
                        str(r.get("match_id", "")).startswith("T-")
                        and "__" in str(r.get("match_id", ""))
                    )
                ]
            else:
                rows_unscoped = all_rows
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("seizoens-events ophalen", err)
        return []

    # Combineer + dedup op (match_id, id_of_row)
    combined = list(rows_scoped) + list(rows_unscoped)

    # Groepeer per match_id
    grouped: dict[str, list] = {}
    first_seen: dict[str, str] = {}
    for r in combined:
        mid = r.get("match_id")
        if not mid:
            continue
        grouped.setdefault(mid, []).append(r)
        ts = r.get("created_at") or ""
        if mid not in first_seen or ts < first_seen[mid]:
            first_seen[mid] = ts

    matches = []
    for mid, events in grouped.items():
        matches.append({
            "match_id": mid,
            "events": events,
            "date": first_seen.get(mid, ""),
        })
    # Sorteer op datum (oudste eerst = chronologisch)
    matches.sort(key=lambda m: m.get("date", ""))
    return matches


def _match_result(events: list[dict], own_team_hint: str = "") -> dict:
    """Bepaal doelpunten-voor/tegen + resultaat voor één wedstrijd.

    Als `own_team_hint` gegeven is (de actieve team_name in sessie), gebruiken
    we die om eigen team te identificeren. Anders nemen we de team-naam met
    de meeste events als 'eigen team'.
    """
    team_counts: dict[str, int] = {}
    goals: dict[str, int] = {}
    for e in events:
        t = str(e.get("team", "")).strip()
        if not t:
            continue
        team_counts[t] = team_counts.get(t, 0) + 1
        if str(e.get("event", "")).strip().lower() == "goal":
            goals[t] = goals.get(t, 0) + 1

    names_sorted = sorted(team_counts.items(), key=lambda kv: kv[1], reverse=True)
    if own_team_hint and own_team_hint in team_counts:
        own = own_team_hint
    elif names_sorted:
        own = names_sorted[0][0]
    else:
        own = ""
    opp = next((n for n, _ in names_sorted if n != own), "")

    gf = goals.get(own, 0)
    ga = goals.get(opp, 0) if opp else 0
    if gf > ga:
        result = "W"
    elif gf < ga:
        result = "V"
    else:
        result = "G"
    return {
        "own": own,
        "opponent": opp or "Tegenstander",
        "goals_for": gf,
        "goals_against": ga,
        "result": result,
    }


def build_season_summary(matches: list[dict], own_team_hint: str = "") -> dict:
    """Seizoensoverzicht samenstellen op basis van alle wedstrijden.

    Output:
      - wins, draws, losses, played
      - goals_for, goals_against, goal_diff
      - win_pct
      - per_match: lijst met per-wedstrijd breakdown (voor trend-grafiek)
      - circle_entries_for, shots_for (totaal aanvallende stats)
    """
    summary = {
        "wins": 0, "draws": 0, "losses": 0, "played": 0,
        "goals_for": 0, "goals_against": 0, "goal_diff": 0,
        "win_pct": 0.0,
        "per_match": [],
        "circle_entries_for": 0,
        "shots_for": 0,
    }

    for m in matches:
        ev = m["events"]
        if not ev:
            continue
        mr = _match_result(ev, own_team_hint)
        summary["played"] += 1
        if mr["result"] == "W":
            summary["wins"] += 1
        elif mr["result"] == "G":
            summary["draws"] += 1
        else:
            summary["losses"] += 1
        summary["goals_for"] += mr["goals_for"]
        summary["goals_against"] += mr["goals_against"]

        # Cirkelentries / shots voor eigen team deze wedstrijd
        own = mr["own"]
        ce = sum(1 for e in ev
                 if str(e.get("team", "")).strip() == own
                 and str(e.get("event", "")).strip().lower() == "cirkelentry")
        sh = sum(1 for e in ev
                 if str(e.get("team", "")).strip() == own
                 and str(e.get("event", "")).strip().lower() in ("shot", "schot", "doelpoging"))
        summary["circle_entries_for"] += ce
        summary["shots_for"] += sh

        summary["per_match"].append({
            "match_id": m["match_id"],
            "pretty_id": unscope_match_id(m["match_id"]) if m.get("match_id") else "",
            "date": (m.get("date") or "")[:10],
            "opponent": mr["opponent"],
            "goals_for": mr["goals_for"],
            "goals_against": mr["goals_against"],
            "result": mr["result"],
            "circle_entries": ce,
            "shots": sh,
        })

    summary["goal_diff"] = summary["goals_for"] - summary["goals_against"]
    if summary["played"]:
        summary["win_pct"] = round((summary["wins"] * 100.0) / summary["played"], 1)
    return summary


def build_player_minutes(matches: list[dict]) -> dict:
    """Speelminuten per speler over het seizoen — uit match_lineups.

    Voor elke wedstrijd wordt de lineup opgehaald en de cells-matrix
    opgeteld naar minuten per player_id. Live wissel-events worden
    (nog) niet meegenomen — dat komt in release 2B.

    Output: {
        "players": {player_id: {"name": "...", "minutes": N, "matches": M}},
        "total_minutes": N,
        "total_matches_with_schema": K,
    }
    """
    players_cache = {p["id"]: p for p in st.session_state.get("subs_players", [])}
    players_stats: dict[str, dict] = {}
    total_minutes = 0
    total_matches_with_schema = 0

    for m in matches:
        mid = m.get("match_id")
        if not mid:
            continue
        lineup = cloud_load_match_lineup(mid)
        if not lineup:
            continue
        schema = lineup.get("schema_json") or {}
        cells = schema.get("cells") or {}
        if not cells:
            continue
        total_matches_with_schema += 1
        for pid, row in cells.items():
            try:
                mins = int(sum(int(x or 0) for x in row))
            except Exception:
                mins = 0
            p = players_stats.setdefault(pid, {
                "name": players_cache.get(pid, {}).get("name", pid),
                "minutes": 0,
                "matches": 0,
            })
            p["minutes"] += mins
            if mins > 0:
                p["matches"] += 1
            total_minutes += mins

    return {
        "players": players_stats,
        "total_minutes": total_minutes,
        "total_matches_with_schema": total_matches_with_schema,
    }


def _load_season_diagnostics() -> dict:
    """Meta-info voor de seizoens-tool: hoeveel matches in scope / onscoped / andere teams."""
    client = get_supabase_client()
    if client is None:
        return {
            "client_ok": False, "prefix": "", "scoped": 0, "unscoped": 0,
            "other_teams": 0, "global": 0, "total_events": 0, "goal_events": 0,
            "entry_events": 0, "team_names": [], "sample": [],
        }
    prefix = _team_match_prefix()
    scoped = 0
    unscoped = 0
    other_teams = 0
    total = 0
    total_events = 0
    goal_events = 0
    entry_events = 0
    team_names: dict[str, int] = {}
    sample: list[dict] = []
    try:
        r_all = client.table("match_events").select(
            "match_id,team,event,quarter,time_sec,created_at"
        ).order("created_at", desc=True).limit(5000).execute().data or []
        ids = {x.get("match_id") for x in r_all if x.get("match_id")}
        total = len(ids)
        for mid in ids:
            if prefix and mid.startswith(prefix):
                scoped += 1
            elif mid.startswith("T-") and "__" in mid:
                other_teams += 1
            else:
                unscoped += 1

        # Event-level diagnostiek — beperkt tot wedstrijden die we gebruiken
        # (scoped + onscoped; niet andere teams).
        for r in r_all:
            mid = r.get("match_id") or ""
            is_other = mid.startswith("T-") and "__" in mid and (not prefix or not mid.startswith(prefix))
            if is_other:
                continue
            total_events += 1
            ev = str(r.get("event", "")).strip().lower()
            if ev == "goal":
                goal_events += 1
            elif ev == "cirkelentry":
                entry_events += 1
            tm = str(r.get("team", "")).strip()
            if tm:
                team_names[tm] = team_names.get(tm, 0) + 1
        # Sample: eerste 3 events voor inspectie
        for r in r_all[:3]:
            sample.append({
                "match_id": r.get("match_id", ""),
                "team": r.get("team", ""),
                "event": r.get("event", ""),
                "quarter": r.get("quarter", ""),
            })
    except Exception as err:
        log_cloud_error("seizoens-diagnose", err)
    # Top 5 meest voorkomende team-namen (met telling)
    top_teams = sorted(team_names.items(), key=lambda kv: -kv[1])[:5]
    return {
        "client_ok": True, "prefix": prefix,
        "scoped": scoped, "unscoped": unscoped,
        "other_teams": other_teams, "global": total,
        "total_events": total_events,
        "goal_events": goal_events,
        "entry_events": entry_events,
        "team_names": top_teams,
        "sample": sample,
    }


def _cloud_health_check() -> dict:
    """Directe diagnose van de cloud-verbinding voor Seizoensoverzicht.

    Geeft terug:
      - ok:           True als alle basisstappen slagen
      - project_ref:  het xxxx deel uit https://xxxx.supabase.co
      - url:          gevonden SUPABASE_URL (of leeg)
      - has_key:      of er een KEY in secrets staat
      - events_count: rij-aantal in match_events (alleen wat anon mag zien)
      - match_ids:    aantal unieke match_id's
      - select_ok:    of de SELECT-call zonder exception terugkwam
      - error:        tekst van de laatste fout (indien aanwezig)
    """
    result = {
        "ok": False, "project_ref": "", "url": "", "has_key": False,
        "events_count": 0, "match_ids": 0, "select_ok": False,
        "error": "",
    }
    try:
        url, _ = _find_secret_value(["SUPABASE_URL", "supabase_url", "URL", "url"])
        key, _ = _find_secret_value([
            "SUPABASE_KEY", "supabase_key",
            "SUPABASE_ANON_KEY", "supabase_anon_key",
            "ANON_KEY", "anon_key", "KEY", "key",
        ])
    except Exception as err:
        result["error"] = f"secret-lookup faalde: {err}"
        return result

    result["url"] = url or ""
    result["has_key"] = bool(key)
    # Extract project_ref uit https://xxxx.supabase.co
    try:
        if url and "://" in url:
            host = url.split("://", 1)[1].split("/", 1)[0]
            result["project_ref"] = host.split(".", 1)[0]
    except Exception:
        pass

    client = get_supabase_client()
    if client is None:
        result["error"] = "get_supabase_client() gaf None — SUPABASE_URL of KEY ontbreekt/ongeldig."
        return result

    # Probeer alle rijen op te halen (alleen id's voor snelheid)
    try:
        rows = client.table("match_events").select("match_id").limit(10000).execute().data or []
        result["select_ok"] = True
        result["events_count"] = len(rows)
        result["match_ids"] = len({r.get("match_id") for r in rows if r.get("match_id")})
        result["ok"] = True
    except Exception as err:
        result["error"] = f"SELECT match_events faalde: {err}"
        return result

    return result


def _compute_top_scorers(matches: list[dict], roster: list[dict], limit: int = 10) -> list[dict]:
    """Tel goals per player_id over alle wedstrijden (alleen eigen team)."""
    id_to_name = {p.get("id"): p.get("name", "?") for p in (roster or [])}
    own_hint = st.session_state.get("team_name", "")
    counts: dict[str, int] = {}
    for m in matches:
        for e in m.get("events", []):
            if str(e.get("event", "")).strip().lower() != "goal":
                continue
            team = str(e.get("team", "")).strip()
            if own_hint and team != own_hint:
                continue
            pid = e.get("player_id")
            if not pid:
                continue
            counts[pid] = counts.get(pid, 0) + 1
    rows = [{"player_id": pid, "name": id_to_name.get(pid, pid), "goals": c}
            for pid, c in counts.items()]
    rows.sort(key=lambda r: -r["goals"])
    return rows[:limit]


def render_season_screen() -> None:
    """Volledige Seizoensoverzicht-pagina als eigen tool.

    Toont wedstrijd-statistieken, topscorers, speelminuten en een PDF-export,
    met een diagnose-blok zodat je ziet hoeveel wedstrijden er gevonden zijn.
    """
    team_name = st.session_state.get("team_name") or st.session_state.get("active_team_name") or "je team"

    # Header
    hc1, hc2 = st.columns([4, 1])
    with hc1:
        st.markdown(f"### 📊 Seizoensoverzicht — {team_name}")
        st.caption("Alle wedstrijden en events van dit team samengevat in één overzicht.")
    with hc2:
        if st.button("🔄  Ververs", use_container_width=True, key="season_refresh"):
            st.rerun()

    # ---- Cloud health check (een-klik diagnose) ----
    with st.expander("☁️ Cloud-verbinding — health check", expanded=False):
        st.caption(
            "Eén-klik test: kijkt direct naar je Supabase-project en telt rijen in `match_events`. "
            "Gebruik dit als het dashboard leeg blijft om te zien of het probleem zit in "
            "(a) de secrets, (b) RLS, of (c) een lege tabel."
        )
        if st.button("🩺  Test cloud-verbinding", key="season_cloud_health_btn"):
            with st.spinner("Cloud-verbinding testen…"):
                health = _cloud_health_check()
            st.session_state["_season_health_last"] = health
        health = st.session_state.get("_season_health_last")
        if health:
            # Project-info
            url = health.get("url") or ""
            ref = health.get("project_ref") or "(onbekend)"
            has_key = health.get("has_key", False)
            st.markdown("**🔧 Configuratie**")
            st.markdown(
                f"- Project ref: `{ref}`\n"
                f"- URL: `{url or '(leeg)'}`\n"
                f"- KEY in secrets: {'✅ ja' if has_key else '❌ nee'}"
            )
            # Verbinding
            st.markdown("**🛰️ Verbinding**")
            if not health.get("ok") and health.get("error"):
                st.error(f"❌ {health['error']}")
            elif health.get("select_ok"):
                cnt = health.get("events_count", 0)
                ids = health.get("match_ids", 0)
                if cnt == 0:
                    st.warning(
                        f"✅ Verbinding OK, maar `match_events` geeft **0 rijen** terug.\n\n"
                        f"Dit betekent één van drie dingen:\n"
                        f"1. **De tabel is écht leeg** — eerdere wedstrijden zijn nooit naar de cloud "
                        f"geschreven (alleen lokaal). Open Supabase → Table editor → `match_events` "
                        f"om dit te bevestigen.\n"
                        f"2. **RLS blokkeert SELECT voor anon** — check Supabase → Authentication → "
                        f"Policies. Voer in SQL Editor `select count(*) from match_events;` uit. Als "
                        f"dat >0 geeft maar deze knop 0, dan is het RLS.\n"
                        f"3. **Verkeerd project** — als je Supabase dashboard zegt dat de tabel "
                        f"vol staat, kijk je hier waarschijnlijk naar een ander project. Vergelijk "
                        f"`{ref}` met de URL in je Supabase dashboard."
                    )
                else:
                    st.success(
                        f"✅ Verbinding OK. **{cnt} events** in {ids} unieke wedstrijden gevonden."
                    )
            else:
                st.error("❌ SELECT op `match_events` faalde — zie de error hierboven.")
            # Vergelijking met team-prefix
            prefix = _team_match_prefix()
            st.markdown("**🏷️ Team-scope**")
            st.markdown(
                f"- Actief team: `{st.session_state.get('team_name', '(geen)')}`\n"
                f"- Verwachte prefix: `{prefix or '(geen — team_id ontbreekt)'}`"
            )
        else:
            st.info("Klik op **Test cloud-verbinding** om een directe diagnose te doen.")

    # Laden + diagnose
    with st.spinner("Seizoensdata laden…"):
        matches = _load_all_team_matches(include_unscoped=True)
        diag = _load_season_diagnostics()

    own_hint = st.session_state.get("team_name", "")

    # Fallback: als de huidige team-naam niet voorkomt in de events,
    # gebruik dan de meest voorkomende team-naam als 'eigen team' zodat
    # de stats toch correct worden berekend.
    team_counts_global: dict[str, int] = {}
    for m in matches:
        for e in m.get("events", []):
            tm = str(e.get("team", "")).strip()
            if tm:
                team_counts_global[tm] = team_counts_global.get(tm, 0) + 1
    if matches and (not own_hint or own_hint not in team_counts_global):
        if team_counts_global:
            most_common = max(team_counts_global.items(), key=lambda kv: kv[1])[0]
            own_hint = most_common

    summary = build_season_summary(matches, own_hint)

    # Diagnose-kaart: laat zien waarom wel/geen wedstrijden
    with st.expander("🔍 Diagnose — welke wedstrijden zijn opgehaald?", expanded=(summary["played"] == 0)):
        if not diag.get("client_ok"):
            st.warning("Geen Supabase-verbinding. Controleer je secrets (SUPABASE_URL / SUPABASE_KEY).")
        else:
            st.markdown(
                f"- Team-prefix: `{diag['prefix'] or '(geen — alles wordt globaal opgehaald)'}`\n"
                f"- Wedstrijden **mét jouw team-prefix**: **{diag['scoped']}** ✅\n"
                f"- Wedstrijden **zonder team-prefix** (oude data): **{diag['unscoped']}** "
                + ("✅ meegenomen" if diag['unscoped'] > 0 else "—") + "\n"
                f"- Wedstrijden van **andere teams** (genegeerd): {diag['other_teams']}\n"
                f"- Totaal in cloud: {diag['global']}\n"
                f"- **Geladen voor dit dashboard: {len(matches)} wedstrijden**"
            )
            if diag["scoped"] > 0 and diag["unscoped"] > 0:
                st.success(
                    f"Alle {diag['scoped']} team-wedstrijden én {diag['unscoped']} oude onscoped wedstrijden "
                    "zijn samengevoegd in dit dashboard."
                )
            elif diag["scoped"] == 0 and diag["unscoped"] > 0:
                st.info(
                    f"Geen wedstrijden met team-prefix gevonden, maar wel {diag['unscoped']} onscoped "
                    "wedstrijden (oude data). Die zijn nu geladen."
                )
            elif diag["scoped"] == 0 and diag["unscoped"] == 0 and diag["other_teams"] > 0:
                st.warning(
                    f"Alleen wedstrijden van andere teams gevonden ({diag['other_teams']}). "
                    "Log in met het juiste team om die te zien, of maak hier een nieuwe wedstrijd aan."
                )
            # ---- Event-level diagnostiek ----
            st.markdown("---")
            st.markdown("**📊 Events in de geladen wedstrijden**")
            st.markdown(
                f"- Totaal events: **{diag.get('total_events', 0)}**\n"
                f"- Goals: **{diag.get('goal_events', 0)}**\n"
                f"- Cirkelentries: **{diag.get('entry_events', 0)}**"
            )

            tn = diag.get("team_names", [])
            if tn:
                st.markdown("**Team-namen in events** (top 5 meest voorkomend):")
                for name, cnt in tn:
                    match_hint = ""
                    own_hint = st.session_state.get("team_name", "")
                    if own_hint and name == own_hint:
                        match_hint = "  ← eigen team (sessie)"
                    st.markdown(f"- `{name}` — {cnt} events{match_hint}")
                # Waarschuw als eigen team-naam niet voorkomt
                own_hint = st.session_state.get("team_name", "")
                if own_hint and not any(name == own_hint for name, _ in tn):
                    st.warning(
                        f"⚠️ Je huidige team-naam **'{own_hint}'** komt niet voor in de events. "
                        "De app raadt automatisch welke team-naam het eigen team is (meest voorkomende), "
                        "maar het kan zijn dat dit niet klopt. Zet je team-naam op "
                        f"'{tn[0][0]}' voor consistente stats."
                    )

            sample = diag.get("sample", [])
            if sample:
                with st.expander("Sample-events (eerste 3 rijen)", expanded=False):
                    for s in sample:
                        st.code(
                            f"match_id: {s['match_id']}\n"
                            f"team:     {s['team']}\n"
                            f"event:    {s['event']}\n"
                            f"quarter:  {s['quarter']}",
                            language="text",
                        )

            if len(matches) > 0:
                with st.expander(f"Toon match-ID's van geladen wedstrijden ({len(matches)})", expanded=False):
                    for m in matches[-30:]:
                        st.code(m.get("match_id", "?"), language="text")

    # ---- Rij 1: KPI kaartjes ----
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        render_info_card(
            "Wedstrijden",
            str(summary["played"]),
            f"W {summary['wins']} · G {summary['draws']} · V {summary['losses']}",
            accent="#3b82f6",
        )
    with c2:
        render_info_card(
            "Win %",
            f"{summary['win_pct']:.0f}%" if summary["played"] else "—",
            "percentage gewonnen",
            accent="#10b981",
        )
    with c3:
        diff = summary["goal_diff"]
        diff_label = f"+{diff}" if diff > 0 else str(diff)
        render_info_card(
            "Doelsaldo",
            diff_label,
            f"{summary['goals_for']} voor · {summary['goals_against']} tegen",
            accent="#f59e0b",
        )
    with c4:
        render_info_card(
            "Cirkelentries",
            str(summary["circle_entries_for"]),
            f"{summary['shots_for']} shots dit seizoen",
            accent="#8b5cf6",
        )

    if summary["played"] == 0:
        st.info("Nog geen wedstrijden met events gevonden — start een wedstrijd en tag events "
                "om hier statistieken te zien.")
        return

    # ---- Rij 2: topscorers (alleen als er scorers gekoppeld zijn) ----
    roster = _active_team_roster()
    top_scorers = _compute_top_scorers(matches, roster, limit=10)
    if top_scorers:
        st.markdown("#### ⚽ Topscorers")
        pills = []
        for i, s in enumerate(top_scorers):
            medal = ["🥇", "🥈", "🥉"][i] if i < 3 else "⚽"
            pills.append(
                f"<span style='background:#eff6ff; color:#1d4ed8; padding:6px 12px; "
                f"border-radius:999px; margin:3px; display:inline-block; font-size:14px; "
                f"font-weight:600;'>{medal} {s['name']} — {s['goals']}</span>"
            )
        st.markdown("<div>" + "".join(pills) + "</div>", unsafe_allow_html=True)
    else:
        st.caption("Nog geen topscorers — koppel de scorer aan goals in het Rapport-scherm of via de popup tijdens de wedstrijd.")

    # ---- Rij 3: wedstrijd-trend (cirkelentries / shots per wedstrijd) ----
    st.markdown(" ")
    lc, rc = st.columns([2, 1])
    with lc:
        st.markdown("#### Aanvallende trend (cirkelentries per wedstrijd)")
        pm = summary["per_match"]
        if pm:
            trend_df = pd.DataFrame([
                {
                    "Wedstrijd": f"{(p['date'] or '')[5:] or '?'} vs {p['opponent']}",
                    "Cirkelentries": p["circle_entries"],
                    "Shots": p["shots"],
                    "Goals voor": p["goals_for"],
                    "Goals tegen": p["goals_against"],
                }
                for p in pm
            ])
            try:
                st.line_chart(trend_df.set_index("Wedstrijd")[["Cirkelentries", "Shots", "Goals voor"]])
            except Exception:
                st.dataframe(trend_df, use_container_width=True, hide_index=True)
        else:
            st.caption("Geen wedstrijden om trend op te baseren.")

    with rc:
        st.markdown("#### Laatste 5 resultaten")
        for p in summary["per_match"][-5:][::-1]:
            res = p["result"]
            color = {"W": "#10b981", "G": "#f59e0b", "V": "#ef4444"}.get(res, "#9ca3af")
            st.markdown(
                f"<div style='display:flex; align-items:center; justify-content:space-between; "
                f"padding:6px 10px; border-bottom:1px solid {CARD_BORDER_SOFT};'>"
                f"<span style='color:{TEXT_MAIN}'>{p['date'] or '?'} · {p['opponent']}</span>"
                f"<span style='color:{color}; font-weight:700;'>{res} {p['goals_for']}-{p['goals_against']}</span>"
                f"</div>",
                unsafe_allow_html=True,
            )

    # ---- Rij 4: volledige wedstrijdlijst ----
    st.markdown(" ")
    st.markdown("#### 🗓️ Alle wedstrijden")
    pm = summary.get("per_match", [])
    if pm:
        matches_df = pd.DataFrame([
            {
                "Datum": p.get("date", "") or "—",
                "Tegenstander": p.get("opponent", "") or "—",
                "Resultaat": p.get("result", ""),
                "Goals": f"{p.get('goals_for', 0)}-{p.get('goals_against', 0)}",
                "Entries": p.get("circle_entries", 0),
                "Shots": p.get("shots", 0),
            }
            for p in pm
        ])
        st.dataframe(matches_df, use_container_width=True, hide_index=True)
    else:
        st.caption("Geen wedstrijden om te tonen.")

    # ---- Rij 5: speelminuten-verdeling ----
    st.markdown(" ")
    st.markdown("#### ⏱️ Speelminuten-verdeling (uit gegenereerde schema's)")
    minutes_info = build_player_minutes(matches)
    if minutes_info["total_matches_with_schema"] == 0:
        st.caption("Nog geen wisselschema's gegenereerd — zodra je ze opslaat, verschijnt hier "
                   "de verdeling over alle spelers.")
    else:
        rows = []
        for pid, p in minutes_info["players"].items():
            rows.append({"Speler": p["name"], "Minuten": p["minutes"], "Wedstrijden": p["matches"]})
        if rows:
            df = pd.DataFrame(rows).sort_values("Minuten", ascending=False)
            try:
                st.bar_chart(df.set_index("Speler")["Minuten"], use_container_width=True)
            except Exception:
                st.dataframe(df, use_container_width=True, hide_index=True)
            with st.expander("Details per speler", expanded=False):
                st.dataframe(df, use_container_width=True, hide_index=True)

    # ---- Rij 4: seizoens-PDF export ----
    st.markdown(" ")
    try:
        pdf_bytes = export_season_pdf(
            summary,
            minutes_info if summary["played"] > 0 else {"players": {}, "total_minutes": 0, "total_matches_with_schema": 0},
            st.session_state.get("team_name", "Ons team"),
        )
        pdf_mime = "application/pdf" if REPORTLAB_AVAILABLE else "text/plain"
    except Exception:
        pdf_bytes = b""
        pdf_mime = "text/plain"
    st.download_button(
        "📄 Download seizoens-PDF",
        data=pdf_bytes,
        file_name=f"seizoensoverzicht_{st.session_state.get('team_name','team')}.pdf",
        mime=pdf_mime,
        key="dl_season_pdf",
        disabled=(not pdf_bytes),
    )


# ==================================================
# SPELERSPROFIEL — notities per speler
# ==================================================

NOTE_CATEGORIES = ["Technisch", "Tactisch", "Fysiek", "Mentaal", "Algemeen"]


def cloud_add_player_note(player_id: str, note_date: str, category: str,
                           rating: int | None, note: str) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return
    try:
        client.table("player_notes").insert({
            "id": str(uuid.uuid4()),
            "team_id": tid,
            "player_id": player_id,
            "note_date": note_date,
            "category": category,
            "rating": rating,
            "note": note.strip(),
        }).execute()
        _fetch_player_notes.clear()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("notitie opslaan", err)


@st.cache_data(ttl=30, show_spinner=False)
def _fetch_player_notes(team_id: str, player_id: str) -> list:
    client = get_supabase_client()
    if not team_id or not player_id or client is None:
        return []
    try:
        response = (
            client.table("player_notes")
            .select("*")
            .eq("team_id", team_id)
            .eq("player_id", player_id)
            .order("note_date", desc=True)
            .execute()
        )
        return response.data or []
    except Exception:
        return []


def cloud_get_player_notes(player_id: str) -> list:
    tid = _active_team_id()
    result = _fetch_player_notes(tid or "", player_id)
    if result is not None:
        mark_cloud_ok()
    return result or []


def cloud_delete_player_note(note_id: str) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return
    try:
        client.table("player_notes").delete().eq("id", note_id).eq("team_id", tid).execute()
        _fetch_player_notes.clear()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("notitie verwijderen", err)


def render_player_profile_screen() -> None:
    st.markdown("### 👤 Spelersprofiel")

    roster = _active_team_roster()
    if not roster:
        st.info("Voeg eerst spelers toe via de Wisselschema-tool.")
        return

    sorted_roster = sorted(roster, key=lambda p: p.get("name", ""))
    player_map = {p["name"]: p for p in sorted_roster}
    selected_name = st.selectbox("Kies speler", list(player_map.keys()), key="profile_player_sel")
    player = player_map[selected_name]
    pid = player["id"]

    # --- Stats uit match events (gebruik al gebouwde df uit session state, geen extra query) ---
    all_events = build_df()  # build_df gebruikt session_state.events — geen Supabase call
    goals_scored = 0
    if not all_events.empty:
        goals_scored = int(
            ((all_events["event"] == "Goal") & (all_events["player_id"] == pid)).sum()
        )

    line_label = {"K": "Keeper", "V": "Verdediger", "M": "Middenvelder", "A": "Aanvaller"}.get(
        player.get("line", "M"), player.get("line", "—")
    )
    prio_label = {"high": "Meer speeltijd ⭐", "low": "Minder speeltijd", "normal": "Normaal"}.get(
        player.get("priority", "normal"), "Normaal"
    )

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Speler", player["name"])
    c2.metric("Linie", line_label)
    c3.metric("Speeltijd voorkeur", prio_label)
    c4.metric("Goals (deze sessie)", goals_scored)

    st.divider()

    col_notes, col_form = st.columns([1.5, 1])

    with col_form:
        st.subheader("Notitie toevoegen")
        with st.form("profile_note_form"):
            note_date = st.date_input("Datum", value=__import__("datetime").date.today())
            category = st.selectbox("Categorie", NOTE_CATEGORIES)
            rating = st.select_slider(
                "Beoordeling (optioneel)",
                options=["—", 1, 2, 3, 4, 5],
                value="—",
            )
            note_text = st.text_area("Notitie", placeholder="Wat viel je op bij deze speler?", height=130)
            submitted = st.form_submit_button("Opslaan", type="primary")

        if submitted:
            if not note_text.strip():
                st.error("Vul een notitie in.")
            else:
                rating_val = None if rating == "—" else int(rating)
                cloud_add_player_note(pid, note_date.strftime("%Y-%m-%d"), category, rating_val, note_text)
                st.success("Notitie opgeslagen.")
                st.rerun()

    with col_notes:
        st.subheader("Notities")
        notes = cloud_get_player_notes(pid)

        if not notes:
            st.info("Nog geen notities voor deze speler.")
        else:
            cat_filter = st.multiselect(
                "Filter op categorie", NOTE_CATEGORIES,
                default=NOTE_CATEGORIES, key="profile_cat_filter",
            )
            filtered = [n for n in notes if n.get("category") in cat_filter]

            for note in filtered:
                rating_val = note.get("rating")
                stars = "⭐" * int(rating_val) if rating_val else ""
                date_fmt = note.get("note_date", "")[:10]
                with st.expander(f"{date_fmt} · {note.get('category', '—')} {stars}"):
                    st.write(note.get("note", ""))
                    if st.button("Verwijder", key=f"del_note_{note['id']}"):
                        cloud_delete_player_note(note["id"])
                        st.rerun()

    st.divider()
    st.subheader("Beoordelingstrend per categorie")
    rated = [n for n in cloud_get_player_notes(pid) if n.get("rating")]
    if not rated:
        st.info("Voeg beoordelingen (1–5) toe om de trend te zien.")
    else:
        import datetime as _dt
        rows = []
        for n in rated:
            rows.append({
                "Datum": n.get("note_date", "")[:10],
                "Categorie": n.get("category", ""),
                "Beoordeling": float(n["rating"]),
            })
        trend_df = pd.DataFrame(rows)
        trend_df["Datum"] = pd.to_datetime(trend_df["Datum"])
        pivot = trend_df.pivot_table(index="Datum", columns="Categorie",
                                     values="Beoordeling", aggfunc="mean")
        st.line_chart(pivot)


# ==================================================
# WEDSTRIJDENBEHEER — uitslagen en tegenstanders
# ==================================================

def cloud_upsert_match_result(match_id: str, opponent: str, home_away: str,
                               location: str, our_score: int | None,
                               opp_score: int | None, notes: str) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return
    try:
        client.table("match_results").upsert({
            "match_id": match_id,
            "team_id": tid,
            "opponent": opponent.strip(),
            "home_away": home_away,
            "location": location.strip(),
            "our_score": our_score,
            "opp_score": opp_score,
            "notes": notes.strip(),
        }, on_conflict="match_id").execute()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("wedstrijdresultaat opslaan", err)


def cloud_get_match_result(match_id: str) -> dict | None:
    client = get_supabase_client()
    if client is None:
        return None
    try:
        response = (
            client.table("match_results")
            .select("*")
            .eq("match_id", match_id)
            .limit(1)
            .execute()
        )
        mark_cloud_ok()
        rows = response.data or []
        return rows[0] if rows else None
    except Exception as err:
        log_cloud_error("wedstrijdresultaat laden", err)
        return None


@st.cache_data(ttl=60, show_spinner=False)
def _fetch_match_results(team_id: str) -> list:
    client = get_supabase_client()
    if not team_id or client is None:
        return []
    try:
        response = (
            client.table("match_results")
            .select("*")
            .eq("team_id", team_id)
            .order("created_at", desc=True)
            .execute()
        )
        return response.data or []
    except Exception:
        return []


def cloud_list_match_results() -> list:
    tid = _active_team_id()
    result = _fetch_match_results(tid or "")
    if result is not None:
        mark_cloud_ok()
    return result or []


def render_match_management_screen() -> None:
    st.markdown("### 🏑 Wedstrijden & uitslagen")
    st.caption(
        "Voer hier de uitslag en details per wedstrijd in. "
        "De score wordt handmatig bevestigd — onafhankelijk van de getagde goals."
    )

    # --- Wedstrijd kiezen ---
    match_ids_cloud = list_match_ids_from_cloud(limit=50)
    meta_map = {}
    _tid = _active_team_id()
    _client = get_supabase_client()
    if _tid and _client:
        try:
            _resp = _client.table("match_meta").select(
                "match_id,opponent_name,own_team_name"
            ).eq("team_id", _tid).execute()
            for _r in (_resp.data or []):
                if _r.get("match_id"):
                    meta_map[_r["match_id"]] = {
                        "opponent_name": _r.get("opponent_name", ""),
                        "own_team_name": _r.get("own_team_name", ""),
                    }
        except Exception:
            pass

    if not match_ids_cloud:
        st.info("Nog geen wedstrijden gevonden. Start een wedstrijd via het startscherm.")
        return

    def _match_label(mid: str) -> str:
        meta = meta_map.get(mid, {})
        opp = meta.get("opponent_name") or "?"
        pretty = unscope_match_id(mid)
        return f"{pretty} — vs {opp}"

    match_options = {_match_label(m): m for m in match_ids_cloud}
    chosen_label = st.selectbox("Kies wedstrijd", list(match_options.keys()), key="mm_match_sel")
    chosen_mid = match_options[chosen_label]

    existing = cloud_get_match_result(chosen_mid)
    meta = meta_map.get(chosen_mid, {})

    st.divider()
    st.subheader("Uitslag & details invullen")

    with st.form("match_result_form"):
        opponent_default = (
            existing.get("opponent") if existing
            else meta.get("opponent_name", "")
        ) or ""
        opponent = st.text_input("Tegenstander", value=opponent_default)

        col1, col2 = st.columns(2)
        with col1:
            home_away_opts = ["Thuis", "Uit", "Neutraal"]
            ha_default = existing.get("home_away", "Thuis") if existing else "Thuis"
            ha_idx = home_away_opts.index(ha_default) if ha_default in home_away_opts else 0
            home_away = st.radio("Thuis / Uit", home_away_opts, index=ha_idx, horizontal=True)
            location = st.text_input(
                "Locatie / veld",
                value=existing.get("location", "") if existing else "",
            )

        with col2:
            our_score = st.number_input(
                "Onze score", min_value=0, max_value=99, step=1,
                value=int(existing["our_score"]) if existing and existing.get("our_score") is not None else 0,
            )
            opp_score = st.number_input(
                "Score tegenstander", min_value=0, max_value=99, step=1,
                value=int(existing["opp_score"]) if existing and existing.get("opp_score") is not None else 0,
            )

        match_notes = st.text_area(
            "Wedstrijdnotities / evaluatie",
            value=existing.get("notes", "") if existing else "",
            height=100,
        )

        saved = st.form_submit_button("💾 Opslaan", type="primary")

    if saved:
        if not opponent.strip():
            st.error("Vul de naam van de tegenstander in.")
        else:
            cloud_upsert_match_result(
                chosen_mid, opponent, home_away, location,
                int(our_score), int(opp_score), match_notes,
            )
            cloud_save_match_meta(chosen_mid, opponent,
                                  st.session_state.get("team_name", ""))
            st.success("Uitslag opgeslagen.")
            st.rerun()

    # --- Seizoensoverzicht ---
    st.divider()
    st.subheader("Seizoensoverzicht")

    all_results = cloud_list_match_results()

    if not all_results:
        st.info("Nog geen uitslagen ingevuld.")
        return

    def _result_label(r: dict) -> str:
        o = r.get("our_score")
        a = r.get("opp_score")
        if o is None or a is None:
            return "—"
        if o > a:
            return "🟢 W"
        elif o == a:
            return "🟡 G"
        return "🔴 V"

    wins = sum(1 for r in all_results if _result_label(r) == "🟢 W")
    draws = sum(1 for r in all_results if _result_label(r) == "🟡 G")
    losses = sum(1 for r in all_results if _result_label(r) == "🔴 V")
    goals_for = sum(int(r["our_score"]) for r in all_results if r.get("our_score") is not None)
    goals_against = sum(int(r["opp_score"]) for r in all_results if r.get("opp_score") is not None)

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Gespeeld", len(all_results))
    m2.metric("Gewonnen", wins)
    m3.metric("Gelijkgespeeld", draws)
    m4.metric("Verloren", losses)
    diff = goals_for - goals_against
    m5.metric("Doelsaldo", f"+{diff}" if diff > 0 else str(diff))

    rows = []
    for r in all_results:
        mid = r.get("match_id", "")
        pretty = unscope_match_id(mid)
        o = r.get("our_score")
        a = r.get("opp_score")
        score_str = f"{o}-{a}" if o is not None and a is not None else "—"
        rows.append({
            "Wedstrijd-ID": pretty,
            "Tegenstander": r.get("opponent", "—"),
            "T/U": r.get("home_away", "—"),
            "Locatie": r.get("location", "—"),
            "Score": score_str,
            "Resultaat": _result_label(r),
        })

    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)


# ==================================================
# SELECTIETOOL — visuele spelersopstelling per wedstrijd
# ==================================================

HOCKEY_POSITIONS = {
    "GK":  ("Keeper",         0.50, 0.93),
    "RB":  ("Rechts achter",  0.82, 0.75),
    "RCB": ("Rechts midden-achter", 0.63, 0.78),
    "LCB": ("Links midden-achter",  0.37, 0.78),
    "LB":  ("Links achter",   0.18, 0.75),
    "RM":  ("Rechts midden",  0.80, 0.52),
    "CM":  ("Centraal midden",0.50, 0.50),
    "LM":  ("Links midden",   0.20, 0.52),
    "RF":  ("Rechts voor",    0.75, 0.25),
    "CF":  ("Spits",          0.50, 0.22),
    "LF":  ("Links voor",     0.25, 0.25),
}

FORMATIES = {
    "4-3-3":  ["GK","LB","LCB","RCB","RB","LM","CM","RM","LF","CF","RF"],
    "4-4-2":  ["GK","LB","LCB","RCB","RB","LM","LCB","RCB","RM","CF","RF"],
    "3-5-2":  ["GK","LCB","CM","RCB","LM","LCB","CM","RM","RF","CF","LF"],
    "4-2-3-1":["GK","LB","LCB","RCB","RB","LM","RM","LF","CM","RF","CF"],
}


def cloud_save_selection(wedstrijd_id: str, selections: list) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or not client:
        return
    try:
        client.table("match_selection").delete().eq("wedstrijd_id", wedstrijd_id).eq("team_id", tid).execute()
        if selections:
            client.table("match_selection").insert(selections).execute()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("selectie opslaan", err)


def cloud_get_selection(wedstrijd_id: str) -> list:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or not client:
        return []
    try:
        r = client.table("match_selection").select("*").eq("wedstrijd_id", wedstrijd_id).eq("team_id", tid).execute()
        mark_cloud_ok()
        return r.data or []
    except Exception as err:
        log_cloud_error("selectie laden", err)
        return []


def _render_field_svg_selection(pos_player_map: dict) -> str:
    """Render hockey veld SVG met spelers op hun positie."""
    circles = ""
    for pos_key, (pos_label, rx, ry) in HOCKEY_POSITIONS.items():
        player_name = pos_player_map.get(pos_key, "")
        if not player_name:
            continue
        cx = rx * 300
        cy = ry * 420
        short = player_name.split()[-1][:10] if player_name else ""
        circles += f"""
        <circle cx="{cx}" cy="{cy}" r="16" fill="#3b82f6" stroke="white" stroke-width="1.5" opacity="0.92"/>
        <text x="{cx}" y="{cy+1}" text-anchor="middle" dominant-baseline="middle"
              fill="white" font-size="9" font-weight="700" font-family="Inter,sans-serif">{short}</text>
        <text x="{cx}" y="{cy+24}" text-anchor="middle"
              fill="#94a3b8" font-size="7.5" font-family="Inter,sans-serif">{pos_key}</text>
        """
    return f"""
    <svg viewBox="0 0 300 420" width="100%" style="max-width:360px;display:block;margin:0 auto">
      <rect width="300" height="420" rx="6" fill="#0a4f2e"/>
      <line x1="0" y1="210" x2="300" y2="210" stroke="white" stroke-width="0.8" opacity="0.4"/>
      <rect x="90" y="0" width="120" height="48" rx="2" fill="none" stroke="white" stroke-width="0.8" opacity="0.4"/>
      <rect x="90" y="372" width="120" height="48" rx="2" fill="none" stroke="white" stroke-width="0.8" opacity="0.4"/>
      <rect x="30" y="0" width="240" height="95" rx="2" fill="none" stroke="white" stroke-width="0.6" opacity="0.3"/>
      <rect x="30" y="325" width="240" height="95" rx="2" fill="none" stroke="white" stroke-width="0.6" opacity="0.3"/>
      <circle cx="150" cy="210" r="30" fill="none" stroke="white" stroke-width="0.8" opacity="0.4"/>
      {circles}
    </svg>"""


def render_selection_screen() -> None:
    st.markdown("### 👥 Selectietool")

    roster = _active_team_roster()
    if not roster:
        st.info("Voeg eerst spelers toe via het Wisselschema.")
        return

    # Kies wedstrijd
    match_ids = list_match_ids_from_cloud(limit=30)
    tid = _active_team_id()
    client = get_supabase_client()
    meta_map = {}
    if tid and client:
        try:
            resp = client.table("match_meta").select("match_id,opponent_name").eq("team_id", tid).execute()
            for r in (resp.data or []):
                if r.get("match_id"):
                    meta_map[r["match_id"]] = r.get("opponent_name", "?")
        except Exception:
            pass

    if not match_ids:
        st.info("Nog geen wedstrijden gevonden. Start er een via de Wedstrijd analyse tool.")
        return

    def _mlabel(mid):
        return f"{unscope_match_id(mid)} — vs {meta_map.get(mid, '?')}"

    opts = {_mlabel(m): m for m in match_ids}
    chosen = st.selectbox("Wedstrijd", list(opts.keys()), key="sel_match_pick")
    wedstrijd_id = opts[chosen]

    formatie = st.selectbox("Formatie", list(FORMATIES.keys()), key="sel_formatie")
    positie_keys = FORMATIES[formatie]

    # Laad bestaande selectie
    bestaand = {s["positie"]: s["speler_naam"] for s in cloud_get_selection(wedstrijd_id)}

    speler_namen = ["—"] + sorted([p["name"] for p in roster])

    st.divider()
    col_form, col_field = st.columns([1, 1], gap="large")

    pos_player_map = {}
    with col_form:
        st.subheader("Basisopstelling")
        with st.form("selectie_form"):
            for pos_key in positie_keys:
                label = HOCKEY_POSITIONS[pos_key][0]
                default_idx = 0
                default_naam = bestaand.get(pos_key, "")
                if default_naam in speler_namen:
                    default_idx = speler_namen.index(default_naam)
                gekozen = st.selectbox(
                    f"{pos_key} — {label}",
                    speler_namen,
                    index=default_idx,
                    key=f"sel_pos_{pos_key}",
                )
                if gekozen != "—":
                    pos_player_map[pos_key] = gekozen

            st.markdown("**Reserves**")
            geselecteerd = set(pos_player_map.values())
            reserves = st.multiselect(
                "Reserves / wisselspelers",
                [p["name"] for p in roster if p["name"] not in geselecteerd],
                default=[p for p in bestaand.get("__reserves__", "").split(",") if p],
                key="sel_reserves",
            )
            opslaan = st.form_submit_button("💾 Opslaan", type="primary")

        if opslaan:
            rows = [
                {"wedstrijd_id": wedstrijd_id, "team_id": _active_team_id(),
                 "positie": pos, "speler_naam": naam, "is_reserve": False}
                for pos, naam in pos_player_map.items()
            ]
            for r in reserves:
                rows.append({"wedstrijd_id": wedstrijd_id, "team_id": _active_team_id(),
                              "positie": "__reserves__", "speler_naam": r, "is_reserve": True})
            cloud_save_selection(wedstrijd_id, rows)
            st.success("Selectie opgeslagen!")
            st.rerun()

    with col_field:
        st.subheader("Veldopstelling")
        svg = _render_field_svg_selection(pos_player_map)
        components.html(
            f'<div style="background:#080c18;padding:8px;border-radius:12px;">{svg}</div>',
            height=450,
        )
        if reserves:
            st.markdown(
                f'<div style="text-align:center;margin-top:12px;color:#94a3b8;font-size:12px;">'
                f'<b>Reserves:</b> {", ".join(reserves)}</div>',
                unsafe_allow_html=True,
            )
        if pos_player_map:
            wa_text = whatsapp_text_selectie(wedstrijd_id, pos_player_map, reserves)
            wa_url = _whatsapp_url(wa_text)
            st.markdown(
                f'<a href="{wa_url}" target="_blank" style="display:inline-flex;align-items:center;'
                f'gap:8px;background:#25d36622;border:1px solid #25d36644;color:#25d366;'
                f'padding:8px 16px;border-radius:10px;font-weight:600;font-size:13px;'
                f'text-decoration:none;margin-top:12px;">'
                f'<span style="font-size:18px;">💬</span> Selectie delen via WhatsApp</a>',
                unsafe_allow_html=True,
            )


# ==================================================
# TRAININGSPLANNING — sessies plannen en oefeningen koppelen
# ==================================================

TRAINING_THEMAS = ["Passing", "Pressing", "Cirkelspel", "Corners", "Vrij spel",
                   "Positiespel", "Conditie", "Warming-up", "Tactiek", "Anders"]
OEFENING_CATEGORIEEN = ["Warming-up", "Techniek", "Tactiek", "Conditie", "Afwerking", "Cooling-down"]


def cloud_save_training(data: dict) -> str | None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or not client:
        return None
    try:
        data["team_id"] = tid
        if not data.get("id"):
            data["id"] = str(uuid.uuid4())
        client.table("training_sessions").upsert(data, on_conflict="id").execute()
        _fetch_trainings.clear()
        mark_cloud_ok()
        return data["id"]
    except Exception as err:
        log_cloud_error("training opslaan", err)
        return None


@st.cache_data(ttl=30, show_spinner=False)
def _fetch_trainings(team_id: str, limit: int) -> list:
    client = get_supabase_client()
    if not team_id or not client:
        return []
    try:
        r = client.table("training_sessions").select("*").eq("team_id", team_id)\
            .order("datum", desc=True).limit(limit).execute()
        return r.data or []
    except Exception:
        return []


def cloud_list_trainings(limit: int = 30) -> list:
    tid = _active_team_id()
    result = _fetch_trainings(tid or "", limit)
    if result is not None:
        mark_cloud_ok()
    return result or []


def cloud_delete_training(training_id: str) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or not client:
        return
    try:
        client.table("training_sessions").delete().eq("id", training_id).eq("team_id", tid).execute()
        _fetch_trainings.clear()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("training verwijderen", err)


def cloud_save_exercise(data: dict) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or not client:
        return
    try:
        data["team_id"] = tid
        if not data.get("id"):
            data["id"] = str(uuid.uuid4())
        client.table("training_exercises").upsert(data, on_conflict="id").execute()
        _fetch_exercises.clear()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("oefening opslaan", err)


@st.cache_data(ttl=60, show_spinner=False)
def _fetch_exercises(team_id: str) -> list:
    client = get_supabase_client()
    if not team_id or not client:
        return []
    try:
        r = client.table("training_exercises").select("*").eq("team_id", team_id)\
            .order("naam").execute()
        return r.data or []
    except Exception:
        return []


def cloud_list_exercises() -> list:
    tid = _active_team_id()
    result = _fetch_exercises(tid or "")
    if result is not None:
        mark_cloud_ok()
    return result or []


def render_training_screen() -> None:
    import datetime as _dt
    st.markdown("### 📅 Trainingsplanning")

    tab_plan, tab_aanwezig, tab_oefeningen, tab_export = st.tabs(
        ["📋 Sessies", "✅ Aanwezigheid", "🏃 Oefeningen bibliotheek", "📄 PDF export"]
    )

    # ─── TAB 1: Sessies ───
    with tab_plan:
        col_list, col_form = st.columns([1.2, 1], gap="large")

        with col_form:
            st.subheader("Nieuwe sessie")
            with st.form("training_form"):
                datum = st.date_input("Datum", value=_dt.date.today(), key="tr_datum")
                c1, c2 = st.columns(2)
                start = c1.time_input("Starttijd", value=_dt.time(18, 0), key="tr_start")
                eind = c2.time_input("Eindtijd", value=_dt.time(19, 30), key="tr_eind")
                locatie = st.text_input("Locatie / veld", placeholder="Hoofdveld 1", key="tr_loc")
                thema = st.selectbox("Thema", TRAINING_THEMAS, key="tr_thema")
                notities = st.text_area("Notities / doelen", height=90, key="tr_notes",
                                        placeholder="Wat wil je bereiken in deze training?")

                # Oefeningen koppelen
                oefeningen = cloud_list_exercises()
                oefen_namen = {o["naam"]: o["id"] for o in oefeningen}
                gekoppeld = st.multiselect("Oefeningen koppelen (optioneel)",
                                           list(oefen_namen.keys()), key="tr_oefen")
                opslaan = st.form_submit_button("➕ Toevoegen", type="primary")

            if opslaan:
                cloud_save_training({
                    "datum": datum.isoformat(),
                    "starttijd": start.strftime("%H:%M"),
                    "eindtijd": eind.strftime("%H:%M"),
                    "locatie": locatie,
                    "thema": thema,
                    "notities": notities,
                    "oefening_ids": [oefen_namen[n] for n in gekoppeld],
                })
                st.success("Training gepland!")
                st.rerun()

        with col_list:
            st.subheader("Geplande sessies")
            trainingen = cloud_list_trainings()
            if not trainingen:
                st.info("Nog geen trainingen gepland.")
            else:
                import datetime as _dt2
                today = _dt2.date.today()
                komend = [t for t in trainingen if t.get("datum", "") >= today.isoformat()]
                verleden = [t for t in trainingen if t.get("datum", "") < today.isoformat()]

                if komend:
                    st.markdown('<div class="cs-section-label">Komende trainingen</div>',
                                unsafe_allow_html=True)
                    for t in komend[:10]:
                        _render_training_card(t)

                if verleden:
                    with st.expander(f"Vorige trainingen ({len(verleden)})"):
                        for t in verleden[:15]:
                            _render_training_card(t)

    # ─── TAB 2: Oefeningen ───
    with tab_oefeningen:
        col_lib, col_add = st.columns([1.2, 1], gap="large")

        with col_add:
            st.subheader("Oefening toevoegen")
            with st.form("oefening_form"):
                naam = st.text_input("Naam", placeholder="Bijv. 4-hoeken passing", key="oe_naam")
                cat = st.selectbox("Categorie", OEFENING_CATEGORIEEN, key="oe_cat")
                duur = st.number_input("Duur (minuten)", min_value=5, max_value=60,
                                       value=15, step=5, key="oe_duur")
                c1, c2 = st.columns(2)
                min_sp = c1.number_input("Min. spelers", min_value=2, value=6, key="oe_min")
                max_sp = c2.number_input("Max. spelers", min_value=2, value=20, key="oe_max")
                beschrijving = st.text_area("Beschrijving / uitleg", height=100, key="oe_desc")
                materiaal = st.text_input("Materiaal", placeholder="Bijv. ballen, pionnen, goals",
                                          key="oe_mat")
                add_oe = st.form_submit_button("➕ Toevoegen", type="primary")

            if add_oe:
                if not naam.strip():
                    st.error("Vul een naam in.")
                else:
                    cloud_save_exercise({
                        "naam": naam.strip(),
                        "categorie": cat,
                        "duur_minuten": int(duur),
                        "min_spelers": int(min_sp),
                        "max_spelers": int(max_sp),
                        "beschrijving": beschrijving.strip(),
                        "materiaal": materiaal.strip(),
                    })
                    st.success("Oefening opgeslagen!")
                    st.rerun()

        with col_lib:
            st.subheader("Oefeningen bibliotheek")
            oefeningen = cloud_list_exercises()
            if not oefeningen:
                st.info("Nog geen oefeningen toegevoegd.")
            else:
                cat_filter = st.multiselect("Filter", OEFENING_CATEGORIEEN,
                                            default=OEFENING_CATEGORIEEN, key="oe_filter")
                for o in oefeningen:
                    if o.get("categorie") not in cat_filter:
                        continue
                    cat_colors = {
                        "Warming-up": "#f59e0b", "Techniek": "#3b82f6",
                        "Tactiek": "#8b5cf6", "Conditie": "#10b981",
                        "Afwerking": "#f43f5e", "Cooling-down": "#64748b",
                    }
                    c = cat_colors.get(o.get("categorie", ""), "#64748b")
                    with st.expander(f"**{o['naam']}** — {o.get('categorie','')}"
                                     f" · {o.get('duur_minuten','')} min"):
                        st.markdown(
                            f'<span style="background:{c}22;color:{c};padding:2px 8px;'
                            f'border-radius:6px;font-size:11px;font-weight:700;">'
                            f'{o.get("categorie","")}</span>',
                            unsafe_allow_html=True,
                        )
                        if o.get("beschrijving"):
                            st.write(o["beschrijving"])
                        col1, col2, col3 = st.columns(3)
                        col1.metric("Duur", f"{o.get('duur_minuten','')} min")
                        col2.metric("Spelers", f"{o.get('min_spelers','')}–{o.get('max_spelers','')}")
                        if o.get("materiaal"):
                            col3.metric("Materiaal", o["materiaal"])

    # ─── TAB: Aanwezigheid ───
    with tab_aanwezig:
        trainingen_all = cloud_list_trainings(limit=50)
        roster = _active_team_roster()
        if not trainingen_all:
            st.info("Nog geen trainingen gepland.")
        elif not roster:
            st.info("Voeg eerst spelers toe via het Wisselschema.")
        else:
            import datetime as _dt2
            today = _dt2.date.today()
            # Toon afgelopen + huidige trainingen voor aanwezigheid
            verleden = [t for t in trainingen_all if t.get("datum", "") <= today.isoformat()]
            if not verleden:
                st.info("Nog geen trainingen geweest. Aanwezigheid bijhouden kan na de training.")
            else:
                gekozen_tr = st.selectbox(
                    "Kies training",
                    verleden,
                    format_func=lambda t: f"{t.get('datum','')} — {t.get('thema','')} ({t.get('starttijd','')})",
                    key="aanw_tr_sel"
                )
                training_id = gekozen_tr["id"]
                aanwezig_records = cloud_get_attendance(training_id)
                aanwezig_ids = {r["player_id"] for r in aanwezig_records}

                sorted_roster = sorted(roster, key=lambda p: p["name"])
                st.markdown(f"**Aanwezigheid voor: {gekozen_tr.get('datum','')} — {gekozen_tr.get('thema','')}**")

                nieuw_aanwezig = []
                cols = st.columns(2)
                for i, speler in enumerate(sorted_roster):
                    was_er = speler["id"] in aanwezig_ids
                    aanw = cols[i % 2].checkbox(
                        speler["name"], value=was_er, key=f"aanw_{training_id}_{speler['id']}"
                    )
                    if aanw:
                        nieuw_aanwezig.append(speler["id"])

                aanwezig_count = len(nieuw_aanwezig)
                st.caption(f"{aanwezig_count} van {len(roster)} spelers aanwezig")

                if st.button("💾 Aanwezigheid opslaan", type="primary", key="aanw_opslaan"):
                    cloud_save_attendance(training_id, nieuw_aanwezig)
                    st.success(f"Aanwezigheid opgeslagen — {aanwezig_count} spelers aanwezig!")
                    st.rerun()

    # ─── TAB: PDF Export ───
    with tab_export:
        st.subheader("Trainingsplan exporteren als PDF")
        trainingen_pdf = cloud_list_trainings(limit=50)
        oefeningen_pdf = cloud_list_exercises()
        if not trainingen_pdf:
            st.info("Nog geen trainingen om te exporteren.")
        else:
            if REPORTLAB_AVAILABLE:
                pdf_bytes = generate_training_pdf(trainingen_pdf, oefeningen_pdf)
                if pdf_bytes:
                    team_name = st.session_state.get("active_team_name", "team")
                    st.download_button(
                        "📄 Download trainingsplan PDF",
                        data=pdf_bytes,
                        file_name=f"trainingsplan_{team_name}.pdf",
                        mime="application/pdf",
                        type="primary",
                        use_container_width=True,
                    )
                    st.caption(f"{len(trainingen_pdf)} trainingen · {len(oefeningen_pdf)} oefeningen")
            else:
                st.warning("ReportLab niet beschikbaar.")


def _render_training_card(t: dict) -> None:
    import datetime as _dt
    datum_str = t.get("datum", "")
    try:
        d = _dt.date.fromisoformat(datum_str)
        dag = ["Ma","Di","Wo","Do","Vr","Za","Zo"][d.weekday()]
        datum_fmt = f"{dag} {d.day} {['jan','feb','mrt','apr','mei','jun','jul','aug','sep','okt','nov','dec'][d.month-1]}"
    except Exception:
        datum_fmt = datum_str

    thema_icons = {
        "Passing": "🎯", "Pressing": "⚡", "Cirkelspel": "⭕",
        "Corners": "🔱", "Vrij spel": "🏑", "Positiespel": "📐",
        "Conditie": "💪", "Warming-up": "🔥", "Tactiek": "🧠", "Anders": "📋",
    }
    icon = thema_icons.get(t.get("thema", ""), "📋")
    start = t.get("starttijd", "")[:5]
    eind = t.get("eindtijd", "")[:5]
    tijd = f"{start}–{eind}" if start and eind else start

    maand_kort = ["jan","feb","mrt","apr","mei","jun","jul","aug","sep","okt","nov","dec"]
    try:
        dag_num = _dt.date.fromisoformat(datum_str).day
        maand_str = maand_kort[_dt.date.fromisoformat(datum_str).month - 1]
    except Exception:
        dag_num = ""
        maand_str = ""

    st.markdown(
        f'<div style="background:#0f1624;border:1px solid #1a2540;border-radius:12px;'
        f'padding:14px 18px;margin-bottom:4px;display:flex;align-items:center;gap:14px;">'
        f'<div style="background:rgba(59,130,246,0.1);border:1px solid rgba(59,130,246,0.2);'
        f'border-radius:10px;padding:10px 14px;text-align:center;min-width:54px;">'
        f'<div style="color:#3b82f6;font-weight:800;font-size:16px;">{dag_num}</div>'
        f'<div style="color:#64748b;font-size:10px;text-transform:uppercase;">{maand_str}</div>'
        f'</div>'
        f'<div style="flex:1;">'
        f'<div style="color:#f1f5f9;font-weight:600;font-size:14px;">{icon} {t.get("thema","Training")}</div>'
        f'<div style="color:#64748b;font-size:12px;margin-top:2px;">{tijd} · {t.get("locatie","")}</div>'
        f'</div>'
        f'</div>',
        unsafe_allow_html=True,
    )
    col_wa, col_del = st.columns([4, 1])
    with col_wa:
        wa_url = _whatsapp_url(whatsapp_text_training(t))
        st.markdown(
            f'<a href="{wa_url}" target="_blank" style="display:inline-flex;align-items:center;'
            f'gap:6px;background:#25d36614;border:1px solid #25d36630;color:#25d366;'
            f'padding:5px 12px;border-radius:8px;font-weight:600;font-size:12px;'
            f'text-decoration:none;">💬 Deel via WhatsApp</a>',
            unsafe_allow_html=True,
        )
    with col_del:
        if st.button("🗑️", key=f"del_tr_{t['id']}", help="Verwijder"):
            cloud_delete_training(t["id"])
            st.rerun()


# ==================================================
# BLESSURE TRACKER — per speler bijhouden
# ==================================================

BLESSURE_TYPES = ["Hamstring", "Enkel", "Knie", "Rug", "Schouder", "Lies",
                  "Scheenbeen", "Voet", "Hoofd/nekk", "Anders"]
BLESSURE_ERNST = ["Licht (< 1 week)", "Matig (1–4 weken)", "Ernstig (> 4 weken)"]


def cloud_add_injury(player_id: str, player_name: str, data: dict) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or not client:
        return
    try:
        data.update({"id": str(uuid.uuid4()), "team_id": tid,
                     "player_id": player_id, "player_name": player_name})
        client.table("player_injuries").insert(data).execute()
        _fetch_injuries.clear()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("blessure opslaan", err)


@st.cache_data(ttl=30, show_spinner=False)
def _fetch_injuries(team_id: str, active_only: bool) -> list:
    client = get_supabase_client()
    if not team_id or not client:
        return []
    try:
        q = client.table("player_injuries").select("*").eq("team_id", team_id)
        if active_only:
            q = q.is_("datum_herstel", "null")
        r = q.order("datum_start", desc=True).execute()
        return r.data or []
    except Exception:
        return []


def cloud_list_injuries(active_only: bool = False) -> list:
    tid = _active_team_id()
    result = _fetch_injuries(tid or "", active_only)
    if result is not None:
        mark_cloud_ok()
    return result or []


def cloud_resolve_injury(injury_id: str) -> None:
    import datetime as _dt
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or not client:
        return
    try:
        client.table("player_injuries").update({
            "datum_herstel": _dt.date.today().isoformat()
        }).eq("id", injury_id).eq("team_id", tid).execute()
        _fetch_injuries.clear()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("blessure oplossen", err)


def render_injury_screen() -> None:
    import datetime as _dt
    st.markdown("### 🩹 Blessure tracker")

    roster = _active_team_roster()
    actieve_blessures = cloud_list_injuries(active_only=True)
    blessure_spelers = {b["player_id"] for b in actieve_blessures}

    # ─── Status overzicht ───
    cols = st.columns(3)
    fit_count = sum(1 for p in roster if p["id"] not in blessure_spelers)
    cols[0].metric("✅ Fit", fit_count)
    cols[1].metric("🩹 Geblesseerd", len(actieve_blessures))
    cols[2].metric("👥 Totaal squad", len(roster))

    st.divider()
    tab_actief, tab_add, tab_historie, tab_pdf = st.tabs(["🔴 Actieve blessures", "➕ Toevoegen", "📋 Historie", "📄 PDF"])

    with tab_actief:
        if not actieve_blessures:
            st.success("Geen actieve blessures — iedereen is fit! 🎉")
        else:
            for b in actieve_blessures:
                ernst_colors = {
                    "Licht (< 1 week)": "#10b981",
                    "Matig (1–4 weken)": "#f59e0b",
                    "Ernstig (> 4 weken)": "#f43f5e",
                }
                color = ernst_colors.get(b.get("ernst", ""), "#64748b")
                start = b.get("datum_start", "")
                verwacht = b.get("verwachte_terugkeer", "—") or "—"

                st.markdown(
                    f'<div style="background:#0f1624;border:1px solid {color}44;'
                    f'border-left:3px solid {color};border-radius:12px;'
                    f'padding:14px 18px;margin-bottom:8px;">'
                    f'<div style="display:flex;justify-content:space-between;align-items:center;">'
                    f'<div><span style="color:#f1f5f9;font-weight:700;font-size:15px;">'
                    f'{b.get("player_name","?")}</span>'
                    f'<span style="background:{color}22;color:{color};padding:2px 8px;'
                    f'border-radius:6px;font-size:11px;font-weight:700;margin-left:10px;">'
                    f'{b.get("blessure_type","")}</span></div>'
                    f'<div style="color:#64748b;font-size:12px;">Sinds {start}</div>'
                    f'</div>'
                    f'<div style="color:#94a3b8;font-size:12px;margin-top:6px;">'
                    f'{b.get("ernst","")} · Verwachte terugkeer: <b style="color:#f1f5f9">{verwacht}</b></div>'
                    f'{"<div style=color:#94a3b8;font-size:12px;margin-top:4px>" + b.get("beschrijving","") + "</div>" if b.get("beschrijving") else ""}'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                if st.button("✅ Hersteld", key=f"herstel_{b['id']}", help="Markeer als hersteld"):
                    cloud_resolve_injury(b["id"])
                    st.success(f"{b.get('player_name','Speler')} is hersteld!")
                    st.rerun()

    with tab_add:
        if not roster:
            st.info("Voeg eerst spelers toe.")
        else:
            with st.form("blessure_form"):
                speler_map = {p["name"]: p["id"] for p in sorted(roster, key=lambda x: x["name"])}
                speler_naam = st.selectbox("Speler", list(speler_map.keys()), key="bl_speler")
                c1, c2 = st.columns(2)
                btype = c1.selectbox("Type blessure", BLESSURE_TYPES, key="bl_type")
                ernst = c2.selectbox("Ernst", BLESSURE_ERNST, key="bl_ernst")
                c3, c4 = st.columns(2)
                datum_start = c3.date_input("Datum blessure", value=_dt.date.today(), key="bl_start")
                datum_terug = c4.date_input("Verwachte terugkeer (schatting)",
                                             value=_dt.date.today() + _dt.timedelta(weeks=2),
                                             key="bl_terug")
                beschrijving = st.text_area("Beschrijving / omstandigheden", height=80, key="bl_desc")
                behandeling = st.text_input("Behandeling", placeholder="Bijv. fysiotherapie, rust",
                                             key="bl_behandeling")
                toevoegen = st.form_submit_button("🩹 Registreren", type="primary")

            if toevoegen:
                cloud_add_injury(speler_map[speler_naam], speler_naam, {
                    "blessure_type": btype,
                    "ernst": ernst,
                    "datum_start": datum_start.isoformat(),
                    "verwachte_terugkeer": datum_terug.isoformat(),
                    "beschrijving": beschrijving.strip(),
                    "behandeling": behandeling.strip(),
                })
                st.success(f"Blessure van {speler_naam} geregistreerd.")
                st.rerun()

    with tab_historie:
        alle = cloud_list_injuries(active_only=False)
        hersteld = [b for b in alle if b.get("datum_herstel")]
        if not hersteld:
            st.info("Nog geen herstelde blessures in de historie.")
        else:
            for b in hersteld:
                st.markdown(
                    f'<div style="background:#0f1624;border:1px solid #1a2540;border-radius:10px;'
                    f'padding:12px 16px;margin-bottom:6px;opacity:0.75;">'
                    f'<b style="color:#f1f5f9">{b.get("player_name","")}</b> · '
                    f'<span style="color:#94a3b8">{b.get("blessure_type","")} · '
                    f'{b.get("datum_start","")} → {b.get("datum_herstel","")}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

    with tab_pdf:
        st.subheader("Blessurerapport exporteren")
        alle_bl = cloud_list_injuries(active_only=False)
        if not alle_bl:
            st.info("Nog geen blessures om te exporteren.")
        elif REPORTLAB_AVAILABLE:
            pdf_bytes = generate_blessures_pdf(alle_bl)
            if pdf_bytes:
                team_name = st.session_state.get("active_team_name", "team")
                st.download_button(
                    "📄 Download blessurerapport PDF",
                    data=pdf_bytes,
                    file_name=f"blessurerapport_{team_name}.pdf",
                    mime="application/pdf",
                    type="primary",
                    use_container_width=True,
                )
                actief_count = len([b for b in alle_bl if not b.get("datum_herstel")])
                st.caption(f"{actief_count} actieve blessures · {len(alle_bl) - actief_count} hersteld")
        else:
            st.warning("ReportLab niet beschikbaar.")


# ==================================================
# SCOUTING — tegenstander analyse
# ==================================================

FORMATIES_SCOUTING = ["4-3-3", "4-4-2", "3-5-2", "4-2-3-1", "5-3-2", "4-1-4-1", "Anders"]
SPEELSTIJLEN = ["Hoog pressing", "Laag blok", "Snel counteren", "Balbezit", "Directe ballen", "Gemengd"]


def cloud_save_scouting(data: dict) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or not client:
        return
    try:
        data["team_id"] = tid
        if not data.get("id"):
            data["id"] = str(uuid.uuid4())
        client.table("scouting_reports").upsert(data, on_conflict="id").execute()
        _fetch_scouting.clear()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("scouting opslaan", err)


@st.cache_data(ttl=60, show_spinner=False)
def _fetch_scouting(team_id: str) -> list:
    client = get_supabase_client()
    if not team_id or not client:
        return []
    try:
        r = client.table("scouting_reports").select("*").eq("team_id", team_id)\
            .order("created_at", desc=True).execute()
        return r.data or []
    except Exception:
        return []


def cloud_list_scouting() -> list:
    tid = _active_team_id()
    result = _fetch_scouting(tid or "")
    if result is not None:
        mark_cloud_ok()
    return result or []


def render_scouting_screen() -> None:
    st.markdown("### 🔍 Tegenstander scouting")

    tab_overzicht, tab_nieuw = st.tabs(["📋 Overzicht", "➕ Nieuw rapport"])
    # PDF download knop bovenaan
    if REPORTLAB_AVAILABLE:
        rapporten_all = cloud_list_scouting()
        if rapporten_all:
            gekozen_sc = st.selectbox(
                "Exporteer rapport als PDF",
                rapporten_all,
                format_func=lambda r: f"{r.get('tegenstander','?')} · {r.get('wedstrijd_datum','')}",
                key="sc_pdf_sel"
            )
            pdf_sc = generate_scouting_pdf(gekozen_sc)
            if pdf_sc:
                teg_naam = gekozen_sc.get("tegenstander", "scouting").replace(" ", "_")
                st.download_button(
                    "📄 Download scoutingrapport PDF",
                    data=pdf_sc,
                    file_name=f"scouting_{teg_naam}.pdf",
                    mime="application/pdf",
                )

    with tab_nieuw:
        with st.form("scouting_form"):
            tegenstander = st.text_input("Tegenstander", placeholder="Bijv. HC Rotterdam", key="sc_teg")
            c1, c2 = st.columns(2)
            wedstrijd_datum = c1.date_input("Geobserveerde wedstrijd", key="sc_datum")
            formatie = c2.selectbox("Formatie tegenstander", FORMATIES_SCOUTING, key="sc_form")
            speelstijl = st.selectbox("Speelstijl", SPEELSTIJLEN, key="sc_stijl")

            st.markdown("**Analyse**")
            c3, c4 = st.columns(2)
            sterk = c3.text_area("💪 Sterke punten", height=100, key="sc_sterk",
                                  placeholder="Waar zijn ze goed in?")
            zwak = c4.text_area("⚠️ Zwakke punten", height=100, key="sc_zwak",
                                 placeholder="Waar kunnen wij van profiteren?")

            st.markdown("**Standaardsituaties**")
            c5, c6 = st.columns(2)
            corner_aan = c5.text_area("🔱 Corners aanval", height=80, key="sc_corn_a",
                                       placeholder="Hoe nemen ze corners?")
            corner_verd = c6.text_area("🛡️ Corners verdediging", height=80, key="sc_corn_v",
                                        placeholder="Hoe verdedigen ze corners?")

            aanpak = st.text_area("🎯 Aanbevolen aanpak voor ons team", height=100,
                                   key="sc_aanpak",
                                   placeholder="Welke tactiek werkt het beste tegen hen?")

            opslaan = st.form_submit_button("💾 Opslaan", type="primary")

        if opslaan:
            if not tegenstander.strip():
                st.error("Vul de naam van de tegenstander in.")
            else:
                cloud_save_scouting({
                    "tegenstander": tegenstander.strip(),
                    "wedstrijd_datum": wedstrijd_datum.isoformat(),
                    "formatie": formatie,
                    "speelstijl": speelstijl,
                    "sterke_punten": sterk.strip(),
                    "zwakke_punten": zwak.strip(),
                    "corners_aanval": corner_aan.strip(),
                    "corners_verdediging": corner_verd.strip(),
                    "aanbevolen_aanpak": aanpak.strip(),
                })
                st.success("Scoutingrapport opgeslagen!")
                st.rerun()

    with tab_overzicht:
        rapporten = cloud_list_scouting()
        if not rapporten:
            st.info("Nog geen scoutingrapporten. Voeg er een toe via de tab hiernaast.")
        else:
            for r in rapporten:
                formatie_badge = r.get("formatie", "")
                stijl_badge = r.get("speelstijl", "")
                datum = r.get("wedstrijd_datum", "")
                with st.expander(f"**{r.get('tegenstander','?')}** · {formatie_badge} · {datum}"):
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(f"**Formatie:** `{formatie_badge}`")
                        st.markdown(f"**Speelstijl:** {stijl_badge}")
                        if r.get("sterke_punten"):
                            st.markdown("**💪 Sterke punten:**")
                            st.info(r["sterke_punten"])
                        if r.get("zwakke_punten"):
                            st.markdown("**⚠️ Zwakke punten:**")
                            st.warning(r["zwakke_punten"])
                    with c2:
                        if r.get("corners_aanval"):
                            st.markdown("**🔱 Corners aanval:**")
                            st.write(r["corners_aanval"])
                        if r.get("corners_verdediging"):
                            st.markdown("**🛡️ Corners verdediging:**")
                            st.write(r["corners_verdediging"])
                        if r.get("aanbevolen_aanpak"):
                            st.markdown("**🎯 Aanpak voor ons:**")
                            st.success(r["aanbevolen_aanpak"])


# ==================================================
# WHATSAPP DELEN
# ==================================================
def _whatsapp_url(text: str) -> str:
    import urllib.parse
    return f"https://wa.me/?text={urllib.parse.quote(text)}"


def whatsapp_text_selectie(wedstrijd_id: str, pos_player_map: dict, reserves: list) -> str:
    team_name = st.session_state.get("active_team_name", "Team")
    mid_label = unscope_match_id(wedstrijd_id)
    lines = [f"⚽ *Selectie {team_name}*", f"📅 Wedstrijd: {mid_label}", ""]
    lines.append("*Basisopstelling:*")
    for pos, naam in pos_player_map.items():
        lines.append(f"  {pos}: {naam}")
    if reserves:
        lines.append("")
        lines.append("*Reserves:*")
        for r in reserves:
            lines.append(f"  • {r}")
    lines.append("\n_Via Coach Studio_")
    return "\n".join(lines)


def whatsapp_text_training(t: dict) -> str:
    team_name = st.session_state.get("active_team_name", "Team")
    lines = [
        f"🏑 *Training {team_name}*",
        f"📅 {t.get('datum', '')} · {t.get('starttijd', '')}–{t.get('eindtijd', '')}",
    ]
    if t.get("locatie"):
        lines.append(f"📍 {t['locatie']}")
    if t.get("thema"):
        lines.append(f"🎯 Thema: {t['thema']}")
    if t.get("notities"):
        lines.append(f"📝 {t['notities']}")
    lines.append("\n_Via Coach Studio_")
    return "\n".join(lines)


# ==================================================
# AANWEZIGHEIDSREGISTRATIE
# ==================================================
@st.cache_data(ttl=30, show_spinner=False)
def _fetch_attendance(team_id: str, training_id: str) -> list:
    client = get_supabase_client()
    if not team_id or not training_id or client is None:
        return []
    try:
        r = client.table("training_attendance").select("*")\
            .eq("team_id", team_id).eq("training_id", training_id).execute()
        return r.data or []
    except Exception:
        return []


def cloud_get_attendance(training_id: str) -> list:
    tid = _active_team_id()
    return _fetch_attendance(tid or "", training_id)


def cloud_save_attendance(training_id: str, aanwezige_ids: list) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or not client:
        return
    try:
        client.table("training_attendance").delete()\
            .eq("team_id", tid).eq("training_id", training_id).execute()
        if aanwezige_ids:
            roster = _active_team_roster()
            id_to_name = {p["id"]: p["name"] for p in roster}
            rows = [
                {"id": str(uuid.uuid4()), "team_id": tid, "training_id": training_id,
                 "player_id": pid, "player_name": id_to_name.get(pid, "?")}
                for pid in aanwezige_ids
            ]
            client.table("training_attendance").insert(rows).execute()
        _fetch_attendance.clear()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("aanwezigheid opslaan", err)


# ==================================================
# PDF EXPORTS — Training, Blessures, Scouting
# ==================================================
def _pdf_styles():
    """Gedeelde ReportLab stijlen."""
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors as _c
    styles = getSampleStyleSheet()
    navy = _c.HexColor("#080c18")
    blue = _c.HexColor("#3b82f6")
    white = _c.white
    gray = _c.HexColor("#64748b")

    title_style = ParagraphStyle("cs_title", parent=styles["Title"],
        fontSize=22, textColor=blue, spaceAfter=4, fontName="Helvetica-Bold")
    sub_style = ParagraphStyle("cs_sub", parent=styles["Normal"],
        fontSize=10, textColor=gray, spaceAfter=16)
    h2_style = ParagraphStyle("cs_h2", parent=styles["Heading2"],
        fontSize=13, textColor=blue, spaceBefore=14, spaceAfter=6, fontName="Helvetica-Bold")
    body_style = ParagraphStyle("cs_body", parent=styles["Normal"],
        fontSize=10, textColor=_c.HexColor("#1e293b"), leading=14, spaceAfter=4)
    label_style = ParagraphStyle("cs_label", parent=styles["Normal"],
        fontSize=9, textColor=gray, fontName="Helvetica-Bold",
        textTransform="uppercase", spaceAfter=2)
    return {"title": title_style, "sub": sub_style, "h2": h2_style,
            "body": body_style, "label": label_style}


def generate_training_pdf(trainingen: list, oefeningen: list) -> bytes | None:
    if not REPORTLAB_AVAILABLE:
        return None
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors as _c
    from io import BytesIO
    import datetime as _dt

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm,
                             topMargin=20*mm, bottomMargin=20*mm)
    S = _pdf_styles()
    team_name = st.session_state.get("active_team_name", "Team")
    story = [
        Paragraph("Coach Studio", S["title"]),
        Paragraph(f"Trainingsplanning — {team_name} · {_dt.date.today().strftime('%d %b %Y')}", S["sub"]),
        HRFlowable(width="100%", thickness=1, color=_c.HexColor("#3b82f6"), spaceAfter=12),
    ]

    oe_map = {o["id"]: o for o in oefeningen}

    for t in trainingen:
        story.append(Paragraph(f"{t.get('datum','')} · {t.get('thema','')}", S["h2"]))
        info = f"{t.get('starttijd','')}–{t.get('eindtijd','')}  |  📍 {t.get('locatie','—')}"
        story.append(Paragraph(info, S["body"]))
        if t.get("notities"):
            story.append(Paragraph(t["notities"], S["body"]))
        oef_ids = t.get("oefening_ids") or []
        if oef_ids and oe_map:
            story.append(Paragraph("Oefeningen:", S["label"]))
            for oid in oef_ids:
                o = oe_map.get(oid)
                if o:
                    story.append(Paragraph(f"• {o['naam']} — {o.get('categorie','')} · {o.get('duur_minuten','')} min", S["body"]))
        story.append(Spacer(1, 6))

    doc.build(story)
    return buf.getvalue()


def generate_blessures_pdf(blessures: list) -> bytes | None:
    if not REPORTLAB_AVAILABLE:
        return None
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors as _c
    import datetime as _dt

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm,
                             topMargin=20*mm, bottomMargin=20*mm)
    S = _pdf_styles()
    team_name = st.session_state.get("active_team_name", "Team")
    story = [
        Paragraph("Coach Studio", S["title"]),
        Paragraph(f"Blessurerapport — {team_name} · {_dt.date.today().strftime('%d %b %Y')}", S["sub"]),
        HRFlowable(width="100%", thickness=1, color=_c.HexColor("#3b82f6"), spaceAfter=12),
    ]

    actief = [b for b in blessures if not b.get("datum_herstel")]
    herstel = [b for b in blessures if b.get("datum_herstel")]

    story.append(Paragraph(f"Actieve blessures ({len(actief)})", S["h2"]))
    if not actief:
        story.append(Paragraph("Geen actieve blessures.", S["body"]))
    for b in actief:
        story.append(Paragraph(f"<b>{b.get('player_name','?')}</b> — {b.get('blessure_type','')}", S["body"]))
        story.append(Paragraph(f"Ernst: {b.get('ernst','')} | Sinds: {b.get('datum_start','')} | Terug: {b.get('verwachte_terugkeer','?')}", S["body"]))
        if b.get("beschrijving"):
            story.append(Paragraph(b["beschrijving"], S["body"]))
        story.append(Spacer(1, 6))

    if herstel:
        story.append(Paragraph(f"Hersteld ({len(herstel)})", S["h2"]))
        for b in herstel:
            story.append(Paragraph(f"<b>{b.get('player_name','?')}</b> — {b.get('blessure_type','')} · Hersteld: {b.get('datum_herstel','')}", S["body"]))

    doc.build(story)
    return buf.getvalue()


def generate_scouting_pdf(rapport: dict) -> bytes | None:
    if not REPORTLAB_AVAILABLE:
        return None
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors as _c
    import datetime as _dt

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm,
                             topMargin=20*mm, bottomMargin=20*mm)
    S = _pdf_styles()
    team_name = st.session_state.get("active_team_name", "Team")
    teg = rapport.get("tegenstander", "?")
    story = [
        Paragraph("Coach Studio", S["title"]),
        Paragraph(f"Scoutingrapport: {teg} — {team_name} · {_dt.date.today().strftime('%d %b %Y')}", S["sub"]),
        HRFlowable(width="100%", thickness=1, color=_c.HexColor("#3b82f6"), spaceAfter=12),
    ]

    velden = [
        ("Formatie", "formatie"), ("Speelstijl", "speelstijl"),
        ("💪 Sterke punten", "sterke_punten"), ("⚠️ Zwakke punten", "zwakke_punten"),
        ("🔱 Corners aanval", "corners_aanval"), ("🛡️ Corners verdediging", "corners_verdediging"),
        ("🎯 Aanbevolen aanpak", "aanbevolen_aanpak"),
    ]
    for label, key in velden:
        val = rapport.get(key, "")
        if val:
            story.append(Paragraph(label, S["label"]))
            story.append(Paragraph(val, S["body"]))
            story.append(Spacer(1, 6))

    doc.build(story)
    return buf.getvalue()


# ==================================================
# SEIZOENSDOELEN
# ==================================================
DOEL_CATEGORIEEN = ["Resultaten", "Aanval", "Verdediging", "Teamontwikkeling", "Individueel", "Anders"]


@st.cache_data(ttl=30, show_spinner=False)
def _fetch_goals(team_id: str) -> list:
    client = get_supabase_client()
    if not team_id or client is None:
        return []
    try:
        r = client.table("season_goals").select("*").eq("team_id", team_id)\
            .order("created_at").execute()
        return r.data or []
    except Exception:
        return []


def cloud_list_goals() -> list:
    tid = _active_team_id()
    result = _fetch_goals(tid or "")
    if result is not None:
        mark_cloud_ok()
    return result or []


def cloud_save_goal(data: dict) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return
    try:
        data["team_id"] = tid
        if not data.get("id"):
            data["id"] = str(uuid.uuid4())
        client.table("season_goals").upsert(data, on_conflict="id").execute()
        _fetch_goals.clear()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("doel opslaan", err)


def cloud_update_goal(goal_id: str, updates: dict) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return
    try:
        client.table("season_goals").update(updates).eq("id", goal_id).eq("team_id", tid).execute()
        _fetch_goals.clear()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("doel bijwerken", err)


def cloud_delete_goal(goal_id: str) -> None:
    tid = _active_team_id()
    client = get_supabase_client()
    if not tid or client is None:
        return
    try:
        client.table("season_goals").delete().eq("id", goal_id).eq("team_id", tid).execute()
        _fetch_goals.clear()
        mark_cloud_ok()
    except Exception as err:
        log_cloud_error("doel verwijderen", err)


def render_goals_screen() -> None:
    st.markdown("### 🎯 Seizoensdoelen")

    doelen = cloud_list_goals()
    voltooid = [d for d in doelen if d.get("voltooid")]
    actief = [d for d in doelen if not d.get("voltooid")]

    # ── KPI strip ──
    c1, c2, c3 = st.columns(3)
    c1.metric("🎯 Totaal doelen", len(doelen))
    c2.metric("✅ Behaald", len(voltooid))
    c3.metric("🔄 In uitvoering", len(actief))

    st.divider()
    tab_actief, tab_nieuw, tab_voltooid = st.tabs(["🔄 Actief", "➕ Nieuw doel", "✅ Behaald"])

    # ── TAB: Actieve doelen ──
    with tab_actief:
        if not actief:
            st.info("Nog geen seizoensdoelen. Voeg er een toe via '➕ Nieuw doel'.")
        for d in actief:
            doel_val = float(d.get("doel_waarde") or 1)
            huidig_val = float(d.get("huidige_waarde") or 0)
            pct = min(int((huidig_val / doel_val) * 100), 100) if doel_val > 0 else 0
            eenheid = d.get("eenheid", "")
            cat = d.get("categorie", "")

            cat_colors = {
                "Resultaten": "#3b82f6", "Aanval": "#10b981", "Verdediging": "#f59e0b",
                "Teamontwikkeling": "#8b5cf6", "Individueel": "#f43f5e", "Anders": "#64748b",
            }
            color = cat_colors.get(cat, "#3b82f6")

            st.markdown(
                f'<div style="background:#0f1624;border:1px solid #1a2540;border-left:3px solid {color};'
                f'border-radius:12px;padding:16px 18px;margin-bottom:10px;">'
                f'<div style="display:flex;justify-content:space-between;align-items:flex-start;">'
                f'<div><span style="color:#f1f5f9;font-weight:700;font-size:15px;">{d.get("titel","")}</span>'
                f'<span style="background:{color}22;color:{color};padding:2px 8px;border-radius:6px;'
                f'font-size:10px;font-weight:700;margin-left:10px;">{cat}</span></div>'
                f'<div style="color:#94a3b8;font-size:13px;font-weight:700;">'
                f'{huidig_val:.0f} / {doel_val:.0f} {eenheid}</div>'
                f'</div>'
                f'<div style="margin-top:10px;background:#1a2540;border-radius:6px;height:8px;">'
                f'<div style="background:{color};width:{pct}%;height:8px;border-radius:6px;'
                f'transition:width 0.4s ease;"></div></div>'
                f'<div style="color:#64748b;font-size:11px;margin-top:4px;">{pct}% behaald</div>'
                f'{"<div style=color:#94a3b8;font-size:12px;margin-top:6px;>" + d.get("beschrijving","") + "</div>" if d.get("beschrijving") else ""}'
                f'</div>',
                unsafe_allow_html=True,
            )

            col_prog, col_vol, col_del = st.columns([2, 1, 1])
            with col_prog:
                nieuwe_waarde = st.number_input(
                    "Huidige stand bijwerken",
                    min_value=0.0, max_value=float(doel_val) * 2,
                    value=float(huidig_val), step=1.0,
                    key=f"goal_prog_{d['id']}",
                    label_visibility="collapsed",
                )
            with col_vol:
                if st.button("📈 Bijwerken", key=f"goal_upd_{d['id']}", use_container_width=True):
                    cloud_update_goal(d["id"], {"huidige_waarde": nieuwe_waarde})
                    st.rerun()
            with col_del:
                if st.button("✅ Behaald!", key=f"goal_done_{d['id']}", use_container_width=True, type="primary"):
                    cloud_update_goal(d["id"], {"voltooid": True, "huidige_waarde": doel_val})
                    st.success(f"🎉 '{d.get('titel','')}' behaald!")
                    st.rerun()

    # ── TAB: Nieuw doel ──
    with tab_nieuw:
        with st.form("goal_form"):
            titel = st.text_input("Titel", placeholder="Bijv. 70% van de wedstrijden winnen", key="gf_titel")
            beschrijving = st.text_area("Beschrijving (optioneel)", height=70, key="gf_desc",
                                         placeholder="Waarom is dit doel belangrijk?")
            c1, c2 = st.columns(2)
            categorie = c1.selectbox("Categorie", DOEL_CATEGORIEEN, key="gf_cat")
            eenheid = c2.text_input("Eenheid", placeholder="bijv. %, goals, punten", key="gf_eenheid")
            c3, c4 = st.columns(2)
            doel_waarde = c3.number_input("Doelwaarde", min_value=1.0, value=10.0, step=1.0, key="gf_doel")
            start_waarde = c4.number_input("Beginstand", min_value=0.0, value=0.0, step=1.0, key="gf_start")
            toevoegen = st.form_submit_button("🎯 Doel toevoegen", type="primary")

        if toevoegen:
            if not titel.strip():
                st.error("Vul een titel in.")
            else:
                cloud_save_goal({
                    "titel": titel.strip(), "beschrijving": beschrijving.strip(),
                    "categorie": categorie, "eenheid": eenheid.strip(),
                    "doel_waarde": float(doel_waarde), "huidige_waarde": float(start_waarde),
                    "voltooid": False,
                })
                st.success("Doel toegevoegd!")
                st.rerun()

    # ── TAB: Behaalde doelen ──
    with tab_voltooid:
        if not voltooid:
            st.info("Nog geen doelen behaald — blijf gaan! 💪")
        for d in voltooid:
            st.markdown(
                f'<div style="background:#0f1624;border:1px solid #10b98133;border-left:3px solid #10b981;'
                f'border-radius:12px;padding:14px 18px;margin-bottom:8px;opacity:0.85;">'
                f'<span style="color:#34d399;font-size:18px;">✅</span> '
                f'<span style="color:#f1f5f9;font-weight:700;">{d.get("titel","")}</span>'
                f'<span style="color:#64748b;font-size:12px;margin-left:10px;">{d.get("categorie","")}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
            if st.button("↩ Heropen", key=f"goal_reopen_{d['id']}", help="Zet terug naar actief"):
                cloud_update_goal(d["id"], {"voltooid": False})
                st.rerun()


# ==================================================
# TOOL SELECTOR — keuze uit 3 tools na team-login
# ==================================================
def render_match_selector_on_home() -> None:
    """Wedstrijd-picker op het tool-overzicht (landingspagina).

    Toont:
    - Welke wedstrijd momenteel geladen is.
    - Mogelijkheid om een eerdere wedstrijd te openen uit de cloud.
    - Mogelijkheid om een nieuwe wedstrijd te starten.
    """
    # Huidig geladen wedstrijd weergeven (zonder team-prefix)
    current_mid = st.session_state.get("match_id") or "—"
    pretty_mid = unscope_match_id(current_mid) if current_mid != "—" else "—"
    event_count = len(st.session_state.get("events") or [])

    st.markdown(
        f"""
        <div style='
            background: {CARD_BG_ELEVATED};
            border: 1px solid {CARD_BORDER_SOFT};
            border-radius: 14px;
            padding: 18px 20px;
            margin: 8px 0 18px 0;
        '>
            <div style='color: {TEXT_SUB}; font-size: 13px; text-transform: uppercase; letter-spacing: .08em;'>
                Actieve wedstrijd
            </div>
            <div style='color: {TEXT_MAIN}; font-size: 20px; font-weight: 700; margin-top: 4px;'>
                📋 {pretty_mid}
            </div>
            <div style='color: {TEXT_MUTED}; font-size: 13px; margin-top: 4px;'>
                {event_count} events geladen
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    c1, c2 = st.columns(2)
    with c1:
        if st.button("📂  Open een eerdere wedstrijd", use_container_width=True, key="home_open_match_btn"):
            st.session_state.home_show_match_picker = not st.session_state.get("home_show_match_picker", False)
            st.session_state.home_show_new_match = False
            st.rerun()
    with c2:
        if st.button("➕  Nieuwe wedstrijd starten", use_container_width=True, key="home_new_match_btn", type="primary"):
            st.session_state.home_show_new_match = not st.session_state.get("home_show_new_match", False)
            st.session_state.home_show_match_picker = False
            st.rerun()

    # Paneel: eerdere wedstrijd kiezen
    if st.session_state.get("home_show_match_picker"):
        with st.container(border=True):
            st.markdown("#### Kies een eerder opgeslagen wedstrijd")
            match_ids = list_match_ids_from_cloud(limit=50)
            if not match_ids:
                st.info("Geen eerdere wedstrijden gevonden voor dit team.")
            else:
                # Toon leesbare namen (zonder prefix), maar open o.b.v. echte match_id
                label_to_id = {unscope_match_id(m): m for m in match_ids}
                picked_label = st.selectbox(
                    "Wedstrijd",
                    list(label_to_id.keys()),
                    key="home_picker_match_label",
                )
                pc1, pc2 = st.columns([1, 1])
                if pc1.button("✅  Open deze wedstrijd", use_container_width=True, type="primary", key="home_picker_open_btn"):
                    real_mid = label_to_id[picked_label]
                    switch_to_match(real_mid)
                    st.session_state.home_show_match_picker = False
                    st.success(f"Geladen: {picked_label} ({len(st.session_state.events)} events).")
                    st.rerun()
                if pc2.button("Annuleer", use_container_width=True, key="home_picker_cancel_btn"):
                    st.session_state.home_show_match_picker = False
                    st.rerun()

    # Paneel: nieuwe wedstrijd starten
    if st.session_state.get("home_show_new_match"):
        with st.container(border=True):
            st.markdown("#### Nieuwe wedstrijd starten")
            nm1, nm2 = st.columns(2)
            with nm1:
                nm_team = st.text_input(
                    "Eigen team",
                    value=st.session_state.team_name or "Ons team",
                    key="home_nm_team",
                )
                nm_opp = st.text_input("Tegenstander", value="", placeholder="Bijv. Kampong D1", key="home_nm_opp")
            with nm2:
                nm_date = st.text_input("Datum", value=time.strftime("%Y-%m-%d"), key="home_nm_date")
                nm_label = st.text_input(
                    "Korte omschrijving (optioneel)",
                    value="",
                    placeholder="Bijv. thuis / competitie",
                    key="home_nm_label",
                )
            sc1, sc2 = st.columns([1, 1])
            if sc1.button("🚀  Start wedstrijd", use_container_width=True, type="primary", key="home_nm_start_btn"):
                clean = lambda s: re.sub(r"\W+", "-", s.strip()).strip("-") or "match"
                parts = [clean(nm_team), clean(nm_opp) if nm_opp else "vs", clean(nm_date)]
                if nm_label:
                    parts.append(clean(nm_label))
                new_id = "-".join(parts)
                # Scope aan huidig team zodat de wedstrijd onder dit team wordt opgeslagen
                new_id = scope_match_id(new_id)

                # State resetten voor een verse wedstrijd
                st.session_state.match_id = new_id
                st.session_state.team_name = nm_team or "Ons team"
                if nm_opp:
                    st.session_state.opponent_name = nm_opp
                st.session_state.events = []
                st.session_state.video_clips = []
                st.session_state.pushoff_offsets = {q: None for q in QUARTERS}
                st.session_state.quarter = "Q1"
                st.session_state.score_team = 0
                st.session_state.score_opponent = 0
                st.session_state.auto_notes = ""
                st.session_state.halftime_report = ""
                # Widget-keys opruimen zodat de setup-bar de juiste namen oppakt
                for wk in ("w_team_name", "w_opponent_name", "w_match_id"):
                    if wk in st.session_state:
                        del st.session_state[wk]
                st.session_state.home_show_new_match = False
                st.success(f"Nieuwe wedstrijd gestart: {unscope_match_id(new_id)}")
                st.rerun()
            if sc2.button("Annuleer", use_container_width=True, key="home_nm_cancel_btn"):
                st.session_state.home_show_new_match = False
                st.rerun()


def render_tool_selector() -> None:
    """Landingspagina met tool-kaarten in een 2×3 grid."""
    team_name = st.session_state.get("active_team_name") or "je team"

    # Welkom-header
    st.markdown(
        f'<div class="cs-welcome">'
        f'<div class="cs-welcome-title">Welkom, {team_name} 👋</div>'
        f'<div class="cs-welcome-sub">Kies een tool om te beginnen</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # --- Wedstrijd-picker (vóór tools, zodat je eerst de wedstrijd kiest) ---
    render_match_selector_on_home()

    tools = [
        {
            "id": "MATCH_ANALYSIS",
            "title": "Wedstrijd analyse",
            "icon": "⚡",
            "desc": "Live taggen tijdens de wedstrijd. Daarna analyseren per kwart, veldweergave en coachrapport.",
            "tabs": "Live · Analyse · Veld · Rapport",
        },
        {
            "id": "VIDEO_ANALYSIS",
            "title": "Video analyse",
            "icon": "🎬",
            "desc": "Knip clips uit video, label ze tactisch en bouw een coachingbundel voor je spelers.",
            "tabs": "Clips · Tags · Highlight reel",
        },
        {
            "id": "SUBSTITUTION",
            "title": "Wisselschema",
            "icon": "🔄",
            "desc": "Automatisch eerlijk wisselschema per minuut en linie. Formatiebeheer en PDF-export.",
            "tabs": "Team · Wedstrijd · Schema · Export",
        },
        {
            "id": "SEASON",
            "title": "Seizoensoverzicht",
            "icon": "📈",
            "desc": "W/G/V statistieken, doelsaldo, topscorers, speelminuten-verdeling en seizoensrapport.",
            "tabs": "KPI · Topscorers · Trend · Minuten",
        },
        {
            "id": "PLAYER_PROFILE",
            "title": "Spelersprofiel",
            "icon": "👤",
            "desc": "Notities per speler: technisch, tactisch, fysiek en mentaal. Beoordeling 1–5 met groeicurve.",
            "tabs": "Profiel · Notities · Trend",
        },
        {
            "id": "MATCH_MGMT",
            "title": "Wedstrijden & uitslagen",
            "icon": "🏆",
            "desc": "Uitslagen bevestigen, tegenstander en locatie invullen. Seizoensoverzicht W/G/V en doelsaldo.",
            "tabs": "Uitslag · Seizoensoverzicht",
        },
        {
            "id": "SELECTION",
            "title": "Selectietool",
            "icon": "👥",
            "desc": "Stel je basisopstelling visueel samen. Kies formatie, wijs spelers toe aan posities en sla reserves op.",
            "tabs": "Formatie · Veld · Reserves",
            "new": True,
        },
        {
            "id": "TRAINING",
            "title": "Trainingsplanning",
            "icon": "📅",
            "desc": "Plan trainingen met datum, tijd en thema. Bouw een oefeningen-bibliotheek en koppel ze aan sessies.",
            "tabs": "Sessies · Oefeningen",
            "new": True,
        },
        {
            "id": "INJURIES",
            "title": "Blessure tracker",
            "icon": "🩹",
            "desc": "Registreer blessures per speler met ernst en verwachte terugkeer. Altijd inzicht in wie fit is.",
            "tabs": "Actief · Toevoegen · Historie",
            "new": True,
        },
        {
            "id": "SCOUTING",
            "title": "Tegenstander scouting",
            "icon": "🔍",
            "desc": "Analyseer tegenstanders op formatie, speelstijl, corners en zwakke punten. Bouw een scoutingdossier.",
            "tabs": "Overzicht · Rapport",
            "new": True,
        },
        {
            "id": "GOALS",
            "title": "Seizoensdoelen",
            "icon": "🎯",
            "desc": "Stel doelen voor het seizoen in, houd de voortgang bij met visuele progressiebars en vier behaalde doelen.",
            "tabs": "Actief · Nieuw · Behaald",
            "new": True,
        },
    ]

    # Grid — 3 kolommen
    st.markdown('<div class="cs-section-label">Tools</div>', unsafe_allow_html=True)
    rows = [tools[i:i+3] for i in range(0, len(tools), 3)]
    for row_tools in rows:
        cols = st.columns(3, gap="medium")
        for col, tool in zip(cols, row_tools):
            with col:
                new_badge = '<div class="tool-card-new">Nieuw</div>' if tool.get("new") else ''
                st.markdown(
                    f'<div class="tool-card">'
                    f'{new_badge}'
                    f'<div class="tool-card-icon-wrap">{tool["icon"]}</div>'
                    f'<div class="tool-card-title">{tool["title"]}</div>'
                    f'<div class="tool-card-desc">{tool["desc"]}</div>'
                    f'<div class="tool-card-tabs">{tool["tabs"]}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                if st.button(
                    f"Open →",
                    key=f"open_tool_{tool['id']}",
                    use_container_width=True,
                    type="primary",
                ):
                    st.session_state.active_tool = tool["id"]
                    st.session_state.active_screen = None
                    st.rerun()
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    st.markdown("<div style='height: 8px;'></div>", unsafe_allow_html=True)

    # Team-beheer onderaan
    with st.expander("⚙️  Team-beheer"):
        teams = list_teams()
        st.markdown(f"Er zijn **{len(teams)}** team(s) geregistreerd.")
        for t in teams:
            is_active = t.get("id") == st.session_state.get("active_team_id")
            c1, c2 = st.columns([4, 1])
            with c1:
                label = f"• {t['name']}" + ("  ✓ (dit team)" if is_active else "")
                st.markdown(label)
            with c2:
                if not is_active and st.button("Verwijder", key=f"del_team_{t['id']}"):
                    delete_team(t["id"])
                    st.rerun()
        st.divider()
        st.markdown("**Nieuw team toevoegen**")
        nn = st.text_input("Naam", key="tm_new_name", placeholder="Bijv. MO16-1 Hockeyclub")
        np1 = st.text_input("Wachtwoord", type="password", key="tm_new_pw")
        if st.button("Team aanmaken", key="tm_create_btn", type="primary"):
            ok, msg = create_team(nn, np1)
            if ok:
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)


# ==================================================
# MAIN
# ==================================================
inject_custom_css()
require_password()
render_hero_header()
render_logout_button()

# Als er nog geen tool gekozen is → tool-selector tonen (en verder niks)
if not st.session_state.get("active_tool"):
    render_tool_selector()
    st.stop()

# Alleen Wedstrijd analyse heeft de setup-bar (team/tegenstander/kwart/match-id)
if st.session_state.active_tool == "MATCH_ANALYSIS":
    render_setup_bar()
render_navigation()
# Auto-sync alleen actief in wedstrijd analyse — niet op andere pagina's pollen
if st.session_state.active_tool == "MATCH_ANALYSIS":
    auto_sync_cloud()

# Vangnet: als team-namen leeg zijn geraakt, terugzetten op standaard
# zodat KPI's en score werken met de namen waarmee de events zijn opgeslagen.
if not (st.session_state.team_name or "").strip():
    st.session_state.team_name = "Ons team"
if not (st.session_state.opponent_name or "").strip():
    st.session_state.opponent_name = "Tegenstander"

# build_df en refresh_derived_state zijn zwaar — alleen uitvoeren bij tools die events nodig hebben
_needs_events = st.session_state.active_tool in ("MATCH_ANALYSIS", "VIDEO_ANALYSIS")
if _needs_events:
    df = build_df()
    if not df.empty:
        refresh_derived_state()
else:
    df = pd.DataFrame()

# Cloud-status: alleen tonen bij echte problemen, niet als groene balk op elke pagina
if not cloud_enabled():
    st.caption("⚠️ Cloud sync inactief — voeg SUPABASE_URL en SUPABASE_KEY toe aan Streamlit secrets.")

# Waarschuwing als er recent een cloud-fout is geweest (zodat je niet stil data verliest)
if st.session_state.cloud_errors and not st.session_state.last_cloud_ok:
    st.error(
        "⚠️ Cloud-probleem opgetreden — je tags zijn wél lokaal bewaard. "
        "Druk op Sync om het opnieuw te proberen.",
        icon="⚠️",
    )
    with st.expander("Details cloud-fouten"):
        for e in reversed(st.session_state.cloud_errors[-5:]):
            st.code(e, language="text")


# ------ Tool-routing ------
tool = st.session_state.active_tool
screen = st.session_state.active_screen

if tool == "MATCH_ANALYSIS":
    if screen == "LIVE":
        render_live_screen(df)
    elif screen == "ANALYSE":
        render_analysis_screen(df)
    elif screen == "VELD":
        render_field_screen(df)
    elif screen == "RAPPORT":
        render_report_screen(df)
    else:
        render_live_screen(df)
elif tool == "VIDEO_ANALYSIS":
    render_video_analysis_screen(df)
elif tool == "SUBSTITUTION":
    render_substitution_screen()
elif tool == "SEASON":
    render_season_screen()
elif tool == "PLAYER_PROFILE":
    render_player_profile_screen()
elif tool == "MATCH_MGMT":
    render_match_management_screen()
elif tool == "SELECTION":
    render_selection_screen()
elif tool == "TRAINING":
    render_training_screen()
elif tool == "INJURIES":
    render_injury_screen()
elif tool == "SCOUTING":
    render_scouting_screen()
elif tool == "GOALS":
    render_goals_screen()
else:
    render_tool_selector()
